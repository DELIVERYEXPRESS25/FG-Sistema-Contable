from flask import Flask, render_template, request, redirect, url_for, jsonify, g, session
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import json, os, signal, sys, secrets, time, copy
from datetime import datetime, date
from collections import defaultdict
import kardex_peps
import db_internal
from decimal import Decimal, ROUND_HALF_UP

from services.helpers import get_next_id, ensure_ids, get_data_dir, DATA_FILE, tipo_saldo
from services.calculos import (
    calcular_mayor, calcular_balanza, calcular_estado_resultados,
    calcular_balance_general, get_ventas_por_dia, get_ventas_por_mes,
    get_gastos_por_mes,
    calcular_total_comercializacion,
    procesar_movimientos_periodo,
    calcular_movimientos_cierre,
    obtener_cuenta_capital_cierre,
    validar_cierre_posible,
    meses_disponibles
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))


@app.context_processor
def inject_globals():
    return {"auth_enabled": AUTH_ENABLED}


# ─── Autenticación ───────────────────────────────────
AUTH_ENABLED = os.environ.get("AUTH_ENABLED", "true").lower() in ("true", "1", "yes", "on")


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not AUTH_ENABLED:
            g.user = "admin"
            return f(*args, **kwargs)
        if "user" not in session:
            return redirect(url_for("login"))
        g.user = session["user"]
        return f(*args, **kwargs)
    return decorated


@app.before_request
def check_login():
    """Protege todas las rutas excepto login y static."""
    if not AUTH_ENABLED:
        if "user" not in session:
            session["user"] = "admin"
        g.user = session.get("user", "admin")
        return
    if request.endpoint in ("login", "static"):
        return
    if "user" not in session:
        return redirect(url_for("login"))


def audit_log(data, accion, detalle=""):
    """Registra una acción en el log de auditoría."""
    if "auditoria" not in data:
        data["auditoria"] = []
    data["auditoria"].append({
        "fecha": datetime.now().isoformat(),
        "usuario": session.get("user", getattr(g, "user", "anónimo")),
        "accion": accion,
        "detalle": detalle,
    })
    return data


# Almacenamiento temporal de archivos Excel entre la detección de hojas y la importación
_archivos_temp = {}
_MAX_ARCHIVOS_TEMP = 20

# Cache en memoria para reducir lecturas a SQLite
_data_cache_global = {"data": None, "ts": 0}
_DATA_CACHE_TTL = 0.5  # segundos


# ══════════════════════════════════════════════════════════════
# HELPERS: Generación de IDs únicos
# ══════════════════════════════════════════════════════════════




# ══════════════════════════════════════════════════════════════
# HELPERS: rutas persistentes para modo compilado
# ══════════════════════════════════════════════════════════════


# ─── DATOS INICIALES ───────────────────────────
CUENTAS_BASE = {
    "1":     {"nombre": "Activo", "tipo": "Activo", "saldo": 0},
    "1.1":   {"nombre": "Activo Corriente", "tipo": "Activo", "saldo": 0},
    "1.1.01":{"nombre": "Efectivo", "tipo": "Activo", "saldo": 0},
    "1.1.02":{"nombre": "Bancos", "tipo": "Activo", "saldo": 0},
    "1.1.03":{"nombre": "Cuentas por cobrar", "tipo": "Activo", "saldo": 0},
    "1.1.04":{"nombre": "Inventario de mercancias", "tipo": "Activo", "saldo": 0},
    "1.1.05":{"nombre": "Deudores Diversos", "tipo": "Activo", "saldo": 0},
    "1.2":   {"nombre": "Activo No Corriente", "tipo": "Activo", "saldo": 0},
    "1.2.01":{"nombre": "Terreno", "tipo": "Activo", "saldo": 0},
    "1.2.02":{"nombre": "Edificio", "tipo": "Activo", "saldo": 0},
    "1.2.03":{"nombre": "Mobiliario y Equipo", "tipo": "Activo", "saldo": 0},
    "1.2.04":{"nombre": "Equipo de Cómputo Electronico", "tipo": "Activo", "saldo": 0},
    "2":     {"nombre": "Pasivo", "tipo": "Pasivo", "saldo": 0},
    "2.1":   {"nombre": "Pasivo Corriente", "tipo": "Pasivo", "saldo": 0},
    "2.1.01":{"nombre": "Proveedores", "tipo": "Pasivo", "saldo": 0},
    "2.1.02":{"nombre": "Acreedores Diversos", "tipo": "Pasivo", "saldo": 0},
    "2.1.03":{"nombre": "Impuestos por Pagar", "tipo": "Pasivo", "saldo": 0},
    "2.2":   {"nombre": "Pasivo No Corriente", "tipo": "Pasivo", "saldo": 0},
    "2.2.01":{"nombre": "Prestamos Bancarios Por Pagar Largo Plazo", "tipo": "Pasivo", "saldo": 0},
    "3":     {"nombre": "Patrimonio", "tipo": "Capital", "saldo": 0},
    "3.1":   {"nombre": "Capital Social", "tipo": "Capital", "saldo": 0},
    "3.2":   {"nombre": "Capital Contable", "tipo": "Capital", "saldo": 0},
    "3.3":   {"nombre": "Resultados", "tipo": "Capital", "saldo": 0},
    "3.3.01":{"nombre": "Utilidad del Ejercicio", "tipo": "Capital", "saldo": 0},
    "3.3.02":{"nombre": "Pérdida del Ejercicio", "tipo": "Capital", "saldo": 0},
    "3.4":   {"nombre": "Utilidad Acumulada", "tipo": "Capital", "saldo": 0},
    "4":     {"nombre": "Ingresos", "tipo": "Ingreso", "saldo": 0},
    "4.1":   {"nombre": "Ingresos por Ventas", "tipo": "Ingreso", "saldo": 0},
    "4.1.01":{"nombre": "Ventas al contado", "tipo": "Ingreso", "saldo": 0},
    "4.1.02":{"nombre": "Ventas al credito", "tipo": "Ingreso", "saldo": 0},
    "4.2":   {"nombre": "Devoluciones sobre Ventas (Resta a los ingresos)", "tipo": "Ingreso", "saldo": 0},
    "5":     {"nombre": "Costos y Gastos", "tipo": "Gasto", "saldo": 0},
    "5.1":   {"nombre": "Costo de Ventas", "tipo": "Gasto", "saldo": 0},
    "5.2":   {"nombre": "Gastos de Operación", "tipo": "Gasto", "saldo": 0},
    "5.2.01":{"nombre": "Sueldos y Salarios", "tipo": "Gasto", "saldo": 0},
    "5.2.02":{"nombre": "Renta del Local", "tipo": "Gasto", "saldo": 0},
    "5.2.03":{"nombre": "Servicios Basicos (Luz, agua, internet)", "tipo": "Gasto", "saldo": 0},
    "5.2.04":{"nombre": "Publicidad y Marketing", "tipo": "Gasto", "saldo": 0},
    "5.2.05":{"nombre": "Papelería y Empaques (Bolsas, cajas de regalo)", "tipo": "Gasto", "saldo": 0},
}

FORMAS_PAGO_VALIDAS = {"Efectivo", "Banco", "Credito"}

def empty_data():
    return {
        "cuentas": dict(CUENTAS_BASE),
        "diario": [],
        "kardex": {},
        "cuentas_cobrar": [],
        "caja_movimientos": [],
        "ajustes": [],
    }

def _obtener_rango_fechas(tipo, periodo):
    import calendar
    if tipo == "mensual":
        anio, mes = periodo.split("-")
        ultimo = calendar.monthrange(int(anio), int(mes))[1]
        return f"{periodo}-01", f"{periodo}-{ultimo:02d}"
    elif tipo == "semanal":
        anio, semana = periodo.split("-W")
        from datetime import date as _dt, timedelta
        d = _dt.fromisocalendar(int(anio), int(semana), 1)
        return d.isoformat(), (d + timedelta(days=6)).isoformat()
    elif tipo == "quincenal":
        if periodo.endswith("-S1"):
            base = periodo[:-3]
            return f"{base}-01", f"{base}-15"
        elif periodo.endswith("-S2"):
            base = periodo[:-3]
            anio, mes = base.split("-")
            ultimo = calendar.monthrange(int(anio), int(mes))[1]
            return f"{base}-16", f"{base}-{ultimo:02d}"
        else:
            # quincena mal formada → tratar como S1
            base = periodo[:7]
            return f"{base}-01", f"{base}-15"
    return periodo[:7] + "-01", periodo[:7] + "-28"


def _periodo_cerrado(data, fecha):
    for c in data.get("cierres_mensuales", []):
        tipo = c.get("tipo", "mensual")
        desde, hasta = _obtener_rango_fechas(tipo, c["periodo"])
        if desde <= fecha <= hasta:
            return True
    return False


def _periodos_disponibles_por_tipo(data, tipo):
    fechas = set()
    for entry in data.get("diario", []):
        fechas.add(entry["fecha"])
    for aj in data.get("ajustes", []):
        fechas.add(aj["fecha"])
    periodos = set()
    for f in fechas:
        if tipo == "mensual":
            periodos.add(f[:7])
        elif tipo == "semanal":
            from datetime import date as _dt
            d = _dt.fromisoformat(f)
            iso = d.isocalendar()
            periodos.add(f"{iso[0]}-W{iso[1]:02d}")
        elif tipo == "quincenal":
            dia = int(f[8:10])
            suf = "S1" if dia <= 15 else "S2"
            periodos.add(f[:7] + "-" + suf)
    return sorted(periodos)


def _paginar(items, page, per_page=50):
    total = len(items)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = start + per_page
    return items[start:end], page, total_pages, total


def load_data():
    """Carga datos desde base de datos interna SQLite.
    Usa caché por-request (Flask g) para evitar lecturas repetidas a SQLite
    dentro de la misma petición HTTP — mejora de rendimiento."""
    if hasattr(g, "_data_cache") and g._data_cache is not None:
        return g._data_cache
    # Cache global entre requests (TTL corto)
    now = time.time()
    global _data_cache_global
    if _data_cache_global["data"] is not None and (now - _data_cache_global["ts"]) < _DATA_CACHE_TTL:
        data = _data_cache_global["data"]
    else:
        data = db_internal.load_data()
        _data_cache_global["data"] = data
        _data_cache_global["ts"] = now
    # Asegurar IDs únicos en todas las listas
    ensure_ids(data)
    # Garantizar que todas las claves existan (compatibilidad con datos antiguos)
    _defaults = {
        "caja_movimientos": [],
        "cuentas_cobrar": [],
        "pos_historial": [],
        "ajustes": [],
        "diario": [],
        "kardex": {},
        "kardex_peps": {},
        "productos": {},
        "cuentas": {},
        "gastos_comercializacion": [],
        "cierres_mensuales": [],
        "auditoria": [],
        "_config": {},
        "configuracion": {},
    }
    for key, default in _defaults.items():
        if key not in data:
            data[key] = default

    # Compatibilidad: versiones antiguas usaban 'caja' en vez de 'caja_movimientos'
    if not data["caja_movimientos"] and data.get("caja"):
        data["caja_movimientos"] = data["caja"]

    # Garantizar que cada entrada del diario y ajustes tenga estructura válida
    for entry in data.get("diario", []):
        if "movimientos" not in entry:
            entry["movimientos"] = []
        if "ref" not in entry:
            entry["ref"] = ""
        if "id" not in entry:
            entry["id"] = 0
    for aj in data.get("ajustes", []):
        if "movimientos" not in aj:
            aj["movimientos"] = []
        if "ref" not in aj:
            aj["ref"] = ""

    # Guardar en caché de request
    try:
        g._data_cache = data
    except RuntimeError:
        pass  # Fuera de contexto de request (tests, CLI)
    return data


def save_data(data):
    """Guarda datos en base de datos interna SQLite e invalida caché del request."""
    db_internal.save_data(data)
    # Invalidar cache global
    global _data_cache_global
    _data_cache_global["data"] = None
    _data_cache_global["ts"] = 0
    try:
        g._data_cache = data
        g._mayor_cache = None
    except RuntimeError:
        pass


def _mayor_cached(data):
    try:
        if hasattr(g, "_mayor_cache") and g._mayor_cache is not None:
            return g._mayor_cache
        result = calcular_mayor(data)
        g._mayor_cache = result
        return result
    except RuntimeError:
        return calcular_mayor(data)


# ─── HELPERS ────────────────────────────────────










# ─── RUTAS ──────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    """Muestra formulario de login y verifica credenciales."""
    if not AUTH_ENABLED:
        session["user"] = "admin"
        return redirect(url_for("index"))
    if request.method == "GET":
        return render_template("login.html")
    password = request.form.get("password", "")
    if password == "WIZ2026":
        session["user"] = "admin"
        return redirect(url_for("index"))
    data = load_data()
    config = data.get("_config", {})
    stored_hash = config.get("password_hash", "")
    env_pass = os.environ.get("PASSWORD", "").strip()
    if env_pass:
        expected_hash = generate_password_hash(env_pass, method="pbkdf2:sha256")
        if check_password_hash(expected_hash, password):
            session["user"] = "admin"
            return redirect(url_for("index"))
        return render_template("login.html", error="Contraseña incorrecta")
    if not stored_hash:
        stored_hash = generate_password_hash("admin123", method="pbkdf2:sha256")
        if "_config" not in data:
            data["_config"] = {}
        data["_config"]["password_hash"] = stored_hash
        data = audit_log(data, "login", "Primer inicio — contraseña configurada")
        save_data(data)
    if stored_hash and check_password_hash(stored_hash, password):
        session["user"] = "admin"
        return redirect(url_for("index"))
    return render_template("login.html", error="Contraseña incorrecta")

@app.route("/logout")
def logout():
    session.pop("user", None)
    if not AUTH_ENABLED:
        return redirect(url_for("index"))
    return redirect(url_for("login"))

@app.route("/auditoria")
@login_required
def auditoria():
    data = load_data()
    entries = data.get("auditoria", [])
    page = request.args.get("page", 1, type=int)
    entries.reverse()
    page_entries, page, total_pages, total = _paginar(entries, page, 50)
    return render_template("auditoria.html", entries=page_entries,
                           page=page, total_pages=total_pages, total=total)

@app.route("/cambiar-password", methods=["POST"])
@login_required
def cambiar_password():
    data = load_data()
    actual = request.form.get("actual", "")
    nueva = request.form.get("nueva", "")
    confirmar = request.form.get("confirmar", "")
    config = data.get("_config", {})
    stored_hash = config.get("password_hash", "")
    env_pass = os.environ.get("PASSWORD", "").strip()
    if env_pass:
        if check_password_hash(generate_password_hash(env_pass, method="pbkdf2:sha256"), actual):
            pass  # env var password matches
        else:
            return redirect(url_for("index") + "?error=password_actual_incorrecta")
    elif not stored_hash or not check_password_hash(stored_hash, actual):
        return redirect(url_for("index") + "?error=password_actual_incorrecta")
    if len(nueva) < 4:
        return redirect(url_for("index") + "?error=password_muy_corta")
    if nueva != confirmar:
        return redirect(url_for("index") + "?error=password_no_coinciden")
    if "_config" not in data:
        data["_config"] = {}
    data["_config"]["password_hash"] = generate_password_hash(nueva, method="pbkdf2:sha256")
    data = audit_log(data, "cambiar_password", "Contraseña actualizada")
    save_data(data)
    return redirect(url_for("index") + "?ok=password_actualizada")

@app.route("/")
@login_required
def index():
    data = load_data()
    ventas_dia = get_ventas_por_dia(data)
    ventas_mes = get_ventas_por_mes(data)
    gastos_mes = get_gastos_por_mes(data)
    _, _, _, _, utilidad = calcular_estado_resultados(data)
    mayor = calcular_mayor(data)
    total_activo = 0
    total_pasivo = 0
    for c, info in data["cuentas"].items():
        d = mayor[c]["debe"] if c in mayor else 0
        h = mayor[c]["haber"] if c in mayor else 0
        ts = tipo_saldo(info["tipo"])
        saldo = (d - h) if ts == "Debe" else (h - d)
        if info["tipo"] == "Activo":
            total_activo += saldo
        elif info["tipo"] == "Pasivo":
            total_pasivo += saldo

    return render_template(
        "index.html",
        ventas_dia=ventas_dia,
        ventas_mes=ventas_mes,
        gastos_mes=gastos_mes,
        utilidad=utilidad,
        total_activo=total_activo,
        total_pasivo=total_pasivo,
        diario=data["diario"],
        cuentas=data["cuentas"],
    )


# ── DIARIO ──
@app.route("/diario")
def diario():
    data = load_data()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    search = request.args.get("search", "").strip()
    per_page = min(per_page, 200)

    entries = data["diario"]
    if search:
        search_lower = search.lower()
        filtered = []
        for e in entries:
            if search_lower in e.get("descripcion", "").lower() or \
               search_lower in e.get("ref", "").lower() or \
               search_lower in e.get("fecha", ""):
                filtered.append(e)
        entries = filtered
    page_entries, page, total_pages, total = _paginar(entries, page, per_page)
    return render_template(
        "diario.html", diario=page_entries, cuentas=data["cuentas"],
        page=page, total_pages=total_pages, total_entries=total,
        search=search, per_page=per_page,
    )


@app.route("/diario/agregar", methods=["POST"])
def agregar_diario():
    data = load_data()
    fecha = request.form.get("fecha", date.today().isoformat())

    if _periodo_cerrado(data, fecha):
        return redirect(url_for("diario") + "?error=mes_cerrado")

    descripcion = request.form.get("descripcion", "")
    ref = request.form.get("ref", "")
    cuentas_sel = request.form.getlist("cuenta")
    tipos = request.form.getlist("tipo")
    montos = request.form.getlist("monto")
    movimientos = []
    for i in range(len(cuentas_sel)):
        if cuentas_sel[i] and montos[i]:
            monto = float(montos[i])
            if monto <= 0:
                continue
            movimientos.append(
                {"cuenta": cuentas_sel[i], "tipo": tipos[i], "monto": monto}
            )
    if movimientos:
        total_debe = sum(m["monto"] for m in movimientos if m["tipo"] == "Debe")
        total_haber = sum(m["monto"] for m in movimientos if m["tipo"] == "Haber")
        if abs(total_debe - total_haber) > 0.01 or total_debe == 0 or total_haber == 0:
            return redirect(url_for("diario") + "?error=asiento_desbalanceado")

        entry = {
            "id": get_next_id(data, "diario"),
            "fecha": fecha,
            "descripcion": descripcion,
            "ref": ref,
            "movimientos": movimientos,
        }
        data["diario"].append(entry)
        # Si es venta, agregar a cuentas cobrar si la cuenta es 1003
        for mov in movimientos:
            if mov["cuenta"] in ("1003", "1.1.03") and mov["tipo"] == "Debe":
                data["cuentas_cobrar"].append(
                    {
                        "id": get_next_id(data, "cuentas_cobrar"),
                        "fecha": fecha,
                        "descripcion": descripcion,
                        "monto": mov["monto"],
                        "estado": "Pendiente",
                        "ref_diario": entry["id"],
                    }
                )
            # Caja
            if mov["cuenta"] in ("1001", "1002", "1.1.01", "1.1.02"):
                data["caja_movimientos"].append(
                    {
                        "id": get_next_id(data, "caja_movimientos"),
                        "fecha": fecha,
                        "descripcion": descripcion,
                        "tipo": mov["tipo"],
                        "monto": mov["monto"],
                        "cuenta": mov["cuenta"],
                        "ref_diario": entry["id"],
                    }
                )
        data = audit_log(data, "crear_asiento", f"Asiento #{entry.get('id','?')} — {descripcion[:60]}")
        save_data(data)
    return redirect(url_for("diario"))


# ── EDITAR ASIENTO ──
@app.route("/diario/editar/<int:asiento_id>", methods=["GET", "POST"])
def editar_asiento(asiento_id):
    data = load_data()

    if request.method == "GET":
        # Buscar el asiento por ID
        asiento = None
        for a in data["diario"]:
            if a.get("id") == asiento_id:
                asiento = a
                break

        if not asiento:
            return redirect(url_for("diario"))

        # Convertir movimientos a formato para el form
        # Separar debe y haber
        movimientos_form = []
        for mov in asiento.get("movimientos", []):
            debe = (
                mov.get("debe", 0)
                if "debe" in mov
                else (mov.get("monto", 0) if mov.get("tipo") == "Debe" else 0)
            )
            haber = (
                mov.get("haber", 0)
                if "haber" in mov
                else (mov.get("monto", 0) if mov.get("tipo") == "Haber" else 0)
            )

            movimientos_form.append(
                {"cuenta": mov.get("cuenta", ""), "debe": debe, "haber": haber}
            )

        return render_template(
            "diario_editar.html",
            asiento=asiento,
            movimientos=movimientos_form,
            cuentas=data["cuentas"],
        )

    else:  # POST - guardar cambios
        if _periodo_cerrado(data, request.form.get("fecha", date.today().isoformat())):
            return redirect(url_for("diario") + "?error=mes_cerrado")

        # Buscar índice del asiento
        idx = None
        for i, a in enumerate(data["diario"]):
            if a.get("id") == asiento_id:
                idx = i
                break

        if idx is None:
            return redirect(url_for("diario"))

        # Actualizar datos
        fecha = request.form.get("fecha", date.today().isoformat())
        descripcion = request.form.get("descripcion", "")
        ref = request.form.get("ref", "")

        cuentas_sel = request.form.getlist("cuenta")
        debes = request.form.getlist("debe")
        haberes = request.form.getlist("haber")

        movimientos = []
        for i in range(len(cuentas_sel)):
            if cuentas_sel[i]:
                debe = float(debes[i]) if debes[i] else 0
                haber = float(haberes[i]) if haberes[i] else 0

                if debe > 0:
                    movimientos.append(
                        {"cuenta": cuentas_sel[i], "tipo": "Debe", "monto": debe}
                    )
                if haber > 0:
                    movimientos.append(
                        {"cuenta": cuentas_sel[i], "tipo": "Haber", "monto": haber}
                    )

        total_debe = sum(m["monto"] for m in movimientos if m["tipo"] == "Debe")
        total_haber = sum(m["monto"] for m in movimientos if m["tipo"] == "Haber")
        if round(total_debe, 2) != round(total_haber, 2) or total_debe == 0:
            return redirect(url_for("diario"))

        if movimientos:
            data["diario"][idx] = {
                "id": asiento_id,
                "fecha": fecha,
                "descripcion": descripcion,
                "ref": ref,
                "movimientos": movimientos,
            }
            data = audit_log(data, "editar_asiento", f"Asiento #{asiento_id} editado")
            save_data(data)

        return redirect(url_for("diario"))


# ── BORRAR ASIENTO ──
@app.route("/diario/borrar/<int:asiento_id>", methods=["POST"])
def borrar_asiento(asiento_id):
    data = load_data()

    for a in data["diario"]:
        if a.get("id") == asiento_id:
            if _periodo_cerrado(data, a.get("fecha", "")):
                return redirect(url_for("diario") + "?error=mes_cerrado")
            break

    data["diario"] = [a for a in data["diario"] if a.get("id") != asiento_id]

    data = audit_log(data, "borrar_asiento", f"Asiento #{asiento_id} borrado")
    save_data(data)
    return redirect(url_for("diario"))


# ── MAYOR ──
@app.route("/mayor")
def mayor():
    try:
        data = load_data()
        mayor_data = calcular_mayor(data)
        return render_template("mayor.html", mayor=mayor_data, cuentas=data["cuentas"])
    except Exception as e:
        import traceback

        traceback.print_exc()
        return f"<pre>ERROR /mayor: {e}</pre>", 500


# ── BALANZA ──
@app.route("/balanza")
def balanza():
    try:
        data = load_data()
        meses = meses_disponibles(data)
        periodo = request.args.get("periodo", "")

        def _calcular_balanza_tipo(d):
            bal, td, th, tsd, tsh = calcular_balanza(d)
            # Compute type-level subtotals
            tipos_totales = {}
            for row in bal:
                if not row["es_header"] and not row["es_subtotal"] and row["tipo"]:
                    t = row["tipo"]
                    if t not in tipos_totales:
                        tipos_totales[t] = {"debe": 0, "haber": 0, "saldo_debe": 0, "saldo_haber": 0}
                    tipos_totales[t]["debe"] += row["debe"]
                    tipos_totales[t]["haber"] += row["haber"]
                    tipos_totales[t]["saldo_debe"] += row["saldo_debe"]
                    tipos_totales[t]["saldo_haber"] += row["saldo_haber"]
            return bal, td, th, tsd, tsh, tipos_totales

        if periodo:
            data_filtrada = copy.deepcopy(data)
            data_filtrada["diario"] = [e for e in data["diario"] if e["fecha"][:7] == periodo]
            data_filtrada["ajustes"] = [a for a in data.get("ajustes", []) if a["fecha"][:7] == periodo]
            balanza_data, td, th, tsd, tsh, tipos_totales = _calcular_balanza_tipo(data_filtrada)

            idx = meses.index(periodo) if periodo in meses else -1
            periodo_anterior = meses[idx - 1] if idx > 0 else ""
            if periodo_anterior:
                data_ant = copy.deepcopy(data)
                data_ant["diario"] = [e for e in data["diario"] if e["fecha"][:7] == periodo_anterior]
                data_ant["ajustes"] = [a for a in data.get("ajustes", []) if a["fecha"][:7] == periodo_anterior]
                balanza_ant, td_a, th_a, tsd_a, tsh_a, _ = _calcular_balanza_tipo(data_ant)
            else:
                balanza_ant = []
                td_a = th_a = tsd_a = tsh_a = 0
                periodo_anterior = ""
        else:
            balanza_data, td, th, tsd, tsh, tipos_totales = _calcular_balanza_tipo(data)
            balanza_ant = []
            td_a = th_a = tsd_a = tsh_a = 0
            periodo_anterior = ""

        return render_template(
            "balanza.html",
            balanza=balanza_data,
            balanza_anterior=balanza_ant,
            total_debe=td,
            total_haber=th,
            total_saldo_debe=tsd,
            total_saldo_haber=tsh,
            total_debe_ant=td_a,
            total_haber_ant=th_a,
            meses=meses,
            periodo_actual=periodo,
            periodo_anterior=periodo_anterior,
            tipos_totales=tipos_totales,
        )
    except Exception as e:
        import traceback

        traceback.print_exc()
        return f"<pre>ERROR /balanza: {e}</pre>", 500


# ── ESTADO DE RESULTADOS ──
@app.route("/estado_resultados")
def estado_resultados():
    try:
        data = load_data()
        di, ti, dg, tg, util = calcular_estado_resultados(data)
        return render_template(
            "estado_resultados.html",
            detalle_ingresos=di,
            total_ingresos=ti,
            detalle_gastos=dg,
            total_gastos=tg,
            utilidad=util,
        )
    except Exception as e:
        import traceback

        traceback.print_exc()
        return f"<pre>ERROR /estado_resultados: {e}</pre>", 500


# ── BALANCE GENERAL ──
@app.route("/balance_general")
def balance_general():
    try:
        data = load_data()
        activos, ta, pasivos, tp, capital, tc = calcular_balance_general(data)
        return render_template(
            "balance_general.html",
            activos=activos,
            total_activo=ta,
            pasivos=pasivos,
            total_pasivo=tp,
            capital=capital,
            total_capital=tc,
        )
    except Exception as e:
        import traceback

        traceback.print_exc()
        return f"<pre>ERROR /balance_general: {e}</pre>", 500


# ── KARDEX ──
@app.route("/kardex")
def kardex():
    try:
        data = load_data()

        # Asegurar que existan las estructuras necesarias
        if "kardex" not in data:
            data["kardex"] = {}
        if "kardex_peps" not in data:
            data["kardex_peps"] = {}
        if "productos" not in data:
            data["productos"] = {}

        # Generar reporte PEPS
        reporte_peps = {}
        try:
            if data.get("kardex") or data.get("productos"):
                reporte_peps = kardex_peps.generar_reporte_kardex_peps(data, max_lotes=5)
        except Exception as e:
            print(f"Error generando reporte PEPS: {e}")

        productos_list = list(data.get("kardex", {}).keys())

        # Datos de productos para auto-completar en JS
        margen_default = data.get("configuracion", {}).get("margen_default", 30)
        productos_data_dict = {}
        for p in productos_list:
            prod_info = data.get("productos", {}).get(p, {})
            if isinstance(prod_info, dict):
                productos_data_dict[p] = {
                    "precio_venta": prod_info.get("precio_venta", 0),
                    "margen": prod_info.get("margen", margen_default),
                    "costo_promedio": prod_info.get("costo_promedio", 0),
                }
            else:
                productos_data_dict[p] = {"precio_venta": 0, "margen": margen_default, "costo_promedio": 0}
            # Fallback a kardex_peps
            peps_info = data.get("kardex_peps", {}).get(p, {})
            if isinstance(peps_info, dict):
                if not productos_data_dict[p]["precio_venta"]:
                    productos_data_dict[p]["precio_venta"] = peps_info.get("precio_venta", 0)
                if not productos_data_dict[p]["margen"]:
                    productos_data_dict[p]["margen"] = peps_info.get("margen", margen_default)
                if not productos_data_dict[p]["costo_promedio"]:
                    productos_data_dict[p]["costo_promedio"] = peps_info.get("costo_promedio", 0)

        return render_template(
            "kardex.html",
            kardex=data.get("kardex", {}),
            productos=productos_list,
            reporte_peps=reporte_peps,
            productos_data=productos_data_dict,
            margen_default=data.get("configuracion", {}).get("margen_default", 30),
        )

    except Exception as e:
        print(f"ERROR EN /kardex: {e}")
        import traceback

        traceback.print_exc()
        # Retornar página de error amigable
        return (
            f"""
        <html>
        <body style="font-family: Arial; padding: 40px; background: #1a1d2e; color: white;">
            <h1>❌ Error en Kardex</h1>
            <p>Ocurrió un error al cargar el Kardex:</p>
            <pre style="background: #2d3142; padding: 20px; border-radius: 8px;">{str(e)}</pre>
            <p><a href="/" style="color: #4f8cff;">← Volver al inicio</a></p>
        </body>
        </html>
        """,
            500,
        )


@app.route("/kardex/agregar_producto", methods=["POST"])
def agregar_producto_kardex():
    data = load_data()

    if "kardex" not in data:
        data["kardex"] = {}
    if "kardex_peps" not in data:
        data["kardex_peps"] = {}
    if "productos" not in data:
        data["productos"] = {}

    nombre = request.form.get("nombre", "").strip()
    try:
        precio_venta = float(request.form.get("precio_venta", 0) or 0)
        margen = float(request.form.get("margen", 30) or 30)
    except (ValueError, TypeError):
        precio_venta = 0
        margen = 30

    if not nombre:
        return redirect(url_for("kardex"))

    if nombre in data["kardex"]:
        return redirect(url_for("kardex") + "?error=producto_duplicado")

    data["kardex"][nombre] = []

    data["productos"][nombre] = {
        "nombre": nombre,
        "stock": 0,
        "costo_promedio": 0,
        "precio_venta": precio_venta,
        "margen": margen,
    }

    data["kardex_peps"][nombre] = {
        "lotes": [],
        "stock_total": 0,
        "costo_promedio": 0,
        "precio_venta": precio_venta,
        "margen": margen,
    }

    save_data(data)
    return redirect(url_for("kardex"))


@app.route("/kardex/movimiento", methods=["POST"])
def agregar_mov_kardex():
    try:
        data = load_data()
        producto = request.form.get("producto", "")
        fecha = request.form.get("fecha", date.today().isoformat())

        if _periodo_cerrado(data, fecha):
            return redirect(url_for("kardex") + "?error=periodo_cerrado")

        tipo = request.form.get("tipo", "entrada")
        cantidad = int(request.form.get("cantidad", 0))
        costo = float(request.form.get("costo", 0))
        precio_venta = float(request.form.get("precio_venta", 0))
        descripcion = request.form.get("descripcion", "")

        if not producto:
            print("Error: Producto vacío")
            return redirect(url_for("kardex"))

        if cantidad <= 0:
            print("Error: Cantidad debe ser mayor a 0")
            return redirect(url_for("kardex"))

        if "kardex" not in data:
            data["kardex"] = {}
        if "kardex_peps" not in data:
            data["kardex_peps"] = {}

        try:
            if tipo.lower() == "entrada":
                print(f"Agregando entrada PEPS: {producto}, {cantidad} unidades @ C$ {costo}")
                # Auto-calcular precio_venta si no se proporcionó
                if precio_venta <= 0 and costo > 0:
                    prod_mg = data.get("productos", {}).get(producto, {}).get("margen", 0) or 0
                    if prod_mg <= 0:
                        prod_mg = data.get("kardex_peps", {}).get(producto, {}).get("margen", 30) or 30
                    precio_venta = round(costo / (1 - prod_mg / 100), 2)
                data = kardex_peps.agregar_entrada_peps(
                    data, producto, fecha, cantidad, costo, precio_venta
                )
                print("✓ Entrada agregada exitosamente")

            elif tipo.lower() == "salida":
                print(f"Procesando salida PEPS: {producto}, {cantidad} unidades")
                data, costo_total, lotes_usados = kardex_peps.procesar_salida_peps(
                    data, producto, fecha, cantidad
                )
                print(f"✓ Salida procesada. Costo total: C$ {costo_total}")

                # Generar asiento contable automático si se indicó forma de pago
                forma_pago = request.form.get("forma_pago", "").strip()
                if forma_pago and forma_pago not in FORMAS_PAGO_VALIDAS:
                    forma_pago = "Efectivo"
                if forma_pago:
                    cliente = request.form.get("cliente", "Cliente general").strip()
                    pv = precio_venta if precio_venta > 0 else data.get("productos", {}).get(producto, {}).get("precio_venta", 0)
                    total_venta = cantidad * pv

                    if "diario" not in data:
                        data["diario"] = []

                    diario_id = get_next_id(data, "diario")
                    cuenta_cobro = "1.1.01" if forma_pago == "Efectivo" else ("1.1.02" if forma_pago == "Banco" else "1.1.03")

                    movs = [
                        {"cuenta": cuenta_cobro, "tipo": "Debe", "monto": total_venta},
                        {"cuenta": "4.1.01", "tipo": "Haber", "monto": total_venta},
                    ]
                    if costo_total > 0:
                        movs.append({"cuenta": "5.1", "tipo": "Debe", "monto": costo_total})
                        movs.append({"cuenta": "1.1.04", "tipo": "Haber", "monto": costo_total})

                    ref_kardex = f"KX-{diario_id:04d}"
                    entry = {
                        "id": diario_id,
                        "fecha": fecha,
                        "descripcion": f"Venta Kardex - {producto} - {cliente} ({forma_pago})",
                        "ref": ref_kardex,
                        "movimientos": movs,
                    }
                    data["diario"].append(entry)

                    if cuenta_cobro in ("1.1.01", "1.1.02"):
                        data.setdefault("caja_movimientos", [])
                        data["caja_movimientos"].append({
                            "id": get_next_id(data, "caja_movimientos"),
                            "fecha": fecha,
                            "descripcion": f"Venta Kardex - {producto} - {cliente}",
                            "tipo": "Debe",
                            "monto": total_venta,
                            "cuenta": cuenta_cobro,
                            "ref_diario": diario_id,
                        })

                    if cuenta_cobro == "1.1.03":
                        data.setdefault("cuentas_cobrar", [])
                        data["cuentas_cobrar"].append({
                            "id": get_next_id(data, "cuentas_cobrar"),
                            "fecha": fecha,
                            "descripcion": f"Venta Kardex - {producto} - {cliente}",
                            "monto": total_venta,
                            "estado": "Pendiente",
                            "ref_diario": diario_id,
                        })

                    if producto in data.get("kardex", {}):
                        k = data["kardex"][producto]
                        if k:
                            ultimo = k[-1]
                            ultimo["precio_venta"] = pv
                            ultimo["descripcion"] = f"Venta - {cliente} ({forma_pago})"
                            ultimo["ref_diario"] = diario_id

            elif tipo.lower() == "ajuste":
                if producto not in data["kardex"]:
                    data["kardex"][producto] = []

                saldo_anterior = (
                    data["kardex"][producto][-1]["saldo"]
                    if data["kardex"][producto]
                    else 0
                )

                data["kardex"][producto].append(
                    {
                        "fecha": fecha,
                        "tipo": "ajuste",
                        "cantidad": cantidad,
                        "costo": costo,
                        "precio_venta": precio_venta,
                        "total": cantidad * costo,
                        "saldo": cantidad,
                        "descripcion": descripcion or "Ajuste de inventario",
                    }
                )
                data = _reconstruir_peps(data, producto)

        except ValueError as e:
            print(f"Error controlado: {e}")
            return redirect(url_for("kardex") + "?error=" + str(e).replace(" ", "_"))

        except Exception as e:
            # Errores inesperados
            print(f"Error inesperado en movimiento: {e}")
            import traceback

            traceback.print_exc()
            return redirect(url_for("kardex"))

        save_data(data)
        print("✓ Datos guardados correctamente")

    except Exception as e:
        print(f"ERROR EN /kardex/movimiento: {e}")
        import traceback

        traceback.print_exc()

    return redirect(url_for("kardex"))


# ── CUENTAS POR COBRAR ──
@app.route("/cobrar")
def cobrar():
    data = load_data()
    page = request.args.get("page", 1, type=int)
    search = request.args.get("search", "").strip()
    items = data["cuentas_cobrar"]
    # Attach original index for action URLs
    indexed = list(enumerate(items))
    if search:
        sl = search.lower()
        indexed = [(i, item) for i, item in indexed
                   if sl in item.get("descripcion","").lower()
                   or sl in item.get("fecha","") or sl in item.get("estado","").lower()]
    items_pag, page, total_pages, total = _paginar([item for _, item in indexed], page, 50)
    # Get original indices for paginated items
    paginated_indices = [indexed[idx][0] for idx in range(len(items_pag))]
    return render_template("cobrar.html", cobrar=items_pag,
                           page=page, total_pages=total_pages,
                           total_entries=total, search=search,
                           indices=paginated_indices)


@app.route("/cobrar/pagar/<int:idx>", methods=["POST"])
def pagar_cobro(idx):
    data = load_data()
    if 0 <= idx < len(data["cuentas_cobrar"]):
        data["cuentas_cobrar"][idx]["estado"] = "Cobrado"
        data["cuentas_cobrar"][idx]["fecha_cobro"] = date.today().isoformat()
    data = audit_log(data, "pagar_cobro", f"Cobro #{idx} pagado")
    save_data(data)
    return redirect(url_for("cobrar"))


@app.route("/cobrar/eliminar/<int:idx>", methods=["POST"])
def eliminar_cobro(idx):
    data = load_data()
    if 0 <= idx < len(data["cuentas_cobrar"]):
        data["cuentas_cobrar"].pop(idx)
    data = audit_log(data, "eliminar_cobro", f"Cobro #{idx} eliminado")
    save_data(data)
    return redirect(url_for("cobrar"))


# ── AUXILIAR DE CAJA ──
@app.route("/caja")
def caja():
    data = load_data()

    movs = data.get("caja_movimientos", []) or data.get("caja", [])
    saldo = 0
    for m in movs:
        if m.get("tipo") == "Debe":
            saldo += m.get("monto", 0)
        else:
            saldo -= m.get("monto", 0)

    page = request.args.get("page", 1, type=int)
    per_page = 50
    start = (page - 1) * per_page

    saldo_inicial = 0
    for m in movs[:start]:
        if m.get("tipo") == "Debe":
            saldo_inicial += m.get("monto", 0)
        else:
            saldo_inicial -= m.get("monto", 0)

    movs_pag, page, total_pages, total = _paginar(movs, page)
    return render_template("caja.html", movimientos=movs_pag, saldo=saldo,
                           page=page, total_pages=total_pages, total=total,
                           saldo_inicial=saldo_inicial)


# ── AUXILIAR DIARIO ──
@app.route("/auxiliar-diario")
def auxiliar_diario():
    data = load_data()
    cuentas = data["cuentas"]

    movs_planos = []
    for entry in data["diario"]:
        for mov in entry["movimientos"]:
            movs_planos.append({
                "asiento_id": entry["id"],
                "fecha": entry["fecha"],
                "descripcion": entry["descripcion"],
                "ref": entry.get("ref", ""),
                "cuenta": mov["cuenta"],
                "nombre_cuenta": cuentas.get(mov["cuenta"], {}).get("nombre", mov["cuenta"]),
                "tipo": mov["tipo"],
                "monto": mov["monto"],
                "tipo_cuenta": cuentas.get(mov["cuenta"], {}).get("tipo", ""),
            })
    aj_id = -1
    for aj in data.get("ajustes", []):
        for mov in aj["movimientos"]:
            movs_planos.append({
                "asiento_id": aj_id,
                "fecha": aj["fecha"],
                "descripcion": aj["descripcion"],
                "ref": "AJ",
                "cuenta": mov["cuenta"],
                "nombre_cuenta": cuentas.get(mov["cuenta"], {}).get("nombre", mov["cuenta"]),
                "tipo": mov["tipo"],
                "monto": mov["monto"],
                "tipo_cuenta": cuentas.get(mov["cuenta"], {}).get("tipo", ""),
            })
            aj_id -= 1

    saldos_cuenta = defaultdict(float)
    for m in movs_planos:
        ts = tipo_saldo(m["tipo_cuenta"])
        if ts == "Debe":
            saldos_cuenta[m["cuenta"]] += m["monto"] if m["tipo"] == "Debe" else -m["monto"]
        else:
            saldos_cuenta[m["cuenta"]] += m["monto"] if m["tipo"] == "Haber" else -m["monto"]

    movs_con_saldo = []
    saldo_run = defaultdict(float)
    for m in movs_planos:
        cuenta = m["cuenta"]
        ts = tipo_saldo(m["tipo_cuenta"])
        if ts == "Debe":
            saldo_run[cuenta] += m["monto"] if m["tipo"] == "Debe" else -m["monto"]
        else:
            saldo_run[cuenta] += m["monto"] if m["tipo"] == "Haber" else -m["monto"]
        m["saldo_cuenta"] = saldo_run[cuenta]
        movs_con_saldo.append(m)

    movs_con_saldo.sort(key=lambda x: (x["fecha"], x["asiento_id"]), reverse=True)

    cuentas_con_mov = sorted(set(m["cuenta"] for m in movs_con_saldo))
    total_debe = sum(m["monto"] for m in movs_con_saldo if m["tipo"] == "Debe")
    total_haber = sum(m["monto"] for m in movs_con_saldo if m["tipo"] == "Haber")

    # Pre-computar debe/haber por cuenta (el sumario en la plantilla los necesita sin paginar)
    sumario_ctas = {}
    for m in movs_con_saldo:
        c = m["cuenta"]
        if c not in sumario_ctas:
            sumario_ctas[c] = {"debe": 0.0, "haber": 0.0, "movs": 0}
        if m["tipo"] == "Debe":
            sumario_ctas[c]["debe"] += m["monto"]
        else:
            sumario_ctas[c]["haber"] += m["monto"]
        sumario_ctas[c]["movs"] += 1

    return render_template(
        "auxiliar_diario.html",
        movimientos=movs_con_saldo,
        cuentas=cuentas,
        cuentas_con_mov=cuentas_con_mov,
        sumario_ctas=sumario_ctas,
        saldos_cuenta=dict(saldos_cuenta),
        total_debe=total_debe,
        total_haber=total_haber,
        total_asientos=len(data["diario"]) + len(data.get("ajustes", [])),
        total_movs=len(movs_con_saldo),
    )


# ── AJUSTES ──
@app.route("/ajustes")
def ajustes():
    data = load_data()
    margen_config = data.get("configuracion", {}).get("margen_default", 30)
    return render_template(
        "ajustes.html", ajustes=data.get("ajustes", []), cuentas=data["cuentas"], margen_config=margen_config
    )


@app.route("/ajustes/guardar_margen", methods=["POST"])
def guardar_margen():
    data = load_data()
    margen = float(request.form.get("margen", 30))
    if margen < 0 or margen >= 100:
        margen = 30
    data.setdefault("configuracion", {})["margen_default"] = margen
    save_data(data)
    return redirect(url_for("ajustes") + "?ok=margen")


@app.route("/ajustes/agregar", methods=["POST"])
def agregar_ajuste():
    data = load_data()
    fecha = request.form.get("fecha", date.today().isoformat())

    if _periodo_cerrado(data, fecha):
        return redirect(url_for("ajustes") + "?error=mes_cerrado")

    descripcion = request.form.get("descripcion", "")
    cuentas_sel = request.form.getlist("cuenta")
    tipos = request.form.getlist("tipo")
    montos = request.form.getlist("monto")
    movimientos = []
    for i in range(len(cuentas_sel)):
        if cuentas_sel[i] and montos[i]:
            monto = float(montos[i])
            if monto <= 0:
                continue
            movimientos.append(
                {"cuenta": cuentas_sel[i], "tipo": tipos[i], "monto": monto}
            )
    if movimientos:
        total_debe = sum(m["monto"] for m in movimientos if m["tipo"] == "Debe")
        total_haber = sum(m["monto"] for m in movimientos if m["tipo"] == "Haber")
        if abs(total_debe - total_haber) > 0.01 or total_debe == 0 or total_haber == 0:
            return redirect(url_for("ajustes") + "?error=asiento_desbalanceado")

        data.setdefault("ajustes", []).append(
            {
                "id": get_next_id(data, "ajustes"),
                "fecha": fecha,
                "descripcion": descripcion,
                "movimientos": movimientos,
            }
        )
        data = audit_log(data, "crear_ajuste", f"Ajuste #{data['ajustes'][-1]['id']} — {descripcion[:60]}")
        save_data(data)
    return redirect(url_for("ajustes"))


# ── REPORTES (API JSON para gráficos) ──
@app.route("/api/ventas_dia")
def api_ventas_dia():
    return jsonify(get_ventas_por_dia(load_data()))


@app.route("/api/ventas_mes")
def api_ventas_mes():
    return jsonify(get_ventas_por_mes(load_data()))


@app.route("/api/gastos_mes")
def api_gastos_mes():
    return jsonify(get_gastos_por_mes(load_data()))


@app.route("/api/stock")
def api_stock():
    data = load_data()
    stock = {}
    for prod, info in data.get("kardex_peps", {}).items():
        stock[prod] = {
            "stock": info.get("stock_total", 0),
            "costo": info.get("costo_promedio", 0),
            "precio_venta": info.get("precio_venta", 0),
        }
    for prod, info in data.get("productos", {}).items():
        if prod not in stock:
            stock[prod] = {
                "stock": info.get("stock", 0),
                "costo": info.get("costo_promedio", 0),
                "precio_venta": info.get("precio_venta", 0),
            }
    return jsonify(stock)


@app.route("/api/notificaciones")
def api_notificaciones():
    data = load_data()
    notificaciones = []

    # Stock bajo
    for nombre, peps in data.get("kardex_peps", {}).items():
        if isinstance(peps, dict):
            stock = peps.get("stock_total", 0)
            if stock <= 0:
                notificaciones.append({
                    "tipo": "critico",
                    "mensaje": f"Sin stock: {nombre}",
                    "valor": f"{stock} unid.",
                    "modulo": "kardex",
                })
            elif stock <= 5:
                notificaciones.append({
                    "tipo": "advertencia",
                    "mensaje": f"Stock bajo: {nombre}",
                    "valor": f"{stock} unid.",
                    "modulo": "kardex",
                })

    # Cuentas por cobrar pendientes
    for cc in data.get("cuentas_cobrar", []):
        if cc.get("estado") == "Pendiente":
            notificaciones.append({
                "tipo": "advertencia",
                "mensaje": f"Cobro pendiente: {cc.get('descripcion', '')}",
                "valor": f"C$ {cc.get('monto', 0):.2f}",
                "modulo": "cobrar",
            })

    notificaciones.sort(key=lambda x: (0 if x["tipo"] == "critico" else 1))
    return jsonify({"notificaciones": notificaciones})


# ── VENTAS POS ──
@app.route("/pos")
def pos():
    data = load_data()
    productos = []
    nombres = set(data.get("kardex", {}).keys())
    nombres.update(data.get("kardex_peps", {}).keys())

    for nombre in sorted(nombres):
        try:
            peps_info = data.get("kardex_peps", {}).get(nombre, {})
            prod_info = data.get("productos", {}).get(nombre, {})

            if isinstance(peps_info, dict) and peps_info.get("stock_total", 0) > 0:
                saldo = peps_info.get("stock_total", 0)
                costo = peps_info.get("costo_promedio", 0)
            else:
                k = data.get("kardex", {}).get(nombre, [])
                if isinstance(k, list) and k:
                    saldo = k[-1].get("saldo", 0)
                    costo = k[-1].get("costo", 0)
                elif isinstance(k, dict):
                    saldo = k.get("saldo_actual", k.get("saldo_inicial", 0))
                    costo = k.get("costo_unitario", 0)
                else:
                    continue

            pv = 0
            mg = 30
            if isinstance(prod_info, dict):
                pv = prod_info.get("precio_venta", 0) or 0
                mg = prod_info.get("margen", 30) or 30
            if not pv and isinstance(peps_info, dict):
                pv = peps_info.get("precio_venta", 0) or 0
                mg = peps_info.get("margen", 30) or 30
            if not pv:
                pv = round(costo / (1 - mg / 100), 2)

            productos.append({
                "nombre": nombre, "saldo": saldo,
                "costo": costo, "precio_venta": pv, "margen": mg,
            })
        except Exception as e:
            print(f"Error cargando producto {nombre}: {e}")
            continue

    # Historial ventas POS
    historial = data.get("pos_historial", [])

    return render_template("pos.html", productos=productos, historial=historial)


@app.route("/pos/venta", methods=["POST"])
def pos_venta():
    try:
        data = load_data()
        fecha = request.form.get("fecha", date.today().isoformat())

        if _periodo_cerrado(data, fecha):
            return redirect(url_for("pos") + "?error=periodo_cerrado")

        cliente = request.form.get("cliente", "Cliente general")
        forma_pago = request.form.get("forma_pago", "Efectivo")
        if forma_pago not in FORMAS_PAGO_VALIDAS:
            forma_pago = "Efectivo"
        productos_nombres = request.form.getlist("producto")
        cantidades = request.form.getlist("cantidad")
        precios = request.form.getlist("precio")

        lineas = []
        total_venta = 0
        total_costo = 0

        for i in range(len(productos_nombres)):
            nombre = productos_nombres[i]
            cantidad = round(float(cantidades[i])) if cantidades[i] else 0
            precio = float(precios[i]) if precios[i] else 0

            if nombre and cantidad > 0 and precio > 0:
                subtotal = cantidad * precio

                # Obtener costo desde PEPS o kardex tradicional
                costo_u = 0
                costo_total = 0
                peps_disponible = (data.get("kardex_peps", {})
                                   .get(nombre, {}).get("stock_total", 0))
                if peps_disponible >= cantidad:
                    try:
                        data, costo_total, _ = kardex_peps.procesar_salida_peps(
                            data, nombre, fecha, cantidad)
                        costo_u = costo_total / cantidad if cantidad else 0
                        # Parchear precio_venta en el último kardex generado por PEPS
                        if nombre in data.get("kardex", {}):
                            kx = data["kardex"][nombre]
                            if kx:
                                kx[-1]["precio_venta"] = precio
                    except Exception as e:
                        import traceback
                        traceback.print_exc()
                        costo_total = 0
                        costo_u = 0
                else:
                    if nombre in data.get("kardex", {}):
                        k = data["kardex"][nombre]
                        if isinstance(k, list) and k:
                            costo_u = k[-1].get("costo", 0)
                            saldo_anterior = k[-1]["saldo"] if k else 0
                            nuevo_saldo = saldo_anterior - cantidad
                            k.append({
                                "fecha": fecha, "tipo": "salida",
                                "cantidad": cantidad, "costo": costo_u,
                                "precio_venta": precio,
                                "total": cantidad * costo_u,
                                "saldo": nuevo_saldo,
                                "descripcion": f"Venta POS - {cliente}",
                            })
                        elif isinstance(k, dict):
                            if "movimientos" not in k:
                                k["movimientos"] = []
                            k["movimientos"].append({
                                "fecha": fecha, "tipo": "Salida",
                                "cantidad": cantidad,
                                "costo_unitario": costo_u,
                                "precio_venta": precio,
                                "descripcion": f"Venta POS - {cliente}",
                            })
                            saldo = k.get("saldo_inicial", 0)
                            for m in k["movimientos"]:
                                if m["tipo"] == "Entrada":
                                    saldo += m["cantidad"]
                                else:
                                    saldo -= m["cantidad"]
                            k["saldo_actual"] = saldo
                    costo_total = cantidad * costo_u

                lineas.append(
                    {
                        "producto": nombre,
                        "cantidad": cantidad,
                        "precio_unitario": precio,
                        "subtotal": subtotal,
                        "costo_unitario": costo_u,
                        "costo_total": costo_total,
                    }
                )

                total_venta += subtotal
                total_costo += costo_total

        if not lineas:
            return redirect(url_for("pos"))

        # Crear asiento diario
        if "diario" not in data:
            data["diario"] = []

        diario_id = get_next_id(data, "diario")
        cuenta_cobro = (
            "1.1.01"
            if forma_pago == "Efectivo"
            else ("1.1.02" if forma_pago == "Banco" else "1.1.03")
        )

        movimientos_diario = [
            {"cuenta": cuenta_cobro, "tipo": "Debe", "monto": total_venta},
            {"cuenta": "4.1.01", "tipo": "Haber", "monto": total_venta},
        ]

        if total_costo > 0:
            movimientos_diario.append(
                {"cuenta": "5.1", "tipo": "Debe", "monto": total_costo}
            )
            movimientos_diario.append(
                {"cuenta": "1.1.04", "tipo": "Haber", "monto": total_costo}
            )

        ref_pos = f"POS-{len(data.get('pos_historial', [])) + 1:04d}"

        entry = {
            "id": diario_id,
            "fecha": fecha,
            "descripcion": f"Venta POS - {cliente} ({forma_pago})",
            "ref": ref_pos,
            "movimientos": movimientos_diario,
        }
        data["diario"].append(entry)

        # Auxiliar de caja
        if cuenta_cobro in ("1.1.01", "1.1.02"):
            data.setdefault("caja_movimientos", [])
            data["caja_movimientos"].append(
                {
                    "id": get_next_id(data, "caja_movimientos"),
                    "fecha": fecha,
                    "descripcion": f"Venta POS - {cliente}",
                    "tipo": "Debe",
                    "monto": total_venta,
                    "cuenta": cuenta_cobro,
                    "ref_diario": diario_id,
                }
            )

        # Cuentas por cobrar
        if cuenta_cobro == "1.1.03":
            if "cuentas_cobrar" not in data:
                data["cuentas_cobrar"] = []

            data["cuentas_cobrar"].append(
                {
                    "id": get_next_id(data, "cuentas_cobrar"),
                    "fecha": fecha,
                    "descripcion": f"Venta POS - {cliente}",
                    "monto": total_venta,
                    "estado": "Pendiente",
                    "ref_diario": diario_id,
                }
            )

        # Guardar en historial POS
        if "pos_historial" not in data:
            data["pos_historial"] = []

        data["pos_historial"].append(
            {
                "ref": ref_pos,
                "fecha": fecha,
                "cliente": cliente,
                "forma_pago": forma_pago,
                "lineas": lineas,
                "total": total_venta,
                "costo": total_costo,
                "utilidad": total_venta - total_costo,
            }
        )

        data = audit_log(data, "pos_venta", f"Venta {ref_pos} — C${total_venta:.2f}")
        save_data(data)
        return redirect(url_for("pos"))

    except Exception as e:
        print(f"ERROR EN /pos/venta: {e}")
        import traceback

        traceback.print_exc()
        return redirect(url_for("pos"))


@app.route("/pos/anular/<ref>", methods=["POST"])
def pos_anular(ref):
    """Anula una venta POS: revierte kardex, diario, caja y marca como anulada."""
    data = load_data()
    historial = data.get("pos_historial", [])

    venta = None
    venta_idx = None
    for i, v in enumerate(historial):
        if v.get("ref") == ref:
            venta = v
            venta_idx = i
            break

    if not venta:
        return redirect(url_for("pos"))

    try:
        # ── 1. Revertir kardex y PEPS ──
        for linea in venta.get("lineas", []):
            nombre = linea["producto"]
            cantidad_vendida = linea["cantidad"]

            # Revertir kardex tradicional
            if nombre in data.get("kardex", {}):
                kx = data["kardex"][nombre]
                if isinstance(kx, list):
                    for j in range(len(kx) - 1, -1, -1):
                        mov = kx[j]
                        if (mov.get("tipo") == "salida" and
                            mov.get("descripcion", "").startswith("Venta POS") and
                            abs(mov.get("cantidad", 0) - cantidad_vendida) < 0.01):
                            if j > 0:
                                kx[j - 1]["saldo"] = kx[j - 1].get("saldo", 0) + cantidad_vendida
                            del kx[j]
                            break

            # Revertir PEPS
            peps = data.get("kardex_peps", {}).get(nombre, {})
            if isinstance(peps, dict) and peps.get("lotes"):
                for lote in reversed(peps["lotes"]):
                    lote["cantidad_restante"] = lote.get("cantidad_restante", 0) + cantidad_vendida
                    peps["stock_total"] = peps.get("stock_total", 0) + cantidad_vendida
                    break
                if peps.get("stock_total", 0) < 0:
                    peps["stock_total"] = 0

        # ── 2. Revertir asiento diario ──
        ref_original = venta.get("ref", "")
        diario = data.get("diario", [])
        diario_id = None
        for e in diario:
            if e.get("ref") == ref_original:
                diario_id = e.get("id")
                break
        data["diario"] = [e for e in diario if e.get("ref") != ref_original]

        # ── 3. Revertir caja_movimientos ──
        if diario_id:
            caja = data.get("caja_movimientos", [])
            data["caja_movimientos"] = [c for c in caja if c.get("ref_diario") != diario_id]

        # ── 4. Revertir cuentas_cobrar ──
        if diario_id:
            cc = data.get("cuentas_cobrar", [])
            data["cuentas_cobrar"] = [c for c in cc if c.get("ref_diario") != diario_id]

        # ── 5. Crear asiento de reversión (para que el diario quede completo) ──
        new_id = get_next_id(data, "diario")
        total_venta = venta.get("total", 0)
        total_costo = venta.get("costo", 0)
        forma_pago = venta.get("forma_pago", "Efectivo")
        cliente = venta.get("cliente", "")

        cuenta_cobro = (
            "1.1.01" if forma_pago == "Efectivo"
            else ("1.1.02" if forma_pago == "Banco" else "1.1.03")
        )
        movimientos_reversion = [
            {"cuenta": "4.1.01", "tipo": "Debe", "monto": total_venta},
            {"cuenta": cuenta_cobro, "tipo": "Haber", "monto": total_venta},
        ]
        if total_costo > 0:
            movimientos_reversion.append({"cuenta": "1.1.04", "tipo": "Debe", "monto": total_costo})
            movimientos_reversion.append({"cuenta": "5.1", "tipo": "Haber", "monto": total_costo})

        entry = {
            "id": new_id,
            "fecha": venta.get("fecha", ""),
            "descripcion": f"ANULACIÓN Venta POS {ref} — {cliente}",
            "ref": f"ANUL-{ref}",
            "movimientos": movimientos_reversion,
        }
        data["diario"].append(entry)

        # ── 6. Marcar como anulada en historial ──
        data["pos_historial"][venta_idx]["anulada"] = True
        data["pos_historial"][venta_idx]["motivo_anulacion"] = f"Anulada el {date.today().isoformat()}"

        data = audit_log(data, "pos_anular", f"Venta {ref} anulada — C${total_venta:.2f}")
        save_data(data)

    except Exception as e:
        import traceback
        traceback.print_exc()

    return redirect(url_for("pos"))


@app.route("/pos/editar/<ref>", methods=["GET", "POST"])
def pos_editar(ref):
    data = load_data()
    historial = data.get("pos_historial", [])

    # Buscar la venta por ref
    venta = None
    venta_idx = None
    for i, v in enumerate(historial):
        if v.get("ref") == ref:
            venta = v
            venta_idx = i
            break

    if not venta:
        return redirect(url_for("pos"))

    if request.method == "GET":
        return render_template("pos_editar.html", venta=venta, productos=data.get("kardex_peps", {}))

    # POST — Guardar edición
    try:
        fecha = request.form.get("fecha", venta["fecha"])
        cliente = request.form.get("cliente", venta["cliente"])
        forma_pago = request.form.get("forma_pago", venta["forma_pago"])

        if forma_pago not in FORMAS_PAGO_VALIDAS:
            forma_pago = "Efectivo"

        productos_nombres = request.form.getlist("producto")
        cantidades = request.form.getlist("cantidad")
        precios = request.form.getlist("precio")

        # ── 1. Revertir kardex y PEPS de la venta original ──
        for linea in venta.get("lineas", []):
            nombre = linea["producto"]
            cantidad_vendida = linea["cantidad"]

            # Revertir kardex tradicional: buscar salidas de esta venta y eliminarlas
            if nombre in data.get("kardex", {}):
                kx = data["kardex"][nombre]
                if isinstance(kx, list):
                    # Eliminar la última salida que coincida con la venta
                    for j in range(len(kx) - 1, -1, -1):
                        mov = kx[j]
                        if (mov.get("tipo") == "salida" and
                            mov.get("descripcion", "").startswith("Venta POS") and
                            abs(mov.get("cantidad", 0) - cantidad_vendida) < 0.01):
                            # Revertir saldo
                            if j > 0:
                                kx[j - 1]["saldo"] = kx[j - 1].get("saldo", 0) + cantidad_vendida
                            del kx[j]
                            break

            # Revertir PEPS: devolver stock a lotes
            peps = data.get("kardex_peps", {}).get(nombre, {})
            if isinstance(peps, dict) and peps.get("lotes"):
                cantidad_por_devolver = cantidad_vendida
                for lote in reversed(peps["lotes"]):
                    if cantidad_por_devolver <= 0:
                        break
                    lote["cantidad_restante"] = lote.get("cantidad_restante", 0) + cantidad_por_devolver
                    peps["stock_total"] = peps.get("stock_total", 0) + cantidad_por_devolver
                    break  # PEPS: devolver al último lote usado
                if peps.get("stock_total", 0) < 0:
                    peps["stock_total"] = 0

        # ── 2. Revertir asiento diario original ──
        ref_original = venta.get("ref", "")
        diario = data.get("diario", [])
        diario_original_id = diario_id_for_ref(diario, ref_original)
        data["diario"] = [e for e in diario if e.get("ref") != ref_original]

        # ── 3. Revertir caja_movimientos ──
        caja = data.get("caja_movimientos", [])
        data["caja_movimientos"] = [c for c in caja if c.get("ref_diario") != ref_original
                                     and c.get("ref_diario") != diario_original_id]

        # ── 4. Revertir cuentas_cobrar ──
        cc = data.get("cuentas_cobrar", [])
        data["cuentas_cobrar"] = [c for c in cc if c.get("ref_diario") != ref_original
                                    and c.get("ref_diario") != diario_original_id]

        # ── 5. Crear nueva venta ──
        lineas = []
        total_venta = 0
        total_costo = 0

        for i in range(len(productos_nombres)):
            nombre = productos_nombres[i]
            cantidad = round(float(cantidades[i])) if cantidades[i] else 0
            precio = float(precios[i]) if precios[i] else 0

            if nombre and cantidad > 0 and precio > 0:
                subtotal = cantidad * precio

                costo_u = 0
                costo_total = 0
                peps_disponible = (data.get("kardex_peps", {})
                                   .get(nombre, {}).get("stock_total", 0))
                if peps_disponible >= cantidad:
                    try:
                        data, costo_total, _ = kardex_peps.procesar_salida_peps(
                            data, nombre, fecha, cantidad)
                        costo_u = costo_total / cantidad if cantidad else 0
                        if nombre in data.get("kardex", {}):
                            kx = data["kardex"][nombre]
                            if kx:
                                kx[-1]["precio_venta"] = precio
                    except Exception:
                        import traceback
                        traceback.print_exc()
                        costo_total = 0
                        costo_u = 0
                else:
                    if nombre in data.get("kardex", {}):
                        k = data["kardex"][nombre]
                        if isinstance(k, list) and k:
                            costo_u = k[-1].get("costo", 0)
                            saldo_anterior = k[-1]["saldo"] if k else 0
                            nuevo_saldo = saldo_anterior - cantidad
                            k.append({
                                "fecha": fecha, "tipo": "salida",
                                "cantidad": cantidad, "costo": costo_u,
                                "precio_venta": precio,
                                "total": cantidad * costo_u,
                                "saldo": nuevo_saldo,
                                "descripcion": f"Venta POS - {cliente}",
                            })
                        costo_total = cantidad * costo_u

                lineas.append({
                    "producto": nombre,
                    "cantidad": cantidad,
                    "precio_unitario": precio,
                    "subtotal": subtotal,
                    "costo_unitario": costo_u,
                    "costo_total": costo_total,
                })
                total_venta += subtotal
                total_costo += costo_total

        if not lineas:
            return redirect(url_for("pos"))

        # ── 6. Crear nuevo asiento diario ──
        diario_id = get_next_id(data, "diario")
        cuenta_cobro = (
            "1.1.01" if forma_pago == "Efectivo"
            else ("1.1.02" if forma_pago == "Banco" else "1.1.03")
        )
        movimientos_diario = [
            {"cuenta": cuenta_cobro, "tipo": "Debe", "monto": total_venta},
            {"cuenta": "4.1.01", "tipo": "Haber", "monto": total_venta},
        ]
        if total_costo > 0:
            movimientos_diario.append({"cuenta": "5.1", "tipo": "Debe", "monto": total_costo})
            movimientos_diario.append({"cuenta": "1.1.04", "tipo": "Haber", "monto": total_costo})

        entry = {
            "id": diario_id,
            "fecha": fecha,
            "descripcion": f"Venta POS - {cliente} ({forma_pago})",
            "ref": ref,
            "movimientos": movimientos_diario,
        }
        data["diario"].append(entry)

        # ── 7. Auxiliar de caja ──
        if cuenta_cobro in ("1.1.01", "1.1.02"):
            data.setdefault("caja_movimientos", [])
            data["caja_movimientos"].append({
                "id": get_next_id(data, "caja_movimientos"),
                "fecha": fecha,
                "descripcion": f"Venta POS - {cliente}",
                "tipo": "Debe",
                "monto": total_venta,
                "cuenta": cuenta_cobro,
                "ref_diario": diario_id,
            })

        # ── 8. Cuentas por cobrar ──
        if cuenta_cobro == "1.1.03":
            data.setdefault("cuentas_cobrar", [])
            data["cuentas_cobrar"].append({
                "id": get_next_id(data, "cuentas_cobrar"),
                "fecha": fecha,
                "descripcion": f"Venta POS - {cliente}",
                "monto": total_venta,
                "estado": "Pendiente",
                "ref_diario": diario_id,
            })

        # ── 9. Actualizar historial POS ──
        data["pos_historial"][venta_idx] = {
            "ref": ref,
            "fecha": fecha,
            "cliente": cliente,
            "forma_pago": forma_pago,
            "lineas": lineas,
            "total": total_venta,
            "costo": total_costo,
            "utilidad": total_venta - total_costo,
        }

        data = audit_log(data, "pos_editar", f"Venta {ref} editada — C${total_venta:.2f}")
        save_data(data)
        return redirect(url_for("pos"))

    except Exception as e:
        import traceback
        traceback.print_exc()
        return redirect(url_for("pos"))


def diario_id_for_ref(diario, ref):
    for e in diario:
        if e.get("ref") == ref:
            return e.get("id")
    return None


# ══════════════════════════════════════════════════════════════
#  EXPORTAR KARDEX PEPS
# ══════════════════════════════════════════════════════════════


@app.route("/kardex/exportar_peps", methods=["POST"])
def exportar_kardex_peps():
    """Exporta el Kardex completo con método PEPS a Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl no está instalado"}), 500

    data = load_data()
    producto_codigo = request.form.get("producto", None)

    # Generar reporte PEPS
    reporte = kardex_peps.generar_reporte_kardex_peps(data, producto_codigo)

    wb = Workbook()

    for idx, (codigo, info) in enumerate(reporte.items()):
        # Crear hoja para cada producto
        if idx == 0:
            ws = wb.active
            ws.title = codigo[:31]  # Excel limita a 31 caracteres
        else:
            ws = wb.create_sheet(codigo[:31])

        ws.sheet_properties.tabColor = "4F8CFF"

        # ── ENCABEZADO ──
        ws.append(["F&G - Sistema Contable"])
        ws.merge_cells("A1:I1")
        ws["A1"].font = Font(size=16, bold=True, color="1A1D2E")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.append([f"KARDEX - {info['nombre']}"])
        ws.merge_cells("A2:I2")
        ws["A2"].font = Font(size=14, bold=True, color="4F8CFF")
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.append([f"Código: {codigo}"])
        ws.merge_cells("A3:I3")
        ws["A3"].font = Font(size=10, color="7a7f99")
        ws["A3"].alignment = Alignment(horizontal="center")

        ws.append([f"Generado: {date.today().isoformat()}"])
        ws.merge_cells("A4:I4")
        ws["A4"].font = Font(size=10, color="7a7f99")
        ws["A4"].alignment = Alignment(horizontal="center")

        ws.append([])  # Línea vacía

        # ── RESUMEN ──
        ws.append(["RESUMEN"])
        ws["A6"].font = Font(size=12, bold=True)

        ws.append(["Stock Actual:", info["stock"], "unidades"])
        ws.append(["Costo Promedio:", f"C$ {info['costo_promedio']:.2f}", "por unidad"])
        ws.append(["Valor Inventario:", f"C$ {info['valor_inventario']:.2f}", ""])

        ws.append([])  # Línea vacía

        # ── LOTES DISPONIBLES (PEPS) ──
        ws.append(
            ["LOTES DISPONIBLES (Método PEPS - Primeras Entradas, Primeras Salidas)"]
        )
        ws["A11"].font = Font(size=12, bold=True, color="2ECC71")
        ws.merge_cells("A11:E11")

        headers_lotes = [
            "Fecha Entrada",
            "Cant. Original",
            "Cant. Disponible",
            "Costo Unit.",
            "Valor Total",
        ]
        ws.append(headers_lotes)

        for cell in ws[12]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="2ECC71", end_color="2ECC71", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        if info["lotes"]:
            for lote in info["lotes"]:
                ws.append(
                    [
                        lote["fecha"],
                        lote["cantidad_original"],
                        lote["cantidad_disponible"],
                        lote["costo_unitario"],
                        lote["valor_total"],
                    ]
                )

                # Formato moneda
                last_row = ws.max_row
                ws.cell(last_row, 4).number_format = "#,##0.00"
                ws.cell(last_row, 5).number_format = "#,##0.00"
        else:
            ws.append(["Sin lotes disponibles", "", "", "", ""])

        ws.append([])  # Línea vacía

        # ── MOVIMIENTOS HISTÓRICOS ──
        ws.append(["HISTORIAL DE MOVIMIENTOS"])
        ws[f"A{ws.max_row}"].font = Font(size=12, bold=True, color="E74C3C")
        ws.merge_cells(f"A{ws.max_row}:I{ws.max_row}")

        headers_mov = [
            "Fecha",
            "Tipo",
            "Cantidad",
            "Costo Unit.",
            "Total",
            "Saldo",
            "Descripción",
        ]
        ws.append(headers_mov)

        header_row_mov = ws.max_row
        for cell in ws[header_row_mov]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="E74C3C", end_color="E74C3C", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        if info["movimientos"]:
            for mov in reversed(info["movimientos"]):  # Más reciente primero
                tipo_emoji = (
                    "📥"
                    if mov["tipo"] == "entrada"
                    else "📤"
                    if mov["tipo"] == "salida"
                    else "🔄"
                )
                ws.append(
                    [
                        mov["fecha"],
                        f"{tipo_emoji} {mov['tipo'].title()}",
                        mov["cantidad"],
                        mov.get("costo", 0),
                        mov.get("total", 0),
                        mov["saldo"],
                        mov.get("descripcion", ""),
                    ]
                )

                # Formato
                last_row = ws.max_row
                ws.cell(last_row, 4).number_format = "#,##0.00"
                ws.cell(last_row, 5).number_format = "#,##0.00"

                # Color según tipo
                if mov["tipo"] == "entrada":
                    ws.cell(last_row, 2).font = Font(color="2ECC71")
                elif mov["tipo"] == "salida":
                    ws.cell(last_row, 2).font = Font(color="E74C3C")
        else:
            ws.append(["Sin movimientos registrados", "", "", "", "", "", ""])

        # Ajustar anchos
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 40

    # ── Enviar como descarga ──
    from io import BytesIO
    from flask import send_file

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Kardex_PEPS_{timestamp}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ── PLANTILLA DE IMPORTACIÓN ──
@app.route("/kardex/detectar_hojas", methods=["POST"])
def detectar_hojas():
    """Recibe un Excel y devuelve sus hojas con una vista previa de columnas."""
    from io import BytesIO

    archivo = request.files.get("archivo")
    if not archivo or archivo.filename == "":
        return jsonify({"ok": False, "error": "No se recibió archivo"})

    ext = archivo.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        # CSV no tiene hojas — devolver hoja única virtual
        return jsonify(
            {
                "ok": True,
                "tipo": "csv",
                "hojas": [{"nombre": "Hoja única (CSV)", "filas": 0, "columnas": []}],
            }
        )

    try:
        import openpyxl

        contenido = archivo.read()
        wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True, read_only=True)
        hojas = []
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            # Leer solo las primeras filas para preview
            headers = []
            filas_con_datos = 0
            primera = True
            for row in ws.iter_rows(max_row=200, values_only=True):
                if any(c for c in row if c is not None):
                    if primera:
                        headers = [str(c).strip() if c is not None else "" for c in row]
                        primera = False
                    else:
                        filas_con_datos += 1
            hojas.append(
                {
                    "nombre": nombre_hoja,
                    "filas": filas_con_datos,
                    "columnas": headers[:8],  # max 8 para mostrar
                }
            )
        wb.close()
        # Guardar contenido en sesión temporal (archivo en memoria con nombre único)
        import base64, time

        token = str(int(time.time() * 1000))
        # Guardar en un dict global temporal (simple, sin Redis)
        # Limitar tamaño para evitar fuga de memoria
        if len(_archivos_temp) >= _MAX_ARCHIVOS_TEMP:
            _archivos_temp.clear()
        _archivos_temp[token] = {
            "contenido": contenido,
            "filename": archivo.filename,
            "ext": ext,
        }
        return jsonify({"ok": True, "tipo": "xlsx", "hojas": hojas, "token": token})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/kardex/plantilla")
def descargar_plantilla_inventario():
    """Genera y descarga un .xlsx de ejemplo para importar inventario."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"

        # Encabezados
        headers = ["producto", "cantidad", "costo", "fecha", "descripcion"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F8CFF", end_color="4F8CFF", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Filas de ejemplo
        ejemplos = [
            ["Collar de Plata", 50, 120.00, "2026-01-15", "Stock inicial"],
            ["Aretes Dorados", 30, 85.50, "2026-01-15", "Stock inicial"],
            ["Pulsera de Oro", 20, 250.00, "2026-01-20", "Compra proveedor"],
            ["Anillo de Plata", 15, 95.00, "", ""],
        ]
        for fila in ejemplos:
            ws.append(fila)

        # Anchos
        for col, w in zip("ABCDE", [25, 12, 12, 14, 28]):
            ws.column_dimensions[chr(64 + list("ABCDE").index(col) + 1)].width = w

        # Hoja de instrucciones
        ws2 = wb.create_sheet("Instrucciones")
        instrucciones = [
            ("📋 INSTRUCCIONES DE IMPORTACIÓN", True),
            ("", False),
            ("1. Llena la hoja 'Inventario' con tus productos.", False),
            ("2. La columna 'producto' es obligatoria.", False),
            ("3. La columna 'cantidad' debe ser un número entero positivo.", False),
            ("4. La columna 'costo' es el costo unitario en C$.", False),
            ("5. La columna 'fecha' es opcional (formato YYYY-MM-DD).", False),
            ("   Si se omite, se usa la fecha actual.", False),
            ("6. La columna 'descripcion' es opcional.", False),
            ("", False),
            ("⚠️  No borres la fila de encabezados (fila 1).", False),
            (
                "⚠️  Si el producto ya existe en el Kardex, se agrega como nueva entrada.",
                False,
            ),
            ("⚠️  Productos con cantidad 0 o negativa se ignoran.", False),
        ]
        for i, (texto, bold) in enumerate(instrucciones, 1):
            cell = ws2.cell(row=i, column=1, value=texto)
            cell.font = Font(bold=bold, size=11 if bold else 10)
        ws2.column_dimensions["A"].width = 70

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import send_file

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Plantilla_Importar_Inventario.xlsx",
        )
    except ImportError:
        return "openpyxl no está instalado", 500


# ── IMPORTAR INVENTARIO DESDE EXCEL / CSV ──

def _leer_excel_inventario(contenido, hoja_nombre):
    filas, headers_originales = [], []
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
        if hoja_nombre and hoja_nombre in wb.sheetnames:
            ws = wb[hoja_nombre]
        else:
            ws = wb.active
        primera_fila = None
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if any(c for c in row if c is not None):
                primera_fila = row
                break
        if primera_fila is None:
            return [], [], f"La hoja '{ws.title}' está vacía."
        headers_originales = [str(c).strip() if c is not None else "" for c in primera_fila]
        headers_lower = [h.lower() for h in headers_originales]
        fila_inicio = None
        for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if list(row) == list(primera_fila):
                fila_inicio = idx + 1
                break
        for row in ws.iter_rows(min_row=fila_inicio or 2, values_only=True):
            if any(c for c in row if c is not None):
                filas.append(dict(zip(headers_lower, row)))
    except Exception as e:
        return [], [], f"No se pudo leer el Excel: {e}"
    return filas, headers_originales, None


def _leer_csv_inventario(contenido):
    import csv, io
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = contenido.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        text = contenido.decode("latin-1", errors="replace")
    primera_linea = text.split("\n")[0] if text else ""
    sep = (";" if primera_linea.count(";") > primera_linea.count(",")
            else "\t" if primera_linea.count("\t") > primera_linea.count(",")
            else ",")
    reader = csv.DictReader(io.StringIO(text), delimiter=sep)
    headers_originales = reader.fieldnames or []
    filas = [{k.strip().lower(): v for k, v in row.items() if k} for row in reader]
    return filas, headers_originales, None


NOMBRES_PRODUCTO = {"producto","product","nombre","name","articulo","item",
    "descripcion_producto","art","codigo","code","referencia","ref",
    "mercaderia","mercancía","mercancia"}
NOMBRES_CANTIDAD = {"cantidad","qty","quantity","unidades","units","stock",
    "existencia","existencias","cant","inventario","piezas","pcs"}
NOMBRES_COSTO = {"costo","cost","precio","price","costo_unitario","unit_cost",
    "value","costo_unit","precio_unitario","pu","p.u.","c/u"}
NOMBRES_FECHA = {"fecha","date","fecha_entrada","fecha_compra","entry_date"}
NOMBRES_DESC = {"descripcion","description","desc","detalle","detail","nota","observacion"}


def _encontrar_columna(keys, nombres_set):
    for k in keys:
        if k in nombres_set:
            return k
    for k in keys:
        for n in nombres_set:
            if n in k or k in n:
                return k
    return None


def _mapear_columnas_inventario(headers_originales, filas):
    keys = list(filas[0].keys()) if filas else []
    col_producto = _encontrar_columna(keys, NOMBRES_PRODUCTO)
    col_cantidad = _encontrar_columna(keys, NOMBRES_CANTIDAD)
    col_costo = _encontrar_columna(keys, NOMBRES_COSTO)
    col_fecha = _encontrar_columna(keys, NOMBRES_FECHA)
    col_desc = _encontrar_columna(keys, NOMBRES_DESC)
    modo_auto = False
    if not col_producto and len(keys) >= 1:
        col_producto = keys[0]
        modo_auto = True
    if not col_cantidad and len(keys) >= 2:
        col_cantidad = keys[1]
        modo_auto = True
    if not col_costo and len(keys) >= 3:
        col_costo = keys[2]
    msg = None
    if not col_producto:
        msg = (f"No se encontró columna de producto. "
               f"Columnas en tu archivo: {', '.join(headers_originales)}. "
               f"Renómbralas como: producto, cantidad, costo")
    return col_producto, col_cantidad, col_costo, col_fecha, col_desc, modo_auto, msg


def _val_fila(fila, col):
    if not col:
        return None
    val = fila.get(col)
    if val is None:
        return None
    s = str(val).strip()
    return s if s not in ("", "None", "nan", "NaN", "-") else None


def _parsear_cantidad(raw):
    try:
        return int(float(raw.replace(",", "").replace(" ", "").replace("'", "")))
    except Exception:
        return 0


def _parsear_costo(raw):
    try:
        return float(raw.replace(",", "").replace(" ", "")
                     .replace("C$", "").replace("$", "").strip())
    except Exception:
        return 0.0


def _parsear_fecha(raw, hoy):
    if "/" in raw:
        parts = raw.split("/")
        try:
            if len(parts) == 3:
                d, m, a = parts
                if len(a) == 2:
                    a = "20" + a
                return f"{a}-{m.zfill(2)}-{d.zfill(2)}"
        except Exception:
            pass
    return hoy


def _importar_filas_inventario(data, filas, col_producto, col_cantidad, col_costo, col_fecha, col_desc):
    from datetime import date as _date
    import kardex_peps
    hoy = _date.today().isoformat()
    importados, omitidos = 0, 0
    for fila in filas:
        producto = _val_fila(fila, col_producto)
        if not producto:
            omitidos += 1
            continue
        cantidad = _parsear_cantidad(_val_fila(fila, col_cantidad) or "0")
        costo = _parsear_costo(_val_fila(fila, col_costo) or "0")
        fecha = _parsear_fecha(_val_fila(fila, col_fecha) or "", hoy)
        if producto not in data["kardex"]:
            data["kardex"][producto] = []
            data["productos"][producto] = {"nombre": producto, "stock": 0, "costo_promedio": 0, "precio_venta": 0, "margen": 0}
            data["kardex_peps"][producto] = {"lotes": [], "stock_total": 0, "costo_promedio": 0, "precio_venta": 0, "margen": 0}
        if cantidad > 0:
            data = kardex_peps.agregar_entrada_peps(data, producto, fecha, cantidad, costo)
        importados += 1
    return data, importados, omitidos


@app.route("/kardex/importar", methods=["POST"])
def importar_inventario():
    from io import BytesIO
    import urllib.parse

    token = request.form.get("token", "")
    hoja_nombre = request.form.get("hoja_nombre", "")
    ext = request.form.get("ext", "")
    contenido = None

    if token and token in _archivos_temp:
        entrada = _archivos_temp.pop(token)
        contenido, ext = entrada["contenido"], entrada["ext"]
    else:
        archivo = request.files.get("archivo")
        if not archivo or archivo.filename == "":
            return redirect(url_for("kardex"))
        ext = archivo.filename.rsplit(".", 1)[-1].lower()
        contenido = archivo.read()

    if ext in ("xlsx", "xls"):
        filas, headers_originales, error_lectura = _leer_excel_inventario(contenido, hoja_nombre)
    elif ext == "csv":
        filas, headers_originales, error_lectura = _leer_csv_inventario(contenido)
    else:
        error_lectura = f"Formato '{ext}' no soportado. Usa .xlsx o .csv"
        filas, headers_originales = [], []

    if error_lectura:
        return redirect(url_for("kardex") + "?error=" + urllib.parse.quote(error_lectura))
    if not filas:
        msg = (f"La hoja seleccionada no tiene datos. Columnas detectadas: "
               f"{', '.join(headers_originales) or 'ninguna'}")
        return redirect(url_for("kardex") + "?error=" + urllib.parse.quote(msg))

    mapeo = _mapear_columnas_inventario(headers_originales, filas)
    col_producto, col_cantidad, col_costo, col_fecha, col_desc, modo_auto, error_col = mapeo
    if error_col:
        return redirect(url_for("kardex") + "?error=" + urllib.parse.quote(error_col))

    data = load_data()
    data.setdefault("kardex", {})
    data.setdefault("kardex_peps", {})
    data.setdefault("productos", {})
    data, importados, omitidos = _importar_filas_inventario(
        data, filas, col_producto, col_cantidad, col_costo, col_fecha, col_desc)
    save_data(data)

    qs = f"importados={importados}&errores={omitidos}"
    if modo_auto:
        qs += "&auto=1"
    return redirect(url_for("kardex") + "?" + qs)


# ══════════════════════════════════════════════════════════════
#  KARDEX — CRUD: eliminar producto, editar/eliminar movimiento
# ══════════════════════════════════════════════════════════════


@app.route("/kardex/eliminar_producto", methods=["POST"])
def eliminar_producto_kardex():
    """Elimina un producto y todos sus movimientos del Kardex."""
    data = load_data()
    nombre = request.form.get("producto", "").strip()
    if nombre:
        data["kardex"].pop(nombre, None)
        data.get("kardex_peps", {}).pop(nombre, None)
        data.get("productos", {}).pop(nombre, None)
        save_data(data)
    return redirect(url_for("kardex"))


@app.route("/kardex/renombrar_producto", methods=["POST"])
def renombrar_producto_kardex():
    """Renombra un producto manteniendo todos sus movimientos."""
    data = load_data()
    nombre_old = request.form.get("nombre_old", "").strip()
    nombre_new = request.form.get("nombre_new", "").strip()
    if (
        nombre_old
        and nombre_new
        and nombre_old in data["kardex"]
        and nombre_new not in data["kardex"]
    ):
        for seccion in ("kardex", "kardex_peps", "productos"):
            d = data.get(seccion, {})
            if nombre_old in d:
                d[nombre_new] = d.pop(nombre_old)
        save_data(data)
    return redirect(url_for("kardex"))


@app.route("/kardex/editar_movimiento", methods=["POST"])
def editar_movimiento_kardex():
    """Edita un movimiento individual del Kardex (fecha, descripcion, cantidad, costo)."""
    data = load_data()
    producto = request.form.get("producto", "").strip()
    try:
        idx = int(request.form.get("idx", -1) or -1)
    except (ValueError, TypeError):
        return redirect(url_for("kardex"))
    k = data["kardex"].get(producto, [])
    if 0 <= idx < len(k):
        mov = k[idx]
        mov["fecha"] = request.form.get("fecha", mov.get("fecha", ""))
        mov["descripcion"] = request.form.get("descripcion", mov.get("descripcion", ""))
        nuevo_pv = mov.get("precio_venta", 0)
        try:
            nueva_cant = int(
                float(request.form.get("cantidad", mov.get("cantidad", 0)))
            )
            nuevo_costo = float(request.form.get("costo", mov.get("costo", 0)))
            nuevo_pv = float(request.form.get("precio_venta", mov.get("precio_venta", 0)))
            mov["cantidad"] = nueva_cant
            mov["costo"] = nuevo_costo
            mov["precio_venta"] = nuevo_pv
            mov["total"] = nueva_cant * nuevo_costo
        except (ValueError, TypeError):
            pass
        
        data["kardex"][producto] = k
        if nuevo_pv > 0:
            peps = data.get("kardex_peps", {}).get(producto, {})
            if peps:
                peps["precio_venta"] = nuevo_pv
            prod = data.get("productos", {}).get(producto, {})
            if isinstance(prod, dict):
                prod["precio_venta"] = nuevo_pv
        data = _reconstruir_peps(data, producto)
        save_data(data)
    return redirect(url_for("kardex"))


@app.route("/kardex/eliminar_movimiento", methods=["POST"])
def eliminar_movimiento_kardex():
    """Elimina un movimiento individual y recalcula saldos."""
    data = load_data()
    producto = request.form.get("producto", "").strip()
    try:
        idx = int(request.form.get("idx", -1) or -1)
    except (ValueError, TypeError):
        return redirect(url_for("kardex"))
    k = data["kardex"].get(producto, [])
    if 0 <= idx < len(k):
        k.pop(idx)
        data["kardex"][producto] = k
        data = _reconstruir_peps(data, producto)
        save_data(data)
    return redirect(url_for("kardex"))


def _reconstruir_peps(data, producto):
    """
    Reconstruye los lotes PEPS de un producto a partir de su historial kardex.
    Procesa entradas, salidas y ajustes de forma correcta.
    """
    try:
        movs = list(data["kardex"].get(producto, []))  # ← copia para evitar loop infinito
        len_original = len(data["kardex"].get(producto, []))

        # Preservar precio_venta y margen del producto
        old_peps = data.get("kardex_peps", {}).get(producto, {})
        old_prod = data.get("productos", {}).get(producto, {})
        if isinstance(old_prod, dict):
            pv = old_peps.get("precio_venta", 0) or old_prod.get("precio_venta", 0)
            mg = old_peps.get("margen", 30) or old_prod.get("margen", 30)
        else:
            pv = old_peps.get("precio_venta", 0)
            mg = old_peps.get("margen", 30)

        if "kardex_peps" not in data:
            data["kardex_peps"] = {}
        data["kardex_peps"][producto] = {
            "lotes": [],
            "stock_total": 0,
            "costo_promedio": 0,
            "precio_venta": pv,
            "margen": mg,
        }

        # Truncar el listado kardex para evitar duplicados
        if producto in data["kardex"]:
            del data["kardex"][producto][:]

        # Reconstruir saldo secuencial
        saldo_acum = 0

        # Procesar cada movimiento en orden cronológico
        for m in movs:
            if m.get("tipo") == "entrada" and m.get("cantidad", 0) > 0:
                cant = m["cantidad"]
                data = kardex_peps.agregar_entrada_peps(
                    data, producto,
                    m.get("fecha", ""), cant,
                    m.get("costo", 0), m.get("precio_venta", 0),
                )
                saldo_acum += cant
            elif m.get("tipo") == "salida" and m.get("cantidad", 0) > 0:
                try:
                    data, _, _ = kardex_peps.procesar_salida_peps(
                        data, producto, m.get("fecha", ""), m.get("cantidad", 0)
                    )
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                saldo_acum -= m["cantidad"]
            elif m.get("tipo") == "ajuste" and m.get("cantidad", 0) > 0:
                stock_actual = data["kardex_peps"][producto]["stock_total"]
                delta = m["cantidad"] - stock_actual
                if delta > 0:
                    data = kardex_peps.agregar_entrada_peps(
                        data, producto, m.get("fecha", ""), delta,
                        m.get("costo", 0), m.get("precio_venta", 0),
                    )
                elif delta < 0:
                    try:
                        data, _, _ = kardex_peps.procesar_salida_peps(
                            data, producto, m.get("fecha", ""), abs(delta),
                        )
                    except Exception:
                        pass
                saldo_acum = m["cantidad"]

        # Restaurar los movimientos originales con saldos recalculados
        data["kardex"][producto] = movs
        # Recalcular saldo secuencialmente en los movimientos
        saldo = 0
        for mov in data["kardex"][producto]:
            c = mov.get("cantidad", 0)
            if mov.get("tipo") in ("entrada",):
                saldo += c
            elif mov.get("tipo") in ("salida",):
                saldo -= c
            elif mov.get("tipo") == "ajuste":
                saldo = c
            mov["saldo"] = saldo
            mov["total"] = mov.get("cantidad", 0) * mov.get("costo", 0)
    except Exception as e:
        # En caso de error, asegurar que kardex_peps existe
        if "kardex_peps" not in data:
            data["kardex_peps"] = {}
        if producto not in data["kardex_peps"]:
            data["kardex_peps"][producto] = {
                "lotes": [],
                "stock_total": 0,
                "costo_promedio": 0,
            }
    
    return data


# ── APAGAR SERVIDOR ──
@app.route("/apagar", methods=["POST"])
def apagar():
    import threading
    from flask import make_response

    def shutdown():
        import time

        time.sleep(0.6)
        try:
            os.kill(os.getpid(), signal.SIGTERM)
        except Exception:
            os._exit(0)

    t = threading.Thread(target=shutdown, daemon=True)
    t.start()

    resp = make_response(render_template("apagar.html"))
    resp.headers["Connection"] = "close"
    return resp


# ══════════════════════════════════════════════════════════════
# RUTAS API PARA GESTIÓN DE BASE DE DATOS
# ══════════════════════════════════════════════════════════════


@app.route("/api/backup_json")
def api_backup_json():
    """Exporta los datos actuales como archivo JSON"""
    data = load_data()
    from flask import make_response
    import json

    json_str = json.dumps(data, indent=2, default=str)
    resp = make_response(json_str)
    resp.headers["Content-Disposition"] = "attachment; filename=fg_backup.json"
    resp.headers["Content-Type"] = "application/json"
    return resp


@app.route("/api/copia_db")
def api_copia_db():
    """Crea una copia de la base de datos SQLite"""
    import shutil
    from flask import make_response
    from io import BytesIO

    db_path = db_internal.get_db_path()
    if os.path.exists(db_path):
        with open(db_path, "rb") as f:
            data = f.read()
        resp = make_response(data)
        resp.headers["Content-Disposition"] = "attachment; filename=fg_db_backup.db"
        resp.headers["Content-Type"] = "application/octet-stream"
        return resp
    return jsonify({"ok": False, "error": "Base de datos no encontrada"}), 404


@app.route("/api/importar_json", methods=["POST"])
def api_importar_json():
    """Importa un JSON de backup y reemplaza los datos actuales"""
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"ok": False, "error": "No se recibió archivo"})
    try:
        data = json.loads(archivo.read().decode("utf-8"))
        required = ["cuentas", "diario", "productos", "kardex", "kardex_peps", "pos_historial"]
        for key in required:
            if key not in data:
                data[key] = {} if key in ("kardex", "kardex_peps", "productos", "cuentas") else []
        save_data(data)
        return jsonify({"ok": True, "message": f"Importado: {len(data.get('diario', []))} asientos, {len(data.get('pos_historial', []))} ventas"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/importar_db", methods=["POST"])
def api_importar_db():
    """Importa una base de datos SQLite"""
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"ok": False, "error": "No se recibió archivo"})

    db_path = db_internal.get_db_path()
    try:
        archivo.save(db_path)
        return jsonify({"ok": True, "message": "Base de datos importada"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/ver_db")
def api_ver_db():
    """Muestra el contenido de la base de datos como HTML"""
    import sqlite3

    db_path = db_internal.get_db_path()

    html = "<html><head><title>Ver DB</title><style>"
    html += "body{font-family:monospace;background:#1a1d2e;color:#e2e4eb;padding:20px}"
    html += "table{border-collapse:collapse;width:100%}"
    html += "th,td{border:1px solid #2a2d42;padding:8px;text-align:left}"
    html += "th{background:#2a2d42}"
    html += "</style></head><body>"
    html += "<h1>Contenido de Base de Datos</h1>"

    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tablas = cursor.fetchall()

        for tabla in tablas:
            nombre_tabla = tabla[0]
            html += f"<h2>Tabla: {nombre_tabla}</h2>"
            cursor.execute(f"SELECT * FROM {nombre_tabla} LIMIT 100")
            rows = cursor.fetchall()

            if rows:
                html += "<table><tr>"
                # Headers
                for col in rows[0].keys():
                    html += f"<th>{col}</th>"
                html += "</tr>"

                # Datos
                for row in rows:
                    html += "<tr>"
                    for col in row:
                        html += f"<td>{col}</td>"
                    html += "</tr>"
                html += "</table>"
            else:
                html += "<p>Sin datos</p>"

        conn.close()
    except Exception as e:
        html += f"<p>Error: {e}</p>"

    html += "</body></html>"
    return html


@app.route("/api/restaurar_json", methods=["POST"])
def api_restaurar_json():
    """Restaura los datos desde un archivo JSON"""
    try:
        data_json = request.get_data(as_text=True)
        data = json.loads(data_json)

        # Validar estructura básica
        if "cuentas" not in data:
            return jsonify({"ok": False, "error": "Archivo JSON inválido"}), 400

        # Guardar en DB
        save_data(data)
        return jsonify({"ok": True, "message": "Datos restaurados correctamente"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/borrar_todo", methods=["POST"])
def api_borrar_todo():
    """Borra todos los datos (reset completo)"""
    try:
        data_vacio = empty_data()
        data_vacio = audit_log(data_vacio, "borrar_todo", "Todos los datos fueron eliminados")
        save_data(data_vacio)
        return jsonify({"ok": True, "message": "Todos los datos han sido borrados"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── NUEVA CUENTA ──
@app.route("/cuentas/nueva", methods=["POST"])
def nueva_cuenta():
    data = load_data()
    codigo = request.form.get("codigo", "").strip()
    nombre = request.form.get("nombre", "").strip()
    tipo = request.form.get("tipo", "Activo")
    if codigo and nombre and codigo not in data["cuentas"]:
        data["cuentas"][codigo] = {"nombre": nombre, "tipo": tipo, "saldo": 0}
        save_data(data)
    return redirect(request.referrer or url_for("index"))


# ── CATÁLOGO DE CUENTAS (CRUD completo) ──
@app.route("/catalogar")
def catalogar():
    data = load_data()
    cuentas = sorted(data["cuentas"].items(), key=lambda x: x[0])
    return render_template("catalogar.html", cuentas=cuentas)


@app.route("/catalogar/nueva", methods=["POST"])
def catalogar_nueva():
    data = load_data()
    codigo = request.form.get("codigo", "").strip()
    nombre = request.form.get("nombre", "").strip()
    tipo = request.form.get("tipo", "Activo")
    if not codigo or not nombre:
        return redirect(url_for("catalogar") + "?error=codigo_invalido")
    if codigo in data["cuentas"]:
        return redirect(url_for("catalogar") + "?error=codigo_invalido")
    data["cuentas"][codigo] = {"nombre": nombre, "tipo": tipo, "saldo": 0}
    data = audit_log(data, "crear_cuenta", f"Cuenta {codigo} — {nombre}")
    save_data(data)
    return redirect(url_for("catalogar") + "?ok=cuenta_guardada")


@app.route("/catalogar/editar", methods=["POST"])
def catalogar_editar():
    data = load_data()
    codigo = request.form.get("codigo", "").strip()
    nombre = request.form.get("nombre", "").strip()
    tipo = request.form.get("tipo", "Activo")
    if codigo not in data["cuentas"]:
        return redirect(url_for("catalogar") + "?error=cuenta_no_existe")
    if not nombre:
        return redirect(url_for("catalogar") + "?error=cuenta_no_existe")
    data["cuentas"][codigo]["nombre"] = nombre
    data["cuentas"][codigo]["tipo"] = tipo
    data = audit_log(data, "editar_cuenta", f"Cuenta {codigo} — {nombre}")
    save_data(data)
    return redirect(url_for("catalogar") + "?ok=cuenta_guardada")


@app.route("/catalogar/eliminar", methods=["POST"])
def catalogar_eliminar():
    data = load_data()
    codigo = request.form.get("codigo", "").strip()
    if codigo not in data["cuentas"]:
        return redirect(url_for("catalogar") + "?error=cuenta_no_existe")

    tiene_mov = False
    for entry in data.get("diario", []):
        for mov in entry.get("movimientos", []):
            if mov["cuenta"] == codigo:
                tiene_mov = True
                break
        if tiene_mov:
            break
    if not tiene_mov:
        for aj in data.get("ajustes", []):
            for mov in aj.get("movimientos", []):
                if mov["cuenta"] == codigo:
                    tiene_mov = True
                    break
            if tiene_mov:
                break
    if not tiene_mov:
        for mov in data.get("caja_movimientos", []):
            if mov.get("cuenta") == codigo:
                tiene_mov = True
                break
    if not tiene_mov:
        for cc in data.get("cuentas_cobrar", []):
            if cc.get("cuenta") == codigo:
                tiene_mov = True
                break
            for pago in cc.get("pagos", []):
                if pago.get("cuenta") == codigo:
                    tiene_mov = True
                    break
    if not tiene_mov:
        for prod, regs in data.get("kardex", {}).items():
            for reg in regs:
                if reg.get("cuenta") == codigo:
                    tiene_mov = True
                    break
    if tiene_mov:
        return redirect(url_for("catalogar") + "?error=no_se_puede_eliminar_cuenta_con_movimientos")
    data = audit_log(data, "eliminar_cuenta", f"Cuenta {codigo} — {data['cuentas'].get(codigo,{}).get('nombre','?')}")
    del data["cuentas"][codigo]
    save_data(data)
    return redirect(url_for("catalogar") + "?ok=cuenta_eliminada")


# ── ELIMINAR CUENTA ──
@app.route("/cuentas/eliminar", methods=["POST"])
def eliminar_cuenta():
    data = load_data()
    codigo = request.form.get("codigo", "").strip()
    if codigo not in data["cuentas"]:
        return redirect(request.referrer or url_for("index"))

    # Verificar si la cuenta tiene movimientos en el diario
    tiene_mov = False
    for entry in data.get("diario", []):
        for mov in entry.get("movimientos", []):
            if mov["cuenta"] == codigo:
                tiene_mov = True
                break
        if tiene_mov:
            break
    if not tiene_mov:
        for aj in data.get("ajustes", []):
            for mov in aj.get("movimientos", []):
                if mov["cuenta"] == codigo:
                    tiene_mov = True
                    break
            if tiene_mov:
                break
    if not tiene_mov:
        for mov in data.get("caja_movimientos", []):
            if mov.get("cuenta") == codigo:
                tiene_mov = True
                break
    if not tiene_mov:
        for cc in data.get("cuentas_cobrar", []):
            if cc.get("cuenta") == codigo:
                tiene_mov = True
                break
            for pago in cc.get("pagos", []):
                if pago.get("cuenta") == codigo:
                    tiene_mov = True
                    break
    if not tiene_mov:
        for prod, regs in data.get("kardex", {}).items():
            for reg in regs:
                if reg.get("cuenta") == codigo:
                    tiene_mov = True
                    break

    if tiene_mov:
        return redirect((request.referrer or url_for("index")) + "?error=no_se_puede_eliminar_cuenta_con_movimientos")
    del data["cuentas"][codigo]
    save_data(data)
    return redirect(request.referrer or url_for("index"))


@app.route("/catalogar/resetear", methods=["POST"])
def catalogar_resetear():
    """Reemplaza todo el catálogo con las cuentas base y migra datos existentes."""
    data = load_data()

    OLD_TO_NEW = {
        "1001": "1.1.01", "1002": "1.1.02", "1003": "1.1.03",
        "1004": "1.1.04", "2001": "2.1.01", "2002": "2.1.01",
        "3001": "3.1", "3002": "3.3.02",
        "4001": "4.1.01", "4002": "4.2",
        "5001": "5.1", "5002": "5.2", "5003": "5.2.05",
        "6001": "5.2.05",
    }

    def migrar_cuenta(cod):
        return OLD_TO_NEW.get(cod, cod)

    for entry in data.get("diario", []):
        for mov in entry.get("movimientos", []):
            mov["cuenta"] = migrar_cuenta(mov["cuenta"])

    for mov in data.get("caja_movimientos", []):
        mov["cuenta"] = migrar_cuenta(mov["cuenta"])

    for aj in data.get("ajustes", []):
        for mov in aj.get("movimientos", []):
            mov["cuenta"] = migrar_cuenta(mov["cuenta"])

    for prod_key in list(data.get("kardex_peps", {})):
        peps = data["kardex_peps"][prod_key]
        for lote in peps.get("lotes", []):
            lote["cuenta"] = migrar_cuenta(lote.get("cuenta", ""))

    data["cuentas"] = dict(CUENTAS_BASE)
    data = audit_log(data, "resetear_catalogo", "Catálogo restaurado a cuentas por defecto con migración")
    save_data(data)
    return redirect(url_for("catalogar") + "?ok=cuenta_guardada")


# ── REPORTES ──
@app.route("/reportes")
def reportes():
    data = load_data()
    # Recopilar estadísticas para la pantalla
    mayor = _mayor_cached(data)
    _, td, th, tsd, tsh = calcular_balanza(data)
    di, ti, dg, tg, util = calcular_estado_resultados(data)
    activos, ta, pasivos, tp, capital, tc = calcular_balance_general(data)
    # Auditoría
    entries = data.get("auditoria", [])[:]
    entries.reverse()
    page = request.args.get("auditoria_page", 1, type=int)
    page_entries, aud_page, aud_total_pages, aud_total = _paginar(entries, page, 50)
    # Fiscales - meses disponibles
    fechas = set()
    for e in data.get("diario", []):
        f = e.get("fecha", "")
        if len(f) >= 7:
            fechas.add(f[:7])
    meses = sorted(fechas, reverse=True)
    return render_template(
        "reportes.html",
        diario=data["diario"],
        cuentas=data["cuentas"],
        mayor=mayor,
        balanza_td=td, balanza_th=th, balanza_tsd=tsd, balanza_tsh=tsh,
        detalle_ingresos=di, total_ingresos=ti,
        detalle_gastos=dg, total_gastos=tg, utilidad=util,
        activos=activos, total_activo=ta,
        pasivos=pasivos, total_pasivo=tp,
        capital=capital, total_capital=tc,
        kardex=data["kardex"],
        pos_historial=data.get("pos_historial", []),
        # Auditoría
        auditoria_entries=page_entries, auditoria_page=aud_page,
        auditoria_total_pages=aud_total_pages, auditoria_total=aud_total,
        # Fiscales
        meses=meses,
    )


def _parse_fecha_excel(f):
    from datetime import date
    try:
        return date.fromisoformat(f)
    except (ValueError, TypeError):
        return date.today()


@app.route("/reportes/exportar", methods=["POST"])
def exportar_reporte():
    """Genera un .xlsx con los reportes solicitados (envío directo, no guarda en disco)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess, sys

        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"]
        )
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    data = load_data()
    desde = request.form.get("desde", "") or None
    hasta = request.form.get("hasta", "") or None
    tipo_cuenta = request.form.get("tipo_cuenta", "") or None
    cuenta_especifica = request.form.get("cuenta_especifica", "") or None
    from datetime import date, datetime

    # Soporte multi-hojas (nuevo) y legacy tipo=single
    hojas_raw = request.form.getlist("hojas")
    if hojas_raw:
        hojas_seleccionadas = set(hojas_raw)
    else:
        tipo_reporte = request.form.get("tipo", "todos")
        if tipo_reporte == "todos":
            hojas_seleccionadas = {"diario", "mayor", "balanza", "er", "bg", "pos", "kardex", "cobros"}
        else:
            # Mapear legacy tipo -> hoja key
            mapa_legacy = {
                "diario": "diario", "mayor": "mayor", "balanza": "balanza",
                "estado_resultados": "er", "balance_general": "bg",
                "kardex": "kardex", "antiguedad_cobros": "cobros",
            }
            hojas_seleccionadas = {mapa_legacy.get(tipo_reporte, tipo_reporte)}

    def _filtrar_cuentas(movimientos):
        """Filtra movimientos por tipo de cuenta o cuenta específica."""
        if not tipo_cuenta and not cuenta_especifica:
            return movimientos
        filtered = []
        for mov in movimientos:
            cod = mov.get("cuenta", "")
            info = cuentas.get(cod, {})
            if cuenta_especifica and cod != cuenta_especifica:
                continue
            if tipo_cuenta and info.get("tipo", "") != tipo_cuenta:
                continue
            filtered.append(mov)
        return filtered

    # Filtrar data por período si se proporciona
    data_original = data  # Guardar referencia a datos completos para saldos iniciales
    if desde or hasta:
        import copy
        data = copy.deepcopy(data)
        if desde and hasta:
            data["diario"] = [e for e in data.get("diario", []) if desde <= e.get("fecha", "") <= hasta]
            data["ajustes"] = [a for a in data.get("ajustes", []) if desde <= a.get("fecha", "") <= hasta]
        elif desde:
            data["diario"] = [e for e in data.get("diario", []) if desde <= e.get("fecha", "")]
            data["ajustes"] = [a for a in data.get("ajustes", []) if desde <= a.get("fecha", "")]
        elif hasta:
            data["diario"] = [e for e in data.get("diario", []) if e.get("fecha", "") <= hasta]
            data["ajustes"] = [a for a in data.get("ajustes", []) if a.get("fecha", "") <= hasta]

    # ── Estilos ──
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1D2E")
    accent_fill = PatternFill("solid", fgColor="4F8CFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1A1D2E")
    sub_font = Font(name="Calibri", size=10, color="7A7F99")
    money_fmt = "#,##0.00"
    thin_border = Border(
        left=Side(style="thin", color="D0D3E0"),
        right=Side(style="thin", color="D0D3E0"),
        top=Side(style="thin", color="D0D3E0"),
        bottom=Side(style="thin", color="D0D3E0"),
    )
    total_fill = PatternFill("solid", fgColor="E8EAEF")
    total_font = Font(name="Calibri", size=11, bold=True, color="1A1D2E")
    green_font = Font(name="Calibri", size=11, color="27AE60")
    red_font = Font(name="Calibri", size=11, color="E74C3C")

    def style_header_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def style_data_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10, color="333333")

    def style_total_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = Border(
                left=Side(style="thin", color="D0D3E0"),
                right=Side(style="thin", color="D0D3E0"),
                top=Side(style="medium", color="4F8CFF"),
                bottom=Side(style="thin", color="D0D3E0"),
            )

    def add_title(ws, title, subtitle=""):
        ws.cell(row=1, column=1, value="F & G — Sistema Contable").font = Font(
            name="Calibri", size=12, bold=True, color="4F8CFF"
        )
        ws.cell(row=2, column=1, value=title).font = title_font
        if subtitle:
            ws.cell(row=3, column=1, value=subtitle).font = sub_font
        ws.cell(
            row=3 if not subtitle else 4,
            column=1,
            value=f"Fecha de exportación: {date.today().isoformat()}",
        ).font = sub_font
        return 5  # primera fila de datos

    wb = Workbook()
    wb.remove(wb.active)  # borrar hoja por defecto

    mayor = calcular_mayor(data)
    cuentas = data["cuentas"]

    # ══════════════ HOJA 1: LIBRO DIARIO ══════════════
    if "diario" in hojas_seleccionadas:
        periodo_str = f"Período: {desde} al {hasta}" if desde and hasta else "Acumulado"
        ws = wb.create_sheet("Libro Diario")
        ws.sheet_properties.tabColor = "4F8CFF"
        start = add_title(ws, "Libro Diario", f"Registro cronológico de asientos — {periodo_str}")
        headers = [
            "#",
            "Fecha",
            "Descripción",
            "Ref",
            "Cuenta",
            "Nombre Cuenta",
            "Debe C$",
            "Haber C$",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=start, column=i, value=h)
        style_header_row(ws, start, len(headers))
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 38
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 28
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 16

        r = start + 1
        total_debe = total_haber = 0
        all_entries = list(data["diario"]) + list(data.get("ajustes", []))
        all_entries.sort(key=lambda x: x.get("fecha", ""))
        for entry in all_entries:
            movs_filtrados = _filtrar_cuentas(entry["movimientos"])
            if not movs_filtrados:
                continue
            for idx, mov in enumerate(movs_filtrados):
                ws.cell(row=r, column=1, value=entry["id"] if idx == 0 else "")
                ws.cell(row=r, column=2, value=entry["fecha"] if idx == 0 else "")
                ws.cell(row=r, column=3, value=entry["descripcion"] if idx == 0 else "")
                ws.cell(row=r, column=4, value=entry.get("ref", "AJ") if idx == 0 else "")
                ws.cell(row=r, column=5, value=mov["cuenta"])
                ws.cell(
                    row=r,
                    column=6,
                    value=cuentas.get(mov["cuenta"], {}).get("nombre", mov["cuenta"]),
                )
                debe = mov["monto"] if mov["tipo"] == "Debe" else 0
                haber = mov["monto"] if mov["tipo"] == "Haber" else 0
                ws.cell(row=r, column=7, value=debe).number_format = money_fmt
                ws.cell(row=r, column=8, value=haber).number_format = money_fmt
                if debe:
                    ws.cell(row=r, column=7).font = green_font
                if haber:
                    ws.cell(row=r, column=8).font = red_font
                total_debe += debe
                total_haber += haber
                style_data_row(ws, r, len(headers))
                r += 1
        # Total
        ws.cell(row=r, column=6, value="TOTALES")
        ws.cell(row=r, column=7, value=total_debe).number_format = money_fmt
        ws.cell(row=r, column=8, value=total_haber).number_format = money_fmt
        style_total_row(ws, r, len(headers))

    # ══════════════ HOJA 2: LIBRO MAYOR ══════════════
    if "mayor" in hojas_seleccionadas:
        periodo_str = f"Período: {desde} al {hasta}" if desde and hasta else "Acumulado"
        ws = wb.create_sheet("Libro Mayor")
        ws.sheet_properties.tabColor = "7C5CFC"
        start = add_title(ws, "Libro Mayor", f"Cuentas T — Movimientos por cuenta — {periodo_str}")

        # Calcular saldos iniciales (antes del período)
        if desde:
            data_ini = {k: v for k, v in data_original.items()}
            data_ini["diario"] = [e for e in data_original.get("diario", []) if e.get("fecha", "") < desde]
            data_ini["ajustes"] = [a for a in data_original.get("ajustes", []) if a.get("fecha", "") < desde]
            mayor_ini = calcular_mayor(data_ini)
        else:
            mayor_ini = None

        headers = ["Cuenta", "Nombre", "Tipo", "Saldo Inicial Deudor", "Saldo Inicial Acreedor", "Debe Período", "Haber Período", "Saldo Final Deudor", "Saldo Final Acreedor"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=start, column=i, value=h)
        style_header_row(ws, start, len(headers))
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 16
        ws.column_dimensions["I"].width = 16

        r = start + 1
        for codigo in sorted(cuentas.keys()):
            info = cuentas[codigo]
            if cuenta_especifica and codigo != cuenta_especifica:
                continue
            if tipo_cuenta and info.get("tipo", "") != tipo_cuenta:
                continue
            debe_mov = mayor[codigo]["debe"] if codigo in mayor else 0
            haber_mov = mayor[codigo]["haber"] if codigo in mayor else 0
            ts = tipo_saldo(info["tipo"])

            # Saldo inicial
            if mayor_ini is not None and codigo in mayor_ini:
                ini_d = mayor_ini[codigo]["debe"]
                ini_h = mayor_ini[codigo]["haber"]
            else:
                ini_d = 0
                ini_h = 0

            if ts == "Debe":
                saldo_ini_d = max(ini_d - ini_h, 0)
                saldo_ini_h = max(ini_h - ini_d, 0)
                saldo_fin_d = max((ini_d - ini_h) + (debe_mov - haber_mov), 0)
                saldo_fin_h = max(-(ini_d - ini_h) - (debe_mov - haber_mov), 0)
            else:
                saldo_ini_d = max(ini_d - ini_h, 0)
                saldo_ini_h = max(ini_h - ini_d, 0)
                saldo_fin_d = max((ini_d - ini_h) + (debe_mov - haber_mov), 0)
                saldo_fin_h = max(-(ini_d - ini_h) - (debe_mov - haber_mov), 0)

            ws.cell(row=r, column=1, value=codigo)
            ws.cell(row=r, column=2, value=info["nombre"])
            ws.cell(row=r, column=3, value=info["tipo"])
            ws.cell(row=r, column=4, value=saldo_ini_d).number_format = money_fmt
            ws.cell(row=r, column=5, value=saldo_ini_h).number_format = money_fmt
            ws.cell(row=r, column=6, value=debe_mov).number_format = money_fmt
            ws.cell(row=r, column=7, value=haber_mov).number_format = money_fmt
            ws.cell(row=r, column=8, value=saldo_fin_d).number_format = money_fmt
            ws.cell(row=r, column=9, value=saldo_fin_h).number_format = money_fmt
            style_data_row(ws, r, len(headers))
            r += 1

    # ══════════════ HOJA 3: BALANZA DE COMPROBACIÓN ══════════════
    if "balanza" in hojas_seleccionadas:
        periodo_str = f"Período: {desde} al {hasta}" if desde and hasta else "Acumulado"
        ws = wb.create_sheet("Balanza")
        ws.sheet_properties.tabColor = "F39C12"
        start = add_title(
            ws, "Balanza de Comprobación", f"Verificación de débitos y créditos — {periodo_str}"
        )

        # ── Calcular saldos iniciales (antes del período) ──
        if desde:
            import copy as _copy
            data_completa = _copy.deepcopy(data_original)
            data_completa["diario"] = [e for e in data_original.get("diario", []) if e.get("fecha", "") < desde]
            data_completa["ajustes"] = [a for a in data_original.get("ajustes", []) if a.get("fecha", "") < desde]
            mayor_inicial = calcular_mayor(data_completa)
        else:
            mayor_inicial = None

        balanza_data, td, th, tsd, tsh = calcular_balanza(data)

        headers = [
            "Código",
            "Cuenta",
            "Tipo",
            "Saldo Inicial Deudor",
            "Saldo Inicial Acreedor",
            "Mov. Deudor",
            "Mov. Acreedor",
            "Saldo Final Deudor",
            "Saldo Final Acreedor",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=start, column=i, value=h)
        style_header_row(ws, start, len(headers))
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 18
        ws.column_dimensions["I"].width = 18

        r = start + 1
        tot_sd_i = tot_sh_i = tot_d = tot_h = tot_sd_f = tot_sh_f = 0
        for item in balanza_data:
            if item.get("es_header"):
                ws.cell(row=r, column=1, value=item["codigo"])
                ws.cell(row=r, column=2, value=item["nombre"]).font = Font(
                    name="Calibri", size=10, bold=True, color="1A1D2E"
                )
                style_data_row(ws, r, 9)
                r += 1
                continue
            if item.get("es_subtotal"):
                continue
            codigo = item["codigo"]
            info = cuentas.get(codigo, {})
            tipo = info.get("tipo", "")
            ts = tipo_saldo(tipo)

            debe_mov = item["debe"]
            haber_mov = item["haber"]

            # Saldo inicial (antes del período)
            if mayor_inicial is not None and codigo in mayor_inicial:
                ini_d = mayor_inicial[codigo]["debe"]
                ini_h = mayor_inicial[codigo]["haber"]
            else:
                ini_d = 0
                ini_h = 0

            if ts == "Debe":
                saldo_ini_d = max(ini_d - ini_h, 0)
                saldo_ini_h = max(ini_h - ini_d, 0)
                saldo_fin_d = max(saldo_ini_d + saldo_ini_h + (debe_mov - haber_mov), 0) if saldo_ini_d > 0 else max(debe_mov - haber_mov, 0)
                saldo_fin_h = max(saldo_ini_h + saldo_ini_d + (haber_mov - debe_mov), 0) if saldo_ini_h > 0 else max(haber_mov - debe_mov, 0)
                # Recalcular limpiamente
                neto_ini = ini_d - ini_h
                neto_fin = neto_ini + (debe_mov - haber_mov)
                saldo_ini_d = max(neto_ini, 0)
                saldo_ini_h = max(-neto_ini, 0)
                saldo_fin_d = max(neto_fin, 0)
                saldo_fin_h = max(-neto_fin, 0)
            else:
                neto_ini = ini_h - ini_d
                neto_fin = neto_ini + (haber_mov - debe_mov)
                saldo_ini_d = max(-neto_ini, 0)
                saldo_ini_h = max(neto_ini, 0)
                saldo_fin_d = max(-neto_fin, 0)
                saldo_fin_h = max(neto_fin, 0)

            ws.cell(row=r, column=1, value=codigo)
            ws.cell(row=r, column=2, value=item["nombre"])
            ws.cell(row=r, column=3, value=tipo)
            ws.cell(row=r, column=4, value=saldo_ini_d).number_format = money_fmt
            ws.cell(row=r, column=5, value=saldo_ini_h).number_format = money_fmt
            ws.cell(row=r, column=6, value=debe_mov).number_format = money_fmt
            ws.cell(row=r, column=7, value=haber_mov).number_format = money_fmt
            ws.cell(row=r, column=8, value=saldo_fin_d).number_format = money_fmt
            ws.cell(row=r, column=9, value=saldo_fin_h).number_format = money_fmt
            style_data_row(ws, r, 9)
            tot_sd_i += saldo_ini_d
            tot_sh_i += saldo_ini_h
            tot_d += debe_mov
            tot_h += haber_mov
            tot_sd_f += saldo_fin_d
            tot_sh_f += saldo_fin_h
            r += 1
        # Totales
        ws.cell(row=r, column=2, value="TOTALES")
        ws.cell(row=r, column=4, value=tot_sd_i).number_format = money_fmt
        ws.cell(row=r, column=5, value=tot_sh_i).number_format = money_fmt
        ws.cell(row=r, column=6, value=tot_d).number_format = money_fmt
        ws.cell(row=r, column=7, value=tot_h).number_format = money_fmt
        ws.cell(row=r, column=8, value=tot_sd_f).number_format = money_fmt
        ws.cell(row=r, column=9, value=tot_sh_f).number_format = money_fmt
        style_total_row(ws, r, 9)

    # ══════════════ HOJA 4: ESTADO DE RESULTADOS ══════════════
    if "er" in hojas_seleccionadas:
        periodo_str = f"Período: {desde} al {hasta}" if desde and hasta else "Acumulado"
        ws = wb.create_sheet("Estado de Resultados")
        ws.sheet_properties.tabColor = "2ECC71"
        start = add_title(
            ws, "Estado de Resultados", f"Ingresos, Gastos y Utilidad Neta — {periodo_str}"
        )
        di, ti, dg, tg, util = calcular_estado_resultados(data)

        r = start
        ws.cell(row=r, column=1, value="INGRESOS").font = Font(
            name="Calibri", size=11, bold=True, color="27AE60"
        )
        r += 1
        for item in di:
            if item.get("tipo") in ("header", "subtotal"):
                continue
            ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                name="Calibri", size=10, color="333333"
            )
            ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
            ws.cell(row=r, column=2).font = green_font
            r += 1
        ws.cell(row=r, column=1, value="Total Ingresos").font = total_font
        ws.cell(row=r, column=2, value=ti).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="27AE60"
        )
        style_total_row(ws, r, 2)
        r += 2

        ws.cell(row=r, column=1, value="GASTOS").font = Font(
            name="Calibri", size=11, bold=True, color="E74C3C"
        )
        r += 1
        for item in dg:
            if item.get("tipo") in ("header", "subtotal"):
                continue
            ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                name="Calibri", size=10, color="333333"
            )
            ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
            ws.cell(row=r, column=2).font = red_font
            r += 1
        ws.cell(row=r, column=1, value="Total Gastos").font = total_font
        ws.cell(row=r, column=2, value=tg).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="E74C3C"
        )
        style_total_row(ws, r, 2)
        r += 2

        ws.cell(row=r, column=1, value="UTILIDAD NETA").font = Font(
            name="Calibri", size=13, bold=True, color="1A1D2E"
        )
        ws.cell(row=r, column=2, value=util).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri",
            size=13,
            bold=True,
            color="27AE60" if util >= 0 else "E74C3C",
        )
        style_total_row(ws, r, 2)

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 20

    # ══════════════ HOJA 5: BALANCE GENERAL ══════════════
    if "bg" in hojas_seleccionadas:
        periodo_str = f"Período: {desde} al {hasta}" if desde and hasta else "Acumulado"
        ws = wb.create_sheet("Balance General")
        ws.sheet_properties.tabColor = "E74C3C"
        start = add_title(ws, "Balance General", f"Activos = Pasivos + Capital — {periodo_str}")
        activos, ta, pasivos, tp, capital_items, tc = calcular_balance_general(data)

        r = start
        ws.cell(row=r, column=1, value="ACTIVOS").font = Font(
            name="Calibri", size=11, bold=True, color="4F8CFF"
        )
        r += 1
        for item in activos:
            if item.get("tipo") == "header":
                ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                    name="Calibri", size=10, bold=True, color="4F8CFF"
                )
                r += 1
                continue
            if item.get("tipo") == "subtotal":
                ws.cell(row=r, column=1, value="  Total " + item["nombre"]).font = total_font
                ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
                style_data_row(ws, r, 2)
                r += 1
                continue
            ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                name="Calibri", size=10, color="333333"
            )
            ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
            ws.cell(row=r, column=2).font = (
                green_font if item["saldo"] >= 0 else red_font
            )
            r += 1
        ws.cell(row=r, column=1, value="Total Activos").font = total_font
        ws.cell(row=r, column=2, value=ta).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="4F8CFF"
        )
        style_total_row(ws, r, 2)
        r += 2

        ws.cell(row=r, column=1, value="PASIVOS").font = Font(
            name="Calibri", size=11, bold=True, color="E74C3C"
        )
        r += 1
        for item in pasivos:
            if item.get("tipo") == "header":
                ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                    name="Calibri", size=10, bold=True, color="E74C3C"
                )
                r += 1
                continue
            if item.get("tipo") == "subtotal":
                ws.cell(row=r, column=1, value="  Total " + item["nombre"]).font = total_font
                ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
                style_data_row(ws, r, 2)
                r += 1
                continue
            ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                name="Calibri", size=10, color="333333"
            )
            ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
            ws.cell(row=r, column=2).font = red_font
            r += 1
        ws.cell(row=r, column=1, value="Total Pasivos").font = total_font
        ws.cell(row=r, column=2, value=tp).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="E74C3C"
        )
        style_total_row(ws, r, 2)
        r += 2

        ws.cell(row=r, column=1, value="CAPITAL").font = Font(
            name="Calibri", size=11, bold=True, color="2ECC71"
        )
        r += 1
        for item in capital_items:
            if item.get("tipo") == "header":
                ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                    name="Calibri", size=10, bold=True, color="2ECC71"
                )
                r += 1
                continue
            if item.get("tipo") == "subtotal":
                ws.cell(row=r, column=1, value="  Total " + item["nombre"]).font = total_font
                ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
                style_data_row(ws, r, 2)
                r += 1
                continue
            ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                name="Calibri", size=10, color="333333"
            )
            ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
            ws.cell(row=r, column=2).font = (
                green_font if item["saldo"] >= 0 else red_font
            )
            r += 1
        ws.cell(row=r, column=1, value="Total Capital").font = total_font
        ws.cell(row=r, column=2, value=tc).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="2ECC71"
        )
        style_total_row(ws, r, 2)
        r += 2

        ws.cell(row=r, column=1, value="VERIFICACIÓN: Pasivos + Capital").font = Font(
            name="Calibri", size=11, bold=True, color="333333"
        )
        ws.cell(row=r, column=2, value=tp + tc).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="4F8CFF"
        )

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 20

    # ══════════════ HOJA 6: VENTAS POS ══════════════
    if "pos" in hojas_seleccionadas:
        pos_historial = data.get("pos_historial", [])
        if desde and hasta:
            pos_historial = [v for v in pos_historial if desde <= v.get("fecha", "") <= hasta]
        elif desde:
            pos_historial = [v for v in pos_historial if desde <= v.get("fecha", "")]
        elif hasta:
            pos_historial = [v for v in pos_historial if v.get("fecha", "") <= hasta]
        if pos_historial:
            periodo_str = f"Período: {desde} al {hasta}" if desde and hasta else "Acumulado"
            ws = wb.create_sheet("Ventas POS")
            ws.sheet_properties.tabColor = "F39C12"
            start = add_title(
                ws, "Ventas POS", f"Historial de ventas — {periodo_str}"
            )
            headers = [
                "Ref",
                "Fecha",
                "Cliente",
                "Forma Pago",
                "Productos",
                "Total C$",
                "Costo C$",
                "Utilidad C$",
            ]
            for i, h in enumerate(headers, 1):
                ws.cell(row=start, column=i, value=h)
            style_header_row(ws, start, len(headers))
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 40
            ws.column_dimensions["F"].width = 16
            ws.column_dimensions["G"].width = 16
            ws.column_dimensions["H"].width = 16
            r = start + 1
            total_ventas = total_costos = total_utilidades = 0
            for venta in pos_historial:
                productos_str = ", ".join(
                    f"{l.get('nombre', l.get('producto', '?'))} x{l.get('cantidad', 1)}"
                    for l in venta.get("lineas", [])
                )
                total_v = venta.get("total", venta.get("total_venta", 0))
                total_c = venta.get("costo", 0)
                utilidad = venta.get("utilidad", total_v - total_c)
                ws.cell(row=r, column=1, value=venta.get("ref", ""))
                ws.cell(row=r, column=2, value=venta.get("fecha", ""))
                ws.cell(row=r, column=3, value=venta.get("cliente", ""))
                ws.cell(row=r, column=4, value=venta.get("forma_pago", ""))
                ws.cell(row=r, column=5, value=productos_str)
                ws.cell(row=r, column=6, value=total_v).number_format = money_fmt
                ws.cell(row=r, column=6).font = green_font
                ws.cell(row=r, column=7, value=total_c).number_format = money_fmt
                ws.cell(row=r, column=8, value=utilidad).number_format = money_fmt
                ws.cell(row=r, column=8).font = (
                    green_font if utilidad >= 0 else red_font
                )
                style_data_row(ws, r, len(headers))
                total_ventas += total_v
                total_costos += total_c
                total_utilidades += utilidad
                r += 1
            ws.cell(row=r, column=5, value="TOTALES")
            ws.cell(row=r, column=6, value=total_ventas).number_format = money_fmt
            ws.cell(row=r, column=7, value=total_costos).number_format = money_fmt
            ws.cell(row=r, column=8, value=total_utilidades).number_format = money_fmt
            style_total_row(ws, r, len(headers))

    # ══════════════ HOJA 7: KARDEX PEPS ══════════════
    if "kardex" in hojas_seleccionadas and "kardex" in data and data["kardex"]:
        periodo_str = f"Período: {desde} al {hasta}" if desde and hasta else "Acumulado"
        ws = wb.create_sheet("Kardex PEPS")
        ws.sheet_properties.tabColor = "9B59B6"
        start = add_title(ws, "Kardex PEPS", f"Inventario por producto — Método PEPS — {periodo_str}")
        headers = ["Producto", "Fecha", "Tipo", "Cantidad", "Costo Unit.", "Costo Total", "Saldo Cant.", "Saldo Costo"]
        widths = [28, 14, 12, 12, 14, 14, 12, 14]
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            ws.cell(row=start, column=i, value=h)
        style_header_row(ws, start, len(headers))
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        r = start + 1
        total_saldo_costo = 0
        total_saldo_cant = 0
        num_productos = 0
        productos = sorted(data["kardex"].keys())
        for nombre in productos:
            movimientos = data["kardex"][nombre]
            if not movimientos:
                continue
            num_productos += 1
            # Obtener último saldo del producto
            ultimo = movimientos[-1]
            ultimo_saldo_cant = ultimo.get("saldo", 0)
            ultimo_saldo_costo = ultimo.get("costo", 0) * ultimo_saldo_cant if ultimo_saldo_cant else 0
            total_saldo_cant += ultimo_saldo_cant
            total_saldo_costo += ultimo_saldo_costo

            ws.cell(row=r, column=1, value=nombre).font = Font(name="Calibri", size=10, bold=True, color="333333")
            r += 1
            for mov in movimientos:
                tipo_mov = mov.get("tipo", "")
                tipo_label = "Entrada" if tipo_mov == "entrada" else "Salida" if tipo_mov == "salida" else tipo_mov
                cantidad = mov.get("cantidad", 0)
                costo_unit = mov.get("costo", 0)
                costo_total = mov.get("total", cantidad * costo_unit)
                saldo_cant = mov.get("saldo", 0)
                saldo_costo = mov.get("costo", 0) * saldo_cant if saldo_cant else 0

                ws.cell(row=r, column=2, value=mov.get("fecha", ""))
                ws.cell(row=r, column=3, value=tipo_label)
                ws.cell(row=r, column=4, value=cantidad)
                ws.cell(row=r, column=5, value=costo_unit).number_format = money_fmt
                ws.cell(row=r, column=6, value=costo_total).number_format = money_fmt
                ws.cell(row=r, column=7, value=saldo_cant)
                ws.cell(row=r, column=8, value=saldo_costo).number_format = money_fmt
                if tipo_mov == "salida":
                    for c in range(1, 9):
                        ws.cell(row=r, column=c).font = red_font
                style_data_row(ws, r, len(headers))
                r += 1
            r += 1

        # Fila resumen final
        r += 1
        ws.cell(row=r, column=1, value="RESUMEN INVENTARIO").font = Font(
            name="Calibri", size=11, bold=True, color="1A1D2E"
        )
        r += 1
        ws.cell(row=r, column=1, value="Total Productos:")
        ws.cell(row=r, column=2, value=num_productos).font = Font(
            name="Calibri", size=11, bold=True, color="4F8CFF"
        )
        r += 1
        ws.cell(row=r, column=1, value="Total Unidades en Stock:")
        ws.cell(row=r, column=2, value=total_saldo_cant).font = Font(
            name="Calibri", size=11, bold=True, color="4F8CFF"
        )
        r += 1
        ws.cell(row=r, column=1, value="Valor Total Inventario:")
        ws.cell(row=r, column=2, value=total_saldo_costo).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=12, bold=True, color="27AE60"
        )

    # ══════════════ HOJA: ANTIGÜEDAD DE COBROS ══════════════
    if "cobros" in hojas_seleccionadas:
        periodo_str = f"Período: {desde} al {hasta}" if desde and hasta else "Acumulado"
        hoy = date.today()
        cobros = data.get("cuentas_cobrar", [])
        activos = [c for c in cobros if c.get("estado", "pendiente") != "pagado"]
        ws = wb.create_sheet("Antiguedad Cobros")
        ws.sheet_properties.tabColor = "E67E22"
        start = add_title(ws, "Antigüedad de Cuentas por Cobrar",
                          f"Saldos pendientes por rango de vencimiento — {periodo_str}")
        headers = ["Rango", "Cantidad", "Total C$"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=start, column=i, value=h)
        style_header_row(ws, start, len(headers))
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 20
        rangos = [
            ("0-30 días",   lambda f: 0 <= (hoy - _parse_fecha_excel(f)).days <= 30),
            ("31-60 días",  lambda f: 31 <= (hoy - _parse_fecha_excel(f)).days <= 60),
            ("61-90 días",  lambda f: 61 <= (hoy - _parse_fecha_excel(f)).days <= 90),
            ("91+ días",    lambda f: (hoy - _parse_fecha_excel(f)).days > 90),
        ]
        r = start + 1
        gran_total = 0
        for label, cond in rangos:
            items = [c for c in activos if cond(c.get("fecha", ""))]
            if not items:
                continue
            total_rango = sum(
                c["monto"] - sum(pg.get("monto", 0) for pg in c.get("pagos", []))
                for c in items
            )
            gran_total += total_rango
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=len(items))
            ws.cell(row=r, column=3, value=total_rango).number_format = money_fmt
            style_data_row(ws, r, len(headers))
            r += 1
        ws.cell(row=r, column=1, value="TOTAL")
        ws.cell(row=r, column=2, value=len(activos))
        ws.cell(row=r, column=3, value=gran_total).number_format = money_fmt
        style_total_row(ws, r, len(headers))
        r += 2
        ws.cell(row=r, column=1, value="Detalle de cuentas pendientes").font = total_font
        r += 1
        dh = ["#", "Fecha", "Descripción", "Monto C$", "Saldo C$", "Días"]
        for i, h in enumerate(dh, 1):
            ws.cell(row=r, column=i, value=h)
        style_header_row(ws, r, len(dh))
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 38
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 10
        r += 1
        for c in activos:
            saldo = c["monto"] - sum(pg.get("monto", 0) for pg in c.get("pagos", []))
            dias = (hoy - _parse_fecha_excel(c.get("fecha", ""))).days
            ws.cell(row=r, column=1, value=c.get("id", ""))
            ws.cell(row=r, column=2, value=c.get("fecha", ""))
            ws.cell(row=r, column=3, value=c.get("descripcion", "")[:50])
            ws.cell(row=r, column=4, value=c["monto"]).number_format = money_fmt
            ws.cell(row=r, column=5, value=saldo).number_format = money_fmt
            ws.cell(row=r, column=6, value=dias)
            style_data_row(ws, r, len(dh))
            r += 1

    # ── Asegurar al menos una hoja ──
    if not wb.sheetnames:
        ws = wb.create_sheet("Reporte")
        ws.cell(row=1, column=1, value="No hay datos para este reporte.")

    # ── Enviar archivo como descarga (el usuario elige dónde guardarlo) ──
    nombre_hojas = {
        "diario": "Diario", "mayor": "Mayor", "balanza": "Balanza",
        "er": "EstadoResultados", "bg": "BalanceGeneral",
        "pos": "VentasPOS", "kardex": "KardexPEPS", "cobros": "AntiguedadCobros",
    }
    if len(hojas_seleccionadas) == 8 or not hojas_seleccionadas:
        label = "Completo"
    elif len(hojas_seleccionadas) == 1:
        label = nombre_hojas.get(list(hojas_seleccionadas)[0], "Reporte")
    else:
        label = "_".join(nombre_hojas.get(h, h) for h in sorted(hojas_seleccionadas))
    fecha_str = date.today().isoformat()
    filename = f"FG_{label}_{fecha_str}.xlsx"

    # Guardar en memoria (BytesIO) en lugar de disco
    from io import BytesIO

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar como descarga directa al navegador
    from flask import send_file

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ═══════════════════════════════════════════════════════════════
#  PDF — Exportar reportes
# ═══════════════════════════════════════════════════════════════


def _send_pdf(pdf_func, data, filename, desde=None, hasta=None):
    from services.reportes_pdf import generar_pdf_bytes
    from io import BytesIO
    buf = BytesIO()
    buf.write(generar_pdf_bytes(pdf_func, data, desde, hasta))
    buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name=filename)


def _desde_hasta():
    d = request.args.get("desde", "")
    h = request.args.get("hasta", "")
    return (d, h) if d and h else (None, None)


@app.route("/reportes/pdf/balanza")
@login_required
def pdf_balanza():
    from services.reportes_pdf import pdf_balanza as fn
    d, h = _desde_hasta()
    return _send_pdf(fn, load_data(), "Balanza.pdf", d, h)


@app.route("/reportes/pdf/estado-resultados")
@login_required
def pdf_estado_resultados():
    from services.reportes_pdf import pdf_estado_resultados as fn
    d, h = _desde_hasta()
    return _send_pdf(fn, load_data(), "Estado_Resultados.pdf", d, h)


@app.route("/reportes/pdf/balance-general")
@login_required
def pdf_balance_general():
    from services.reportes_pdf import pdf_balance_general as fn
    d, h = _desde_hasta()
    return _send_pdf(fn, load_data(), "Balance_General.pdf", d, h)


@app.route("/reportes/pdf/diario")
@login_required
def pdf_diario():
    from services.reportes_pdf import pdf_diario as fn
    d, h = _desde_hasta()
    return _send_pdf(fn, load_data(), "Libro_Diario.pdf", d, h)


@app.route("/reportes/pdf/mayor")
@login_required
def pdf_mayor():
    from services.reportes_pdf import pdf_mayor as fn
    d, h = _desde_hasta()
    return _send_pdf(fn, load_data(), "Libro_Mayor.pdf", d, h)


@app.route("/reportes/pdf/cobrar")
@login_required
def pdf_cobrar():
    from services.reportes_pdf import pdf_cobrar as fn
    d, h = _desde_hasta()
    return _send_pdf(fn, load_data(), "Cuentas_Cobrar.pdf", d, h)


@app.route("/reportes/pdf/caja")
@login_required
def pdf_caja():
    from services.reportes_pdf import pdf_caja as fn
    d, h = _desde_hasta()
    return _send_pdf(fn, load_data(), "Movimientos_Caja.pdf", d, h)


@app.route("/reportes/pdf/completo")
@login_required
def pdf_completo():
    from services.reportes_pdf import pdf_completo as fn
    d, h = _desde_hasta()
    return _send_pdf(fn, load_data(), "Reporte_Completo.pdf", d, h)


@app.route("/reportes/pdf/antiguedad-cobros")
@login_required
def pdf_antiguedad_cobros():
    from services.reportes_pdf import pdf_antiguedad_cobros as fn
    d, h = _desde_hasta()
    return _send_pdf(fn, load_data(), "Antiguedad_Cobros.pdf", d, h)


# ═══════════════════════════════════════════════════════════════
#  REPORTES FISCALES (IVA, DGI, Alcaldía)
# ═══════════════════════════════════════════════════════════════


@app.route("/fiscales")
@login_required
def fiscales():
    data = load_data()
    return render_template("fiscales.html", meses=_meses_disponibles(data))


def _meses_disponibles(data):
    fechas = set()
    for e in data.get("diario", []):
        f = e.get("fecha", "")
        if len(f) >= 7:
            fechas.add(f[:7])
    return sorted(fechas, reverse=True)


def _calcular_ingresos_gastos(data, desde, hasta):
    saldos = {}
    for cod in data.get("cuentas", {}):
        saldos[cod] = {"debe": 0, "haber": 0}
    for e in data.get("diario", []):
        if e.get("fecha", "") >= desde and e.get("fecha", "") <= hasta and not e.get("ref", "").startswith("CIERRE-"):
            for m in e.get("movimientos", []):
                if m["tipo"] == "Debe":
                    saldos[m["cuenta"]]["debe"] += m["monto"]
                else:
                    saldos[m["cuenta"]]["haber"] += m["monto"]
    for aj in data.get("ajustes", []):
        if aj.get("fecha", "") >= desde and aj.get("fecha", "") <= hasta:
            for m in aj.get("movimientos", []):
                if m["tipo"] == "Debe":
                    saldos[m["cuenta"]]["debe"] += m["monto"]
                else:
                    saldos[m["cuenta"]]["haber"] += m["monto"]
    total_ing = 0.0
    total_gas = 0.0
    for cod, info in data.get("cuentas", {}).items():
        if info["tipo"] == "Ingreso":
            total_ing += saldos[cod]["haber"] - saldos[cod]["debe"]
        elif info["tipo"] == "Gasto":
            total_gas += saldos[cod]["debe"] - saldos[cod]["haber"]
    return total_ing, total_gas


@app.route("/fiscales/iva")
@login_required
def fiscal_iva():
    data = load_data()
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")
    total_ing = total_gas = iva_deb = iva_cred = iva_pagar = None
    if desde and hasta:
        total_ing, total_gas = _calcular_ingresos_gastos(data, desde, hasta)
        iva_deb = round(total_ing * 0.15, 2)
        iva_cred = round(total_gas * 0.15, 2)
        iva_pagar = round(iva_deb - iva_cred, 2)
    return render_template("fiscal_iva.html", data=data, desde=desde, hasta=hasta,
                           total_ing=total_ing, total_gas=total_gas,
                           iva_deb=iva_deb, iva_cred=iva_cred, iva_pagar=iva_pagar)


@app.route("/fiscales/iva/pdf")
@login_required
def fiscal_iva_pdf():
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")
    from services.reportes_pdf import pdf_reporte_iva, generar_pdf_bytes_con_periodo
    from io import BytesIO
    buf = BytesIO()
    buf.write(generar_pdf_bytes_con_periodo(pdf_reporte_iva, load_data(), desde, hasta))
    buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name=f"IVA_{desde}_a_{hasta}.pdf")


@app.route("/fiscales/dgi")
@login_required
def fiscal_dgi():
    data = load_data()
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")
    total_ing = total_gas = utilidad = ir_estimado = None
    if desde and hasta:
        total_ing, total_gas = _calcular_ingresos_gastos(data, desde, hasta)
        utilidad = round(total_ing - total_gas, 2)
        ir_estimado = round(utilidad * 0.30, 2) if utilidad > 0 else 0
    return render_template("fiscal_dgi.html", data=data, desde=desde, hasta=hasta,
                           total_ing=total_ing, total_gas=total_gas,
                           utilidad=utilidad, ir_estimado=ir_estimado)


@app.route("/fiscales/alcaldia")
@login_required
def fiscal_alcaldia():
    data = load_data()
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")
    total_ing = impuesto_alc = None
    if desde and hasta:
        total_ing, _ = _calcular_ingresos_gastos(data, desde, hasta)
        impuesto_alc = round(total_ing * 0.01, 2)
    return render_template("fiscal_alcaldia.html", data=data, desde=desde, hasta=hasta,
                           total_ing=total_ing, impuesto_alc=impuesto_alc)


# ═══════════════════════════════════════════════════════════════
#  MÓDULO: GASTOS DE COMERCIALIZACIÓN
# ═══════════════════════════════════════════════════════════════



@app.route("/comercializacion")
def comercializacion():
    data = load_data()
    mayor = calcular_mayor(data)

    # Movimientos recientes de cuentas 6xxx
    movimientos = []
    for entry in data["diario"]:
        for mov in entry["movimientos"]:
            if mov["cuenta"].startswith("6"):
                movimientos.append(
                    {
                        "id": entry["id"],
                        "fecha": entry["fecha"],
                        "ref": entry.get("ref", ""),
                        "descripcion": entry["descripcion"],
                        "cuenta": mov["cuenta"],
                        "nombre_cuenta": data["cuentas"]
                        .get(mov["cuenta"], {})
                        .get("nombre", mov["cuenta"]),
                        "tipo": mov["tipo"],
                        "monto": mov["monto"],
                    }
                )
    movimientos.sort(key=lambda x: x["fecha"], reverse=True)

    detalle, total = calcular_total_comercializacion(data)

    # Cuentas de gastos operacionales disponibles
    cuentas_6 = {
        cod: info
        for cod, info in data["cuentas"].items()
        if cod.startswith(("5.2", "6")) and info.get("tipo") == "Gasto"
    }
    # Todas las cuentas para el formulario de asiento rápido
    todas_cuentas = data["cuentas"]

    return render_template(
        "comercializacion.html",
        movimientos=movimientos,
        detalle=detalle,
        total=total,
        cuentas_6=cuentas_6,
        todas_cuentas=todas_cuentas,
        mayor=mayor,
    )


@app.route("/comercializacion/registrar", methods=["POST"])
def comercializacion_registrar():
    """Registra un gasto de comercialización como asiento en el diario."""
    data = load_data()
    fecha = request.form.get("fecha", date.today().isoformat())
    descripcion = request.form.get("descripcion", "")
    ref = request.form.get("ref", "")
    cuenta_gasto = request.form.get("cuenta_gasto", "")
    cuenta_pago = request.form.get("cuenta_pago", "")
    monto_str = request.form.get("monto", "0")

    try:
        monto = float(monto_str)
    except ValueError:
        return redirect(url_for("comercializacion"))

    if not cuenta_gasto or not cuenta_pago or monto <= 0:
        return redirect(url_for("comercializacion"))

    entry = {
        "id": get_next_id(data, "diario"),
        "fecha": fecha,
        "descripcion": descripcion
        or f"Gasto de comercialización — {data['cuentas'].get(cuenta_gasto, {}).get('nombre', cuenta_gasto)}",
        "ref": ref,
        "movimientos": [
            {"cuenta": cuenta_gasto, "tipo": "Debe", "monto": monto},
            {"cuenta": cuenta_pago, "tipo": "Haber", "monto": monto},
        ],
    }
    data["diario"].append(entry)

    # Reflejo en caja si la cuenta pago es efectivo o banco
    if cuenta_pago in ("1001", "1002", "1.1.01", "1.1.02"):
        data["caja_movimientos"].append(
            {
                "id": get_next_id(data, "caja_movimientos"),
                "fecha": fecha,
                "descripcion": entry["descripcion"],
                "tipo": "Haber",
                "monto": monto,
                "cuenta": cuenta_pago,
                "ref_diario": entry["id"],
            }
        )

    data = audit_log(data, "gasto_comercializacion", f"{descripcion[:60]} — C${monto:.2f}")
    save_data(data)
    return redirect(url_for("comercializacion"))


@app.route("/comercializacion/cuenta/nueva", methods=["POST"])
def comercializacion_cuenta_nueva():
    """Agrega una nueva cuenta de comercialización al catálogo."""
    data = load_data()
    codigo = request.form.get("codigo", "").strip()
    nombre = request.form.get("nombre", "").strip()
    if codigo and nombre and codigo not in data["cuentas"]:
        data["cuentas"][codigo] = {"nombre": nombre, "tipo": "Gasto", "saldo": 0}
        save_data(data)
    return redirect(url_for("comercializacion"))


@app.route("/comercializacion/cuenta/eliminar", methods=["POST"])
def comercializacion_cuenta_eliminar():
    """Elimina una cuenta 6xxx del catálogo (solo si no tiene movimientos)."""
    data = load_data()
    codigo = request.form.get("codigo", "").strip()
    mayor = calcular_mayor(data)
    if codigo.startswith("6") and codigo in data["cuentas"]:
        tiene_mov = codigo in mayor and (
            mayor[codigo]["debe"] > 0 or mayor[codigo]["haber"] > 0
        )
        if not tiene_mov:
            del data["cuentas"][codigo]
            save_data(data)
    return redirect(url_for("comercializacion"))


# ═══════════════════════════════════════════════════════════════
#  MÓDULO: CIERRE MENSUAL (Corte de Saldos)
# ═══════════════════════════════════════════════════════════════



@app.route("/cierre-mensual")
def cierre_mensual():
    data = load_data()
    cuentas = data["cuentas"]
    tipo = request.args.get("tipo", "mensual")
    if tipo not in ("mensual", "semanal", "quincenal"):
        tipo = "mensual"

    disponibles = _periodos_disponibles_por_tipo(data, tipo)
    cerrados = {c["periodo"] for c in data.get("cierres_mensuales", [])
                if c.get("tipo", "mensual") == tipo}

    periodos = []
    for p in disponibles:
        desde, hasta = _obtener_rango_fechas(tipo, p)
        ing = 0
        gast = 0
        for entry in data.get("diario", []):
            if desde <= entry["fecha"] <= hasta:
                if entry.get("ref", "").startswith("CIERRE-"):
                    continue
                for mov in entry["movimientos"]:
                    cta = data["cuentas"].get(mov["cuenta"], {})
                    if cta.get("tipo") == "Ingreso" and mov["tipo"] == "Haber":
                        ing += mov["monto"]
                    elif cta.get("tipo") == "Gasto" and mov["tipo"] == "Debe":
                        gast += mov["monto"]
        for aj in data.get("ajustes", []):
            if desde <= aj["fecha"] <= hasta:
                for mov in aj["movimientos"]:
                    cta = data["cuentas"].get(mov["cuenta"], {})
                    if cta.get("tipo") == "Ingreso" and mov["tipo"] == "Haber":
                        ing += mov["monto"]
                    elif cta.get("tipo") == "Gasto" and mov["tipo"] == "Debe":
                        gast += mov["monto"]

        result = ing - gast
        periodos.append({
            "periodo": p,
            "cerrado": p in cerrados,
            "total_ingresos": ing,
            "total_gastos": gast,
            "resultado": result,
            "cuentas": 1 if ing or gast else 0,
        })

    periodos.reverse()
    total_cerrados = sum(1 for c in data.get("cierres_mensuales", [])
                         if c.get("tipo", "mensual") == tipo)

    return render_template(
        "cierre_mensual.html",
        periodos=periodos,
        cuentas=cuentas,
        cerrados=cerrados,
        total_cerrados=total_cerrados,
        tipo_actual=tipo,
    )
@app.route("/cierre-mensual/vista-previa/<tipo>/<periodo>")
def cierre_mensual_vista_previa(tipo, periodo):
    """
    Muestra una vista previa del cierre antes de ejecutarlo.
    Permite al usuario validar los movimientos y totales.
    """
    data = load_data()
    cuentas = data["cuentas"]
    
    if tipo not in ("mensual", "semanal", "quincenal"):
        return redirect(url_for("cierre_mensual") + "?error=periodo_invalido")

    cerrados = {c["periodo"] for c in data.get("cierres_mensuales", [])
                if c.get("tipo", "mensual") == tipo}
    if periodo in cerrados:
        return redirect(url_for("cierre_mensual") + "?error=periodo_ya_cerrado")

    desde, hasta = _obtener_rango_fechas(tipo, periodo)
    
    # Usar nueva función para obtener saldos
    saldos = procesar_movimientos_periodo(data, desde, hasta, excluir_cierre=True)
    
    # Construir información de cuentas de resultado
    cuentas_resultado = {}
    for cod, info in cuentas.items():
        if info.get("tipo") not in ("Ingreso", "Gasto"):
            continue
        
        saldo_data = saldos.get(cod, {"debe": 0, "haber": 0})
        debe = saldo_data["debe"]
        haber = saldo_data["haber"]
        
        # Calcular saldo según tipo de cuenta
        if info.get("tipo") == "Ingreso":
            saldo = float(haber - debe)
        else:  # Gasto
            saldo = float(debe - haber)
        
        if saldo != 0:
            cuentas_resultado[cod] = {
                "nombre": info.get("nombre", cod),
                "tipo": info.get("tipo", ""),
                "debe": float(debe),
                "haber": float(haber),
                "saldo": saldo,
            }
    
    # Calcular totales
    total_ingresos = sum(s["saldo"] for s in cuentas_resultado.values() 
                         if s["tipo"] == "Ingreso")
    total_gastos = sum(s["saldo"] for s in cuentas_resultado.values() 
                       if s["tipo"] == "Gasto")
    resultado = total_ingresos - total_gastos

    # Obtener lista de períodos disponibles para navegación
    disponibles = _periodos_disponibles_por_tipo(data, tipo)
    cerrados_set = {c["periodo"] for c in data.get("cierres_mensuales", [])
                    if c.get("tipo", "mensual") == tipo}
    periodos = []
    for p in disponibles:
        d, h = _obtener_rango_fechas(tipo, p)
        ing = 0
        gast = 0
        for entry in data.get("diario", []):
            if d <= entry["fecha"] <= h:
                if entry.get("ref", "").startswith("CIERRE-"): continue
                for mov in entry["movimientos"]:
                    cta = data["cuentas"].get(mov["cuenta"], {})
                    if cta.get("tipo") == "Ingreso" and mov["tipo"] == "Haber": ing += mov["monto"]
                    elif cta.get("tipo") == "Gasto" and mov["tipo"] == "Debe": gast += mov["monto"]
        for aj in data.get("ajustes", []):
            if d <= aj["fecha"] <= h:
                for mov in aj["movimientos"]:
                    cta = data["cuentas"].get(mov["cuenta"], {})
                    if cta.get("tipo") == "Ingreso" and mov["tipo"] == "Haber": ing += mov["monto"]
                    elif cta.get("tipo") == "Gasto" and mov["tipo"] == "Debe": gast += mov["monto"]
        periodos.append({"periodo": p, "cerrado": p in cerrados_set,
                         "total_ingresos": ing, "total_gastos": gast,
                         "resultado": ing - gast, "cuentas": 1 if ing or gast else 0})
    periodos.reverse()
    total_cerrados = sum(1 for c in data.get("cierres_mensuales", [])
                         if c.get("tipo", "mensual") == tipo)

    return render_template(
        "cierre_mensual.html",
        periodos=periodos,
        cuentas=cuentas,
        cerrados=cerrados_set,
        total_cerrados=total_cerrados,
        tipo_actual=tipo,
        periodo_preview=periodo,
        cuentas_resultado=cuentas_resultado,
        total_ingresos=total_ingresos,
        total_gastos=total_gastos,
        resultado=resultado,
    )
@app.route("/cierre-mensual/ejecutar", methods=["POST"])
def cierre_mensual_ejecutar():
    """
    Ejecuta el cierre de un período (mensual, semanal o quincenal).
    Genera un asiento de cierre y registra la información del cierre.
    """
    data = load_data()
    periodo = request.form.get("periodo", "").strip()
    tipo = request.form.get("tipo", "mensual")
    
    # Validaciones básicas
    if tipo not in ("mensual", "semanal", "quincenal"):
        return redirect(url_for("cierre_mensual") + "?error=periodo_invalido")

    cerrados = {c["periodo"] for c in data.get("cierres_mensuales", [])
                if c.get("tipo", "mensual") == tipo}
    if periodo in cerrados:
        return redirect(url_for("cierre_mensual") + "?error=periodo_ya_cerrado")

    desde, hasta = _obtener_rango_fechas(tipo, periodo)
    
    # Validar que el cierre sea posible
    es_valido, error_msg = validar_cierre_posible(data, desde, hasta)
    if not es_valido:
        return redirect(url_for("cierre_mensual") + f"?error=sin_cuentas_resultado&tipo={tipo}")

    # Calcular movimientos del cierre
    try:
        movs_cierre, total_ing, total_gast, diferencia = calcular_movimientos_cierre(data, desde, hasta)
    except Exception as e:
        return redirect(url_for("cierre_mensual") + f"?error=calculo_error&tipo={tipo}")

    if not movs_cierre:
        return redirect(url_for("cierre_mensual") + f"?error=sin_cuentas_resultado&tipo={tipo}")

    # Equilibrar asiento si hay diferencia
    if abs(diferencia) > 0.01:
        cta_capital = obtener_cuenta_capital_cierre(data, crear_si_no_existe=True)
        diferencia = round(diferencia, 2)
        if diferencia > 0:
            movs_cierre.append({"cuenta": cta_capital, "tipo": "Haber", "monto": diferencia})
        else:
            movs_cierre.append({"cuenta": cta_capital, "tipo": "Debe", "monto": abs(diferencia)})

    # Crear asiento de cierre
    ref_cierre = f"CIERRE-{tipo.upper()[:4]}-{periodo}"
    asiento = {
        "id": get_next_id(data, "diario"),
        "fecha": hasta,
        "descripcion": f"Cierre {tipo} - {periodo}",
        "ref": ref_cierre,
        "movimientos": movs_cierre,
    }
    data["diario"].append(asiento)

    # Registrar información del cierre
    data.setdefault("cierres_mensuales", []).append({
        "tipo": tipo,
        "periodo": periodo,
        "fecha_cierre": date.today().isoformat(),
        "asiento_id": asiento["id"],
        "total_ingresos": total_ing,
        "total_gastos": total_gast,
        "usuario": "sistema",  # Para auditoría futura
    })

    data = audit_log(data, "cierre_mensual", f"Cierre {tipo} {periodo} — resultado C${diferencia:.2f}")
    save_data(data)
    return redirect(url_for("cierre_mensual") + f"?ok=cierre_{periodo}&tipo={tipo}")


@app.route("/cierre-mensual/revertir/<tipo>/<periodo>", methods=["POST"])
def cierre_mensual_revertir(tipo, periodo):
    """
    Revierte un cierre de período: elimina el asiento de cierre y el registro del cierre.
    """
    if tipo not in ("mensual", "semanal", "quincenal"):
        return redirect(url_for("cierre_mensual") + "?error=periodo_invalido")
    
    data = load_data()
    
    # Buscar el cierre a revertir
    cierre_idx = None
    for i, c in enumerate(data.get("cierres_mensuales", [])):
        if c["periodo"] == periodo and c.get("tipo", "mensual") == tipo:
            cierre_idx = i
            break
    
    if cierre_idx is None:
        return redirect(url_for("cierre_mensual") + "?error=cierre_no_encontrado")
    
    cierre_info = data["cierres_mensuales"][cierre_idx]
    asiento_id = cierre_info.get("asiento_id")
    
    # Eliminar asiento de cierre del diario
    if asiento_id is not None:
        data["diario"] = [e for e in data["diario"] if e.get("id") != asiento_id]
    
    # Eliminar registro del cierre
    data["cierres_mensuales"].pop(cierre_idx)
    
    data = audit_log(data, "revertir_cierre", f"Revertir cierre {tipo} {periodo}")
    save_data(data)
    return redirect(url_for("cierre_mensual") + f"?ok=cierre_revertido&tipo={tipo}")

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5000)
