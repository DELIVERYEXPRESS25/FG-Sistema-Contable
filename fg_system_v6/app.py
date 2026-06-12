from flask import Flask, render_template, request, redirect, url_for, jsonify, g
import json, os, signal, sys
from datetime import datetime, date
from collections import defaultdict
import kardex_peps
import db_internal  # Base de datos interna embebida (sin archivos externos)

app = Flask(__name__)

# Almacenamiento temporal de archivos Excel entre la detección de hojas y la importación
_archivos_temp = {}


# ══════════════════════════════════════════════════════════════
# HELPERS: Generación de IDs únicos
# ══════════════════════════════════════════════════════════════
def get_next_id(data, key):
    """Retorna el siguiente ID único para una lista en data."""
    if key not in data or not data[key]:
        return 1
    items = data[key]
    max_id = 0
    for item in items:
        if isinstance(item, dict) and "id" in item:
            if item["id"] > max_id:
                max_id = item["id"]
    return max_id + 1


def ensure_ids(data):
    """Asegura que todas las listas tengan IDs únicos."""
    for key, lista in data.items():
        if isinstance(lista, list) and lista:
            seen = set()
            next_id_for_key = 1
            for item in lista:
                if isinstance(item, dict):
                    if "id" not in item:
                        while next_id_for_key in seen:
                            next_id_for_key += 1
                        item["id"] = next_id_for_key
                        seen.add(next_id_for_key)
                        next_id_for_key += 1
                    else:
                        if item["id"] in seen:
                            while next_id_for_key in seen:
                                next_id_for_key += 1
                            item["id"] = next_id_for_key
                            next_id_for_key += 1
                        seen.add(item["id"])


# ══════════════════════════════════════════════════════════════
# HELPERS: rutas persistentes para modo compilado
# ══════════════════════════════════════════════════════════════
def get_data_dir():
    """
    Retorna la carpeta donde deben guardarse archivos persistentes
    (data.json, reportes/, etc).

    En modo compilado (--onefile), PyInstaller extrae a _MEIPASS temporal,
    pero los datos deben persistir → usamos la carpeta del .exe.
    En desarrollo, usamos la carpeta actual.
    """
    if getattr(sys, "frozen", False):
        # Compilado: carpeta del ejecutable
        return os.path.dirname(sys.executable)
    else:
        # Desarrollo: carpeta donde está este archivo
        return os.path.dirname(os.path.abspath(__file__))


# ── DATA_FILE: ubicación persistente ──
DATA_FILE = os.path.join(get_data_dir(), "data.json")


# ─── DATOS INICIALES ───────────────────────────
def empty_data():
    return {
        "cuentas": {
            "1001": {"nombre": "Efectivo", "tipo": "Activo", "saldo": 0},
            "1002": {"nombre": "Bancos", "tipo": "Activo", "saldo": 0},
            "1003": {"nombre": "Cuentas por Cobrar", "tipo": "Activo", "saldo": 0},
            "1100": {"nombre": "Inventarios", "tipo": "Activo", "saldo": 0},
            "1200": {"nombre": "Equipos", "tipo": "Activo", "saldo": 0},
            "1300": {"nombre": "Mobiliario y Enseres", "tipo": "Activo", "saldo": 0},
            "1400": {"nombre": "Vehículos", "tipo": "Activo", "saldo": 0},
            "1500": {"nombre": "Edificios", "tipo": "Activo", "saldo": 0},
            "2001": {"nombre": "Cuentas por Pagar", "tipo": "Pasivo", "saldo": 0},
            "2002": {"nombre": "Préstamos Bancarios", "tipo": "Pasivo", "saldo": 0},
            "2100": {"nombre": "Impuestos por Pagar", "tipo": "Pasivo", "saldo": 0},
            "2200": {"nombre": "Sueldos por Pagar", "tipo": "Pasivo", "saldo": 0},
            "3001": {"nombre": "Capital", "tipo": "Capital", "saldo": 0},
            "3002": {"nombre": "Utilidades Retenidas", "tipo": "Capital", "saldo": 0},
            "4001": {"nombre": "Ventas", "tipo": "Ingreso", "saldo": 0},
            "4002": {"nombre": "Otros Ingresos", "tipo": "Ingreso", "saldo": 0},
            "5001": {"nombre": "Costo de Ventas", "tipo": "Gasto", "saldo": 0},
            "5002": {"nombre": "Gastos de Administración", "tipo": "Gasto", "saldo": 0},
            "5003": {"nombre": "Gastos de Ventas", "tipo": "Gasto", "saldo": 0},
            "5004": {"nombre": "Gastos por Intereses", "tipo": "Gasto", "saldo": 0},
            "5005": {"nombre": "Gastos Varios", "tipo": "Gasto", "saldo": 0},
            "5006": {"nombre": "Gastos de Depreciación", "tipo": "Gasto", "saldo": 0},
            "6001": {"nombre": "Publicidad y Propaganda", "tipo": "Gasto", "saldo": 0},
            "6002": {"nombre": "Comisiones sobre Ventas", "tipo": "Gasto", "saldo": 0},
            "6003": {
                "nombre": "Transporte y Distribución",
                "tipo": "Gasto",
                "saldo": 0,
            },
            "6004": {"nombre": "Empaque y Embalaje", "tipo": "Gasto", "saldo": 0},
            "6005": {"nombre": "Promociones y Descuentos", "tipo": "Gasto", "saldo": 0},
            "6006": {
                "nombre": "Gastos de Ferias y Eventos",
                "tipo": "Gasto",
                "saldo": 0,
            },
            "6007": {"nombre": "Sueldos de Vendedores", "tipo": "Gasto", "saldo": 0},
            "6008": {
                "nombre": "Otros Gastos de Comercialización",
                "tipo": "Gasto",
                "saldo": 0,
            },
        },
        "diario": [],
        "kardex": {},
        "cuentas_cobrar": [],
        "caja_movimientos": [],
        "ajustes": [],
    }


def load_data():
    """Carga datos desde base de datos interna SQLite.
    Usa caché por-request (Flask g) para evitar lecturas repetidas a SQLite
    dentro de la misma petición HTTP — mejora de rendimiento."""
    if hasattr(g, "_data_cache") and g._data_cache is not None:
        return g._data_cache
    data = db_internal.load_data()
    # Asegurar IDs únicos en todas las listas
    ensure_ids(data)
    # Garantizar que todas las claves existan (compatibilidad con datos antiguos)
    _defaults = {
        "caja_movimientos": [],
        "cuentas_cobrar": [],
        "pos_historial": [],
        "ajustes": [],
        "diario": [],
        "kardex": {},
        "cuentas": {},
        "gastos_comercializacion": [],
    }
    for key, default in _defaults.items():
        if key not in data:
            data[key] = default

    # Compatibilidad: versiones antiguas usaban 'caja' en vez de 'caja_movimientos'
    if not data["caja_movimientos"] and data.get("caja"):
        data["caja_movimientos"] = data["caja"]

    # Garantizar que cada entrada del diario y ajustes tenga estructura válida
    for entry in data.get("diario", []):
        if "movimientos" not in entry:
            entry["movimientos"] = []
        if "ref" not in entry:
            entry["ref"] = ""
        if "id" not in entry:
            entry["id"] = 0
    for aj in data.get("ajustes", []):
        if "movimientos" not in aj:
            aj["movimientos"] = []
        if "ref" not in aj:
            aj["ref"] = ""

    # Garantizar cuentas de comercialización (6xxx) en catálogos existentes
    CUENTAS_6_DEFAULT = {
        "6001": {"nombre": "Publicidad y Propaganda", "tipo": "Gasto", "saldo": 0},
        "6002": {"nombre": "Comisiones sobre Ventas", "tipo": "Gasto", "saldo": 0},
        "6003": {"nombre": "Transporte y Distribución", "tipo": "Gasto", "saldo": 0},
        "6004": {"nombre": "Empaque y Embalaje", "tipo": "Gasto", "saldo": 0},
        "6005": {"nombre": "Promociones y Descuentos", "tipo": "Gasto", "saldo": 0},
        "6006": {"nombre": "Gastos de Ferias y Eventos", "tipo": "Gasto", "saldo": 0},
        "6007": {"nombre": "Sueldos de Vendedores", "tipo": "Gasto", "saldo": 0},
        "6008": {
            "nombre": "Otros Gastos de Comercialización",
            "tipo": "Gasto",
            "saldo": 0,
        },
    }
    _changed = False
    for cod, info in CUENTAS_6_DEFAULT.items():
        if cod not in data["cuentas"]:
            data["cuentas"][cod] = info
            _changed = True
    if _changed:
        db_internal.save_data(data)  # persistir inmediatamente

    # Guardar en caché de request
    try:
        g._data_cache = data
    except RuntimeError:
        pass  # Fuera de contexto de request (tests, CLI)
    return data


def save_data(data):
    """Guarda datos en base de datos interna SQLite e invalida caché del request."""
    db_internal.save_data(data)
    # Actualizar caché para que el resto del request vea los datos nuevos
    try:
        g._data_cache = data
    except RuntimeError:
        pass


# ─── HELPERS ────────────────────────────────────
def tipo_saldo(tipo_cuenta):
    return "Debe" if tipo_cuenta in ["Activo", "Gasto"] else "Haber"


def calcular_mayor(data):
    mayor = defaultdict(lambda: {"debe": 0, "haber": 0, "movimientos": []})
    for entry in data["diario"]:
        for mov in entry["movimientos"]:
            cuenta = mov["cuenta"]
            if mov["tipo"] == "Debe":
                mayor[cuenta]["debe"] += mov["monto"]
            else:
                mayor[cuenta]["haber"] += mov["monto"]
            mayor[cuenta]["movimientos"].append(
                {
                    "fecha": entry["fecha"],
                    "descripcion": entry["descripcion"],
                    "tipo": mov["tipo"],
                    "monto": mov["monto"],
                    "ref": entry.get("ref", ""),
                }
            )
    # Ajustes
    for aj in data.get("ajustes", []):
        for mov in aj["movimientos"]:
            cuenta = mov["cuenta"]
            if mov["tipo"] == "Debe":
                mayor[cuenta]["debe"] += mov["monto"]
            else:
                mayor[cuenta]["haber"] += mov["monto"]
            mayor[cuenta]["movimientos"].append(
                {
                    "fecha": aj["fecha"],
                    "descripcion": aj["descripcion"],
                    "tipo": mov["tipo"],
                    "monto": mov["monto"],
                    "ref": "AJ",
                }
            )
    return mayor


def calcular_balanza(data):
    mayor = calcular_mayor(data)
    cuentas = data["cuentas"]
    balanza = []
    total_debe = 0
    total_haber = 0
    total_saldo_d = 0
    total_saldo_h = 0
    for codigo in sorted(cuentas.keys()):
        info = cuentas[codigo]
        debe = mayor[codigo]["debe"] if codigo in mayor else 0
        haber = mayor[codigo]["haber"] if codigo in mayor else 0
        ts = tipo_saldo(info["tipo"])
        if ts == "Debe":
            saldo = debe - haber
            saldo_d = max(saldo, 0)
            saldo_h = max(-saldo, 0)
        else:
            saldo = haber - debe
            saldo_h = max(saldo, 0)
            saldo_d = max(-saldo, 0)
        balanza.append(
            {
                "codigo": codigo,
                "nombre": info["nombre"],
                "tipo": info["tipo"],
                "debe": debe,
                "haber": haber,
                "saldo_debe": saldo_d,
                "saldo_haber": saldo_h,
            }
        )
        total_debe += debe
        total_haber += haber
        total_saldo_d += saldo_d
        total_saldo_h += saldo_h
    return balanza, total_debe, total_haber, total_saldo_d, total_saldo_h


def calcular_estado_resultados(data):
    mayor = calcular_mayor(data)
    cuentas = data["cuentas"]
    ingresos = 0
    gastos = 0
    detalle_ingresos = []
    detalle_gastos = []
    for codigo, info in cuentas.items():
        debe = mayor[codigo]["debe"] if codigo in mayor else 0
        haber = mayor[codigo]["haber"] if codigo in mayor else 0
        if info["tipo"] == "Ingreso":
            monto = haber - debe
            ingresos += monto
            detalle_ingresos.append({"nombre": info["nombre"], "monto": monto})
        elif info["tipo"] == "Gasto":
            monto = debe - haber
            gastos += monto
            detalle_gastos.append({"nombre": info["nombre"], "monto": monto})
    utilidad = ingresos - gastos
    return detalle_ingresos, ingresos, detalle_gastos, gastos, utilidad


def calcular_balance_general(data):
    mayor = calcular_mayor(data)
    cuentas = data["cuentas"]
    _, _, _, _, utilidad = calcular_estado_resultados(data)
    activos = []
    pasivos = []
    capital_items = []
    total_activo = 0
    total_pasivo = 0
    total_capital = 0
    for codigo in sorted(cuentas.keys()):
        info = cuentas[codigo]
        debe = mayor[codigo]["debe"] if codigo in mayor else 0
        haber = mayor[codigo]["haber"] if codigo in mayor else 0
        ts = tipo_saldo(info["tipo"])
        if ts == "Debe":
            saldo = debe - haber
        else:
            saldo = haber - debe
        if info["tipo"] == "Activo":
            activos.append({"nombre": info["nombre"], "saldo": saldo})
            total_activo += saldo
        elif info["tipo"] == "Pasivo":
            pasivos.append({"nombre": info["nombre"], "saldo": saldo})
            total_pasivo += saldo
        elif info["tipo"] == "Capital":
            capital_items.append({"nombre": info["nombre"], "saldo": saldo})
            total_capital += saldo
    total_capital += utilidad
    capital_items.append({"nombre": "Utilidad del Periodo", "saldo": utilidad})
    return activos, total_activo, pasivos, total_pasivo, capital_items, total_capital


def get_ventas_por_dia(data):
    ventas = defaultdict(float)
    for entry in data["diario"]:
        for mov in entry["movimientos"]:
            cuenta = mov["cuenta"]
            if cuenta == "4001" and mov["tipo"] == "Haber":
                ventas[entry["fecha"]] += mov["monto"]
    return dict(sorted(ventas.items()))


def get_ventas_por_mes(data):
    ventas = defaultdict(float)
    for entry in data["diario"]:
        for mov in entry["movimientos"]:
            if mov["cuenta"] == "4001" and mov["tipo"] == "Haber":
                mes = entry["fecha"][:7]
                ventas[mes] += mov["monto"]
    return dict(sorted(ventas.items()))


def get_gastos_por_mes(data):
    gastos = defaultdict(float)
    for entry in data["diario"]:
        for mov in entry["movimientos"]:
            cuenta = mov["cuenta"]
            if (
                data["cuentas"].get(cuenta, {}).get("tipo") == "Gasto"
                and mov["tipo"] == "Debe"
            ):
                mes = entry["fecha"][:7]
                gastos[mes] += mov["monto"]
    return dict(sorted(gastos.items()))


# ─── RUTAS ──────────────────────────────────────
@app.route("/")
def index():
    data = load_data()
    ventas_dia = get_ventas_por_dia(data)
    ventas_mes = get_ventas_por_mes(data)
    gastos_mes = get_gastos_por_mes(data)
    _, _, _, _, utilidad = calcular_estado_resultados(data)
    mayor = calcular_mayor(data)
    # Total activos / pasivos para resumen
    total_activo = 0
    total_pasivo = 0
    for c, info in data["cuentas"].items():
        d = mayor[c]["debe"] if c in mayor else 0
        h = mayor[c]["haber"] if c in mayor else 0
        ts = tipo_saldo(info["tipo"])
        saldo = (d - h) if ts == "Debe" else (h - d)
        if info["tipo"] == "Activo":
            total_activo += saldo
        elif info["tipo"] == "Pasivo":
            total_pasivo += saldo
    return render_template(
        "index.html",
        ventas_dia=ventas_dia,
        ventas_mes=ventas_mes,
        gastos_mes=gastos_mes,
        utilidad=utilidad,
        total_activo=total_activo,
        total_pasivo=total_pasivo,
        diario=data["diario"],
        cuentas=data["cuentas"],
    )


# ── DIARIO ──
@app.route("/diario")
def diario():
    data = load_data()
    return render_template(
        "diario.html", diario=data["diario"], cuentas=data["cuentas"]
    )


@app.route("/diario/agregar", methods=["POST"])
def agregar_diario():
    data = load_data()
    fecha = request.form.get("fecha", date.today().isoformat())
    descripcion = request.form.get("descripcion", "")
    ref = request.form.get("ref", "")
    cuentas_sel = request.form.getlist("cuenta")
    tipos = request.form.getlist("tipo")
    montos = request.form.getlist("monto")
    movimientos = []
    for i in range(len(cuentas_sel)):
        if cuentas_sel[i] and montos[i]:
            movimientos.append(
                {"cuenta": cuentas_sel[i], "tipo": tipos[i], "monto": float(montos[i])}
            )
    if movimientos:
        entry = {
            "id": get_next_id(data, "diario"),
            "fecha": fecha,
            "descripcion": descripcion,
            "ref": ref,
            "movimientos": movimientos,
        }
        data["diario"].append(entry)
        # Si es venta, agregar a cuentas cobrar si la cuenta es 1003
        for mov in movimientos:
            if mov["cuenta"] == "1003" and mov["tipo"] == "Debe":
                data["cuentas_cobrar"].append(
                    {
                        "id": get_next_id(data, "cuentas_cobrar"),
                        "fecha": fecha,
                        "descripcion": descripcion,
                        "monto": mov["monto"],
                        "estado": "Pendiente",
                        "ref_diario": entry["id"],
                    }
                )
            # Caja
            if mov["cuenta"] in ["1001", "1002"]:
                data["caja_movimientos"].append(
                    {
                        "id": get_next_id(data, "caja_movimientos"),
                        "fecha": fecha,
                        "descripcion": descripcion,
                        "tipo": mov["tipo"],
                        "monto": mov["monto"],
                        "cuenta": mov["cuenta"],
                        "ref_diario": entry["id"],
                    }
                )
        save_data(data)
    return redirect(url_for("diario"))


# ── EDITAR ASIENTO ──
@app.route("/diario/editar/<int:asiento_id>", methods=["GET", "POST"])
def editar_asiento(asiento_id):
    data = load_data()

    if request.method == "GET":
        # Buscar el asiento por ID
        asiento = None
        for a in data["diario"]:
            if a.get("id") == asiento_id:
                asiento = a
                break

        if not asiento:
            return redirect(url_for("diario"))

        # Convertir movimientos a formato para el form
        # Separar debe y haber
        movimientos_form = []
        for mov in asiento.get("movimientos", []):
            debe = (
                mov.get("debe", 0)
                if "debe" in mov
                else (mov.get("monto", 0) if mov.get("tipo") == "Debe" else 0)
            )
            haber = (
                mov.get("haber", 0)
                if "haber" in mov
                else (mov.get("monto", 0) if mov.get("tipo") == "Haber" else 0)
            )

            movimientos_form.append(
                {"cuenta": mov.get("cuenta", ""), "debe": debe, "haber": haber}
            )

        return render_template(
            "diario_editar.html",
            asiento=asiento,
            movimientos=movimientos_form,
            cuentas=data["cuentas"],
        )

    else:  # POST - guardar cambios
        # Buscar índice del asiento
        idx = None
        for i, a in enumerate(data["diario"]):
            if a.get("id") == asiento_id:
                idx = i
                break

        if idx is None:
            return redirect(url_for("diario"))

        # Actualizar datos
        fecha = request.form.get("fecha", date.today().isoformat())
        descripcion = request.form.get("descripcion", "")
        ref = request.form.get("ref", "")

        cuentas_sel = request.form.getlist("cuenta")
        debes = request.form.getlist("debe")
        haberes = request.form.getlist("haber")

        movimientos = []
        for i in range(len(cuentas_sel)):
            if cuentas_sel[i]:
                debe = float(debes[i]) if debes[i] else 0
                haber = float(haberes[i]) if haberes[i] else 0

                if debe > 0 or haber > 0:
                    movimientos.append(
                        {"cuenta": cuentas_sel[i], "debe": debe, "haber": haber}
                    )

        if movimientos:
            data["diario"][idx] = {
                "id": asiento_id,
                "fecha": fecha,
                "descripcion": descripcion,
                "ref": ref,
                "movimientos": movimientos,
            }
            save_data(data)

        return redirect(url_for("diario"))


# ── BORRAR ASIENTO ──
@app.route("/diario/borrar/<int:asiento_id>", methods=["POST"])
def borrar_asiento(asiento_id):
    data = load_data()

    # Buscar y eliminar el asiento
    data["diario"] = [a for a in data["diario"] if a.get("id") != asiento_id]

    save_data(data)
    return redirect(url_for("diario"))


# ── MAYOR ──
@app.route("/mayor")
def mayor():
    try:
        data = load_data()
        mayor_data = calcular_mayor(data)
        return render_template("mayor.html", mayor=mayor_data, cuentas=data["cuentas"])
    except Exception as e:
        import traceback

        traceback.print_exc()
        return f"<pre>ERROR /mayor: {e}</pre>", 500


# ── BALANZA ──
@app.route("/balanza")
def balanza():
    try:
        data = load_data()
        balanza_data, td, th, tsd, tsh = calcular_balanza(data)
        return render_template(
            "balanza.html",
            balanza=balanza_data,
            total_debe=td,
            total_haber=th,
            total_saldo_debe=tsd,
            total_saldo_haber=tsh,
        )
    except Exception as e:
        import traceback

        traceback.print_exc()
        return f"<pre>ERROR /balanza: {e}</pre>", 500


# ── ESTADO DE RESULTADOS ──
@app.route("/estado_resultados")
def estado_resultados():
    try:
        data = load_data()
        di, ti, dg, tg, util = calcular_estado_resultados(data)
        return render_template(
            "estado_resultados.html",
            detalle_ingresos=di,
            total_ingresos=ti,
            detalle_gastos=dg,
            total_gastos=tg,
            utilidad=util,
        )
    except Exception as e:
        import traceback

        traceback.print_exc()
        return f"<pre>ERROR /estado_resultados: {e}</pre>", 500


# ── BALANCE GENERAL ──
@app.route("/balance_general")
def balance_general():
    try:
        data = load_data()
        activos, ta, pasivos, tp, capital, tc = calcular_balance_general(data)
        return render_template(
            "balance_general.html",
            activos=activos,
            total_activo=ta,
            pasivos=pasivos,
            total_pasivo=tp,
            capital=capital,
            total_capital=tc,
        )
    except Exception as e:
        import traceback

        traceback.print_exc()
        return f"<pre>ERROR /balance_general: {e}</pre>", 500


# ── KARDEX ──
@app.route("/kardex")
def kardex():
    try:
        data = load_data()

        # Asegurar que existan las estructuras necesarias
        if "kardex" not in data:
            data["kardex"] = {}
        if "kardex_peps" not in data:
            data["kardex_peps"] = {}
        if "productos" not in data:
            data["productos"] = {}

        # Generar reporte PEPS solo si hay productos
        reporte_peps = {}
        try:
            if data.get("kardex") or data.get("productos"):
                reporte_peps = kardex_peps.generar_reporte_kardex_peps(data)
        except Exception as e:
            print(f"Error generando reporte PEPS: {e}")
            import traceback

            traceback.print_exc()
            reporte_peps = {}

        productos_list = list(data.get("kardex", {}).keys())

        return render_template(
            "kardex.html",
            kardex=data.get("kardex", {}),
            productos=productos_list,
            reporte_peps=reporte_peps,
        )

    except Exception as e:
        print(f"ERROR EN /kardex: {e}")
        import traceback

        traceback.print_exc()
        # Retornar página de error amigable
        return (
            f"""
        <html>
        <body style="font-family: Arial; padding: 40px; background: #1a1d2e; color: white;">
            <h1>❌ Error en Kardex</h1>
            <p>Ocurrió un error al cargar el Kardex:</p>
            <pre style="background: #2d3142; padding: 20px; border-radius: 8px;">{str(e)}</pre>
            <p><a href="/" style="color: #4f8cff;">← Volver al inicio</a></p>
        </body>
        </html>
        """,
            500,
        )


@app.route("/kardex/agregar_producto", methods=["POST"])
def agregar_producto_kardex():
    data = load_data()

    # Asegurar estructuras existen
    if "kardex" not in data:
        data["kardex"] = {}
    if "kardex_peps" not in data:
        data["kardex_peps"] = {}
    if "productos" not in data:
        data["productos"] = {}

    nombre = request.form.get("nombre", "").strip()

    if nombre and nombre not in data["kardex"]:
        # Inicializar kardex tradicional
        data["kardex"][nombre] = []

        # Inicializar productos
        data["productos"][nombre] = {"nombre": nombre, "stock": 0, "costo_promedio": 0}

        # Inicializar kardex_peps
        data["kardex_peps"][nombre] = {
            "lotes": [],
            "stock_total": 0,
            "costo_promedio": 0,
        }

    save_data(data)
    return redirect(url_for("kardex"))


@app.route("/kardex/movimiento", methods=["POST"])
def agregar_mov_kardex():
    try:
        data = load_data()
        producto = request.form.get("producto", "")
        fecha = request.form.get("fecha", date.today().isoformat())
        tipo = request.form.get("tipo", "entrada")
        cantidad = int(request.form.get("cantidad", 0))
        costo = float(request.form.get("costo", 0))
        descripcion = request.form.get("descripcion", "")

        # Validaciones básicas
        if not producto:
            print("Error: Producto vacío")
            return redirect(url_for("kardex"))

        if cantidad <= 0:
            print("Error: Cantidad debe ser mayor a 0")
            return redirect(url_for("kardex"))

        # Asegurar que las estructuras existan
        if "kardex" not in data:
            data["kardex"] = {}
        if "kardex_peps" not in data:
            data["kardex_peps"] = {}

        try:
            if tipo.lower() == "entrada":
                # Agregar entrada usando PEPS
                print(
                    f"Agregando entrada PEPS: {producto}, {cantidad} unidades @ C$ {costo}"
                )
                data = kardex_peps.agregar_entrada_peps(
                    data, producto, fecha, cantidad, costo
                )
                print("✓ Entrada agregada exitosamente")

            elif tipo.lower() == "salida":
                # Procesar salida usando PEPS
                print(f"Procesando salida PEPS: {producto}, {cantidad} unidades")
                data, costo_total, lotes_usados = kardex_peps.procesar_salida_peps(
                    data, producto, fecha, cantidad
                )
                print(f"✓ Salida procesada. Costo total: C$ {costo_total}")

            elif tipo.lower() == "ajuste":
                # Para ajustes, usar el método tradicional
                if producto not in data["kardex"]:
                    data["kardex"][producto] = []

                saldo_anterior = (
                    data["kardex"][producto][-1]["saldo"]
                    if data["kardex"][producto]
                    else 0
                )

                data["kardex"][producto].append(
                    {
                        "fecha": fecha,
                        "tipo": "ajuste",
                        "cantidad": cantidad,
                        "costo": costo,
                        "precio": 0,
                        "total": 0,
                        "saldo": cantidad,  # El ajuste establece un saldo absoluto
                        "descripcion": descripcion or "Ajuste de inventario",
                    }
                )

                # También actualizar PEPS si es ajuste positivo
                if cantidad > saldo_anterior:
                    diferencia = cantidad - saldo_anterior
                    data = kardex_peps.agregar_entrada_peps(
                        data, producto, fecha, diferencia, costo
                    )

        except ValueError as e:
            # Errores controlados (ej: stock insuficiente)
            print(f"Error controlado: {e}")
            # Por ahora solo redirigir, podrías agregar flash messages
            return redirect(url_for("kardex"))

        except Exception as e:
            # Errores inesperados
            print(f"Error inesperado en movimiento: {e}")
            import traceback

            traceback.print_exc()
            return redirect(url_for("kardex"))

        save_data(data)
        print("✓ Datos guardados correctamente")

    except Exception as e:
        print(f"ERROR EN /kardex/movimiento: {e}")
        import traceback

        traceback.print_exc()

    return redirect(url_for("kardex"))


# ── CUENTAS POR COBRAR ──
@app.route("/cobrar")
def cobrar():
    data = load_data()
    return render_template("cobrar.html", cobrar=data["cuentas_cobrar"])


@app.route("/cobrar/pagar/<int:idx>", methods=["POST"])
def pagar_cobro(idx):
    data = load_data()
    if 0 <= idx < len(data["cuentas_cobrar"]):
        data["cuentas_cobrar"][idx]["estado"] = "Cobrado"
        data["cuentas_cobrar"][idx]["fecha_cobro"] = date.today().isoformat()
    save_data(data)
    return redirect(url_for("cobrar"))


# ── AUXILIAR DE CAJA ──
@app.route("/caja")
def caja():
    data = load_data()

    # Asegurar que exista la clave caja
    if "caja" not in data:
        data["caja"] = []

    movs = data["caja"]
    saldo = 0
    for m in movs:
        if m.get("tipo") == "Debe":
            saldo += m.get("monto", 0)
        else:
            saldo -= m.get("monto", 0)

    return render_template("caja.html", movimientos=movs, saldo=saldo)


# ── AJUSTES ──
@app.route("/ajustes")
def ajustes():
    data = load_data()
    return render_template(
        "ajustes.html", ajustes=data.get("ajustes", []), cuentas=data["cuentas"]
    )


@app.route("/ajustes/agregar", methods=["POST"])
def agregar_ajuste():
    data = load_data()
    fecha = request.form.get("fecha", date.today().isoformat())
    descripcion = request.form.get("descripcion", "")
    cuentas_sel = request.form.getlist("cuenta")
    tipos = request.form.getlist("tipo")
    montos = request.form.getlist("monto")
    movimientos = []
    for i in range(len(cuentas_sel)):
        if cuentas_sel[i] and montos[i]:
            movimientos.append(
                {"cuenta": cuentas_sel[i], "tipo": tipos[i], "monto": float(montos[i])}
            )
    if movimientos:
        data.setdefault("ajustes", []).append(
            {
                "id": get_next_id(data, "ajustes"),
                "fecha": fecha,
                "descripcion": descripcion,
                "movimientos": movimientos,
            }
        )
        save_data(data)
    return redirect(url_for("ajustes"))


# ── REPORTES (API JSON para gráficos) ──
@app.route("/api/ventas_dia")
def api_ventas_dia():
    return jsonify(get_ventas_por_dia(load_data()))


@app.route("/api/ventas_mes")
def api_ventas_mes():
    return jsonify(get_ventas_por_mes(load_data()))


@app.route("/api/gastos_mes")
def api_gastos_mes():
    return jsonify(get_gastos_por_mes(load_data()))


# ── VENTAS POS ──
@app.route("/pos")
def pos():
    data = load_data()
    productos = []

    # Obtener productos del kardex
    for nombre, k in data.get("kardex", {}).items():
        try:
            # Manejar kardex como lista o diccionario
            if isinstance(k, list):
                # Kardex como lista de movimientos
                saldo = 0
                costo = 0
                if k:
                    ultimo_mov = k[-1]
                    saldo = ultimo_mov.get("saldo", 0)
                    costo = ultimo_mov.get("costo", 0)

                productos.append({"nombre": nombre, "saldo": saldo, "costo": costo})
            elif isinstance(k, dict):
                # Kardex como diccionario (formato antiguo)
                productos.append(
                    {
                        "nombre": nombre,
                        "saldo": k.get("saldo_actual", k.get("saldo_inicial", 0)),
                        "costo": k.get("costo_unitario", 0),
                    }
                )
        except Exception as e:
            print(f"Error procesando producto {nombre}: {e}")
            continue

    # Historial ventas POS
    historial = data.get("pos_historial", [])

    return render_template("pos.html", productos=productos, historial=historial)


@app.route("/pos/venta", methods=["POST"])
def pos_venta():
    try:
        data = load_data()
        fecha = request.form.get("fecha", date.today().isoformat())
        cliente = request.form.get("cliente", "Cliente general")
        forma_pago = request.form.get("forma_pago", "Efectivo")
        productos_nombres = request.form.getlist("producto")
        cantidades = request.form.getlist("cantidad")
        precios = request.form.getlist("precio")

        lineas = []
        total_venta = 0
        total_costo = 0

        for i in range(len(productos_nombres)):
            nombre = productos_nombres[i]
            cantidad = float(cantidades[i]) if cantidades[i] else 0
            precio = float(precios[i]) if precios[i] else 0

            if nombre and cantidad > 0 and precio > 0:
                # Obtener costo unitario del kardex (maneja ambos formatos)
                costo_u = 0
                if nombre in data.get("kardex", {}):
                    k = data["kardex"][nombre]
                    if isinstance(k, list) and k:
                        # Formato lista: tomar último movimiento
                        costo_u = k[-1].get("costo", 0)
                    elif isinstance(k, dict):
                        # Formato diccionario
                        costo_u = k.get("costo_unitario", 0)

                subtotal = cantidad * precio
                costo_total = cantidad * costo_u

                lineas.append(
                    {
                        "producto": nombre,
                        "cantidad": cantidad,
                        "precio_unitario": precio,
                        "subtotal": subtotal,
                        "costo_unitario": costo_u,
                        "costo_total": costo_total,
                    }
                )

                total_venta += subtotal
                total_costo += costo_total

                # Actualizar kardex
                if nombre in data.get("kardex", {}):
                    k = data["kardex"][nombre]

                    if isinstance(k, list):
                        # Formato lista: agregar movimiento
                        saldo_anterior = k[-1]["saldo"] if k else 0
                        nuevo_saldo = saldo_anterior - cantidad

                        k.append(
                            {
                                "fecha": fecha,
                                "tipo": "salida",
                                "cantidad": cantidad,
                                "costo": costo_u,
                                "total": costo_total,
                                "saldo": nuevo_saldo,
                                "descripcion": f"Venta POS - {cliente}",
                            }
                        )
                    elif isinstance(k, dict):
                        # Formato diccionario antiguo
                        if "movimientos" not in k:
                            k["movimientos"] = []

                        k["movimientos"].append(
                            {
                                "fecha": fecha,
                                "tipo": "Salida",
                                "cantidad": cantidad,
                                "costo_unitario": costo_u,
                                "descripcion": f"Venta POS - {cliente}",
                            }
                        )

                        saldo = k.get("saldo_inicial", 0)
                        for m in k["movimientos"]:
                            if m["tipo"] == "Entrada":
                                saldo += m["cantidad"]
                            else:
                                saldo -= m["cantidad"]
                        k["saldo_actual"] = saldo

        if not lineas:
            return redirect(url_for("pos"))

        # Crear asiento diario
        if "diario" not in data:
            data["diario"] = []

        diario_id = get_next_id(data, "diario")
        cuenta_cobro = (
            "1001"
            if forma_pago == "Efectivo"
            else ("1002" if forma_pago == "Banco" else "1003")
        )

        movimientos_diario = [
            {"cuenta": cuenta_cobro, "tipo": "Debe", "monto": total_venta},
            {"cuenta": "4001", "tipo": "Haber", "monto": total_venta},
        ]

        if total_costo > 0:
            movimientos_diario.append(
                {"cuenta": "5001", "tipo": "Debe", "monto": total_costo}
            )
            movimientos_diario.append(
                {"cuenta": "1004", "tipo": "Haber", "monto": total_costo}
            )

        ref_pos = f"POS-{len(data.get('pos_historial', [])) + 1:04d}"

        entry = {
            "id": diario_id,
            "fecha": fecha,
            "descripcion": f"Venta POS - {cliente} ({forma_pago})",
            "ref": ref_pos,
            "movimientos": movimientos_diario,
        }
        data["diario"].append(entry)

        # Auxiliar de caja
        if cuenta_cobro in ["1001", "1002"]:
            if "caja" not in data:
                data["caja"] = []

            data["caja"].append(
                {
                    "id": get_next_id(data, "caja"),
                    "fecha": fecha,
                    "descripcion": f"Venta POS - {cliente}",
                    "tipo": "Debe",
                    "monto": total_venta,
                    "cuenta": cuenta_cobro,
                    "ref_diario": diario_id,
                }
            )

        # Cuentas por cobrar
        if cuenta_cobro == "1003":
            if "cuentas_cobrar" not in data:
                data["cuentas_cobrar"] = []

            data["cuentas_cobrar"].append(
                {
                    "id": get_next_id(data, "cuentas_cobrar"),
                    "fecha": fecha,
                    "descripcion": f"Venta POS - {cliente}",
                    "monto": total_venta,
                    "estado": "Pendiente",
                    "ref_diario": diario_id,
                }
            )

        # Guardar en historial POS
        if "pos_historial" not in data:
            data["pos_historial"] = []

        data["pos_historial"].append(
            {
                "ref": ref_pos,
                "fecha": fecha,
                "cliente": cliente,
                "forma_pago": forma_pago,
                "lineas": lineas,
                "total": total_venta,
                "costo": total_costo,
                "utilidad": total_venta - total_costo,
            }
        )

        save_data(data)
        return redirect(url_for("pos"))

    except Exception as e:
        print(f"ERROR EN /pos/venta: {e}")
        import traceback

        traceback.print_exc()
        return redirect(url_for("pos"))


# ══════════════════════════════════════════════════════════════
#  EXPORTAR KARDEX PEPS
# ══════════════════════════════════════════════════════════════


@app.route("/kardex/exportar_peps", methods=["POST"])
def exportar_kardex_peps():
    """Exporta el Kardex completo con método PEPS a Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl no está instalado"}), 500

    data = load_data()
    producto_codigo = request.form.get("producto", None)

    # Generar reporte PEPS
    reporte = kardex_peps.generar_reporte_kardex_peps(data, producto_codigo)

    wb = Workbook()

    for idx, (codigo, info) in enumerate(reporte.items()):
        # Crear hoja para cada producto
        if idx == 0:
            ws = wb.active
            ws.title = codigo[:31]  # Excel limita a 31 caracteres
        else:
            ws = wb.create_sheet(codigo[:31])

        ws.sheet_properties.tabColor = "4F8CFF"

        # ── ENCABEZADO ──
        ws.append(["Colectivo FG - Joyería y Muchas Más"])
        ws.merge_cells("A1:I1")
        ws["A1"].font = Font(size=16, bold=True, color="1A1D2E")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.append([f"KARDEX - {info['nombre']}"])
        ws.merge_cells("A2:I2")
        ws["A2"].font = Font(size=14, bold=True, color="4F8CFF")
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.append([f"Código: {codigo}"])
        ws.merge_cells("A3:I3")
        ws["A3"].font = Font(size=10, color="7a7f99")
        ws["A3"].alignment = Alignment(horizontal="center")

        ws.append([f"Generado: {date.today().isoformat()}"])
        ws.merge_cells("A4:I4")
        ws["A4"].font = Font(size=10, color="7a7f99")
        ws["A4"].alignment = Alignment(horizontal="center")

        ws.append([])  # Línea vacía

        # ── RESUMEN ──
        ws.append(["RESUMEN"])
        ws["A6"].font = Font(size=12, bold=True)

        ws.append(["Stock Actual:", info["stock"], "unidades"])
        ws.append(["Costo Promedio:", f"C$ {info['costo_promedio']:.2f}", "por unidad"])
        ws.append(["Valor Inventario:", f"C$ {info['valor_inventario']:.2f}", ""])

        ws.append([])  # Línea vacía

        # ── LOTES DISPONIBLES (PEPS) ──
        ws.append(
            ["LOTES DISPONIBLES (Método PEPS - Primeras Entradas, Primeras Salidas)"]
        )
        ws["A11"].font = Font(size=12, bold=True, color="2ECC71")
        ws.merge_cells("A11:E11")

        headers_lotes = [
            "Fecha Entrada",
            "Cant. Original",
            "Cant. Disponible",
            "Costo Unit.",
            "Valor Total",
        ]
        ws.append(headers_lotes)

        for cell in ws[12]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="2ECC71", end_color="2ECC71", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        if info["lotes"]:
            for lote in info["lotes"]:
                ws.append(
                    [
                        lote["fecha"],
                        lote["cantidad_original"],
                        lote["cantidad_disponible"],
                        lote["costo_unitario"],
                        lote["valor_total"],
                    ]
                )

                # Formato moneda
                last_row = ws.max_row
                ws.cell(last_row, 4).number_format = "#,##0.00"
                ws.cell(last_row, 5).number_format = "#,##0.00"
        else:
            ws.append(["Sin lotes disponibles", "", "", "", ""])

        ws.append([])  # Línea vacía

        # ── MOVIMIENTOS HISTÓRICOS ──
        ws.append(["HISTORIAL DE MOVIMIENTOS"])
        ws[f"A{ws.max_row}"].font = Font(size=12, bold=True, color="E74C3C")
        ws.merge_cells(f"A{ws.max_row}:I{ws.max_row}")

        headers_mov = [
            "Fecha",
            "Tipo",
            "Cantidad",
            "Costo Unit.",
            "Total",
            "Saldo",
            "Descripción",
        ]
        ws.append(headers_mov)

        header_row_mov = ws.max_row
        for cell in ws[header_row_mov]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="E74C3C", end_color="E74C3C", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        if info["movimientos"]:
            for mov in reversed(info["movimientos"]):  # Más reciente primero
                tipo_emoji = (
                    "📥"
                    if mov["tipo"] == "entrada"
                    else "📤"
                    if mov["tipo"] == "salida"
                    else "🔄"
                )
                ws.append(
                    [
                        mov["fecha"],
                        f"{tipo_emoji} {mov['tipo'].title()}",
                        mov["cantidad"],
                        mov.get("costo", 0),
                        mov.get("total", 0),
                        mov["saldo"],
                        mov.get("descripcion", ""),
                    ]
                )

                # Formato
                last_row = ws.max_row
                ws.cell(last_row, 4).number_format = "#,##0.00"
                ws.cell(last_row, 5).number_format = "#,##0.00"

                # Color según tipo
                if mov["tipo"] == "entrada":
                    ws.cell(last_row, 2).font = Font(color="2ECC71")
                elif mov["tipo"] == "salida":
                    ws.cell(last_row, 2).font = Font(color="E74C3C")
        else:
            ws.append(["Sin movimientos registrados", "", "", "", "", "", ""])

        # Ajustar anchos
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 40

    # ── Enviar como descarga ──
    from io import BytesIO
    from flask import send_file

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Kardex_PEPS_{timestamp}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ── PLANTILLA DE IMPORTACIÓN ──
@app.route("/kardex/detectar_hojas", methods=["POST"])
def detectar_hojas():
    """Recibe un Excel y devuelve sus hojas con una vista previa de columnas."""
    from io import BytesIO

    archivo = request.files.get("archivo")
    if not archivo or archivo.filename == "":
        return jsonify({"ok": False, "error": "No se recibió archivo"})

    ext = archivo.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        # CSV no tiene hojas — devolver hoja única virtual
        return jsonify(
            {
                "ok": True,
                "tipo": "csv",
                "hojas": [{"nombre": "Hoja única (CSV)", "filas": 0, "columnas": []}],
            }
        )

    try:
        import openpyxl

        contenido = archivo.read()
        wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True, read_only=True)
        hojas = []
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            # Leer solo las primeras filas para preview
            headers = []
            filas_con_datos = 0
            primera = True
            for row in ws.iter_rows(max_row=200, values_only=True):
                if any(c for c in row if c is not None):
                    if primera:
                        headers = [str(c).strip() if c is not None else "" for c in row]
                        primera = False
                    else:
                        filas_con_datos += 1
            hojas.append(
                {
                    "nombre": nombre_hoja,
                    "filas": filas_con_datos,
                    "columnas": headers[:8],  # max 8 para mostrar
                }
            )
        wb.close()
        # Guardar contenido en sesión temporal (archivo en memoria con nombre único)
        import base64, time

        token = str(int(time.time() * 1000))
        # Guardar en un dict global temporal (simple, sin Redis)
        _archivos_temp[token] = {
            "contenido": contenido,
            "filename": archivo.filename,
            "ext": ext,
        }
        return jsonify({"ok": True, "tipo": "xlsx", "hojas": hojas, "token": token})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/kardex/plantilla")
def descargar_plantilla_inventario():
    """Genera y descarga un .xlsx de ejemplo para importar inventario."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"

        # Encabezados
        headers = ["producto", "cantidad", "costo", "fecha", "descripcion"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F8CFF", end_color="4F8CFF", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Filas de ejemplo
        ejemplos = [
            ["Collar de Plata", 50, 120.00, "2026-01-15", "Stock inicial"],
            ["Aretes Dorados", 30, 85.50, "2026-01-15", "Stock inicial"],
            ["Pulsera de Oro", 20, 250.00, "2026-01-20", "Compra proveedor"],
            ["Anillo de Plata", 15, 95.00, "", ""],
        ]
        for fila in ejemplos:
            ws.append(fila)

        # Anchos
        for col, w in zip("ABCDE", [25, 12, 12, 14, 28]):
            ws.column_dimensions[chr(64 + list("ABCDE").index(col) + 1)].width = w

        # Hoja de instrucciones
        ws2 = wb.create_sheet("Instrucciones")
        instrucciones = [
            ("📋 INSTRUCCIONES DE IMPORTACIÓN", True),
            ("", False),
            ("1. Llena la hoja 'Inventario' con tus productos.", False),
            ("2. La columna 'producto' es obligatoria.", False),
            ("3. La columna 'cantidad' debe ser un número entero positivo.", False),
            ("4. La columna 'costo' es el costo unitario en C$.", False),
            ("5. La columna 'fecha' es opcional (formato YYYY-MM-DD).", False),
            ("   Si se omite, se usa la fecha actual.", False),
            ("6. La columna 'descripcion' es opcional.", False),
            ("", False),
            ("⚠️  No borres la fila de encabezados (fila 1).", False),
            (
                "⚠️  Si el producto ya existe en el Kardex, se agrega como nueva entrada.",
                False,
            ),
            ("⚠️  Productos con cantidad 0 o negativa se ignoran.", False),
        ]
        for i, (texto, bold) in enumerate(instrucciones, 1):
            cell = ws2.cell(row=i, column=1, value=texto)
            cell.font = Font(bold=bold, size=11 if bold else 10)
        ws2.column_dimensions["A"].width = 70

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import send_file

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Plantilla_Importar_Inventario.xlsx",
        )
    except ImportError:
        return "openpyxl no está instalado", 500


# ── IMPORTAR INVENTARIO DESDE EXCEL / CSV ──
@app.route("/kardex/importar", methods=["POST"])
def importar_inventario():
    """
    Importa inventario desde Excel (.xlsx/.xls) o CSV.
    Acepta:
      - token + hoja_nombre: usa archivo previamente cargado en detectar_hojas
      - archivo directo: para CSV o flujo directo sin selección de hoja
    """
    from io import BytesIO
    from datetime import date as _date
    import urllib.parse

    token = request.form.get("token", "")
    hoja_nombre = request.form.get("hoja_nombre", "")
    ext = request.form.get("ext", "")
    contenido = None

    # ── Origen del archivo ────────────────────────────────────────
    if token and token in _archivos_temp:
        entrada = _archivos_temp.pop(token)  # consumir y limpiar
        contenido = entrada["contenido"]
        ext = entrada["ext"]
    else:
        archivo = request.files.get("archivo")
        if not archivo or archivo.filename == "":
            return redirect(url_for("kardex"))
        ext = archivo.filename.rsplit(".", 1)[-1].lower()
        contenido = archivo.read()

    filas = []
    headers_originales = []
    error_lectura = None

    # ── 1. LEER EL ARCHIVO ────────────────────────────────────────
    if ext in ("xlsx", "xls"):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
            # Seleccionar hoja: la pedida o la activa
            if hoja_nombre and hoja_nombre in wb.sheetnames:
                ws = wb[hoja_nombre]
            else:
                ws = wb.active

            primera_fila = None
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if any(c for c in row if c is not None):
                    primera_fila = row
                    break
            if primera_fila is None:
                error_lectura = f"La hoja '{ws.title}' está vacía."
            else:
                headers_originales = [
                    str(c).strip() if c is not None else "" for c in primera_fila
                ]
                headers_lower = [h.lower() for h in headers_originales]
                fila_inicio = None
                for idx, row in enumerate(
                    ws.iter_rows(min_row=1, values_only=True), start=1
                ):
                    if list(row) == list(primera_fila):
                        fila_inicio = idx + 1
                        break
                fila_inicio = fila_inicio or 2
                for row in ws.iter_rows(min_row=fila_inicio, values_only=True):
                    if any(c for c in row if c is not None):
                        filas.append(dict(zip(headers_lower, row)))
        except Exception as e:
            error_lectura = f"No se pudo leer el Excel: {e}"

    elif ext == "csv":
        try:
            import csv, io

            for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
                try:
                    text = contenido.decode(enc)
                    break
                except Exception:
                    text = contenido.decode("latin-1", errors="replace")
            primera_linea = text.split("\n")[0] if text else ""
            sep = (
                ";"
                if primera_linea.count(";") > primera_linea.count(",")
                else "\t"
                if primera_linea.count("\t") > primera_linea.count(",")
                else ","
            )
            reader = csv.DictReader(io.StringIO(text), delimiter=sep)
            headers_originales = reader.fieldnames or []
            for row in reader:
                filas.append({k.strip().lower(): v for k, v in row.items() if k})
        except Exception as e:
            error_lectura = f"No se pudo leer el CSV: {e}"
    else:
        error_lectura = f"Formato '{ext}' no soportado. Usa .xlsx o .csv"

    if error_lectura:
        return redirect(
            url_for("kardex") + "?error=" + urllib.parse.quote(error_lectura)
        )
    if not filas:
        return redirect(
            url_for("kardex")
            + "?error="
            + urllib.parse.quote(
                f"La hoja seleccionada no tiene datos. Columnas detectadas: {', '.join(headers_originales) or 'ninguna'}"
            )
        )

    # ── 2. MAPEAR COLUMNAS ────────────────────────────────────────
    NOMBRES_PRODUCTO = {
        "producto",
        "product",
        "nombre",
        "name",
        "articulo",
        "item",
        "descripcion_producto",
        "art",
        "codigo",
        "code",
        "referencia",
        "ref",
        "mercaderia",
        "mercancía",
        "mercancia",
    }
    NOMBRES_CANTIDAD = {
        "cantidad",
        "qty",
        "quantity",
        "unidades",
        "units",
        "stock",
        "existencia",
        "existencias",
        "cant",
        "inventario",
        "piezas",
        "pcs",
    }
    NOMBRES_COSTO = {
        "costo",
        "cost",
        "precio",
        "price",
        "costo_unitario",
        "unit_cost",
        "valor",
        "value",
        "costo_unit",
        "precio_unitario",
        "pu",
        "p.u.",
        "c/u",
    }
    NOMBRES_FECHA = {"fecha", "date", "fecha_entrada", "fecha_compra", "entry_date"}
    NOMBRES_DESC = {
        "descripcion",
        "description",
        "desc",
        "detalle",
        "detail",
        "nota",
        "observacion",
    }

    keys = list(filas[0].keys()) if filas else []

    def encontrar_col(keys, nombres_set):
        for k in keys:
            if k in nombres_set:
                return k
        for k in keys:
            for n in nombres_set:
                if n in k or k in n:
                    return k
        return None

    col_producto = encontrar_col(keys, NOMBRES_PRODUCTO)
    col_cantidad = encontrar_col(keys, NOMBRES_CANTIDAD)
    col_costo = encontrar_col(keys, NOMBRES_COSTO)
    col_fecha = encontrar_col(keys, NOMBRES_FECHA)
    col_desc = encontrar_col(keys, NOMBRES_DESC)

    modo_auto = False
    if not col_producto and len(keys) >= 1:
        col_producto = keys[0]
        modo_auto = True
    if not col_cantidad and len(keys) >= 2:
        col_cantidad = keys[1]
        modo_auto = True
    if not col_costo and len(keys) >= 3:
        col_costo = keys[2]

    if not col_producto:
        msg = (
            f"No se encontró columna de producto. "
            f"Columnas en tu archivo: {', '.join(headers_originales)}. "
            f"Renómbralas como: producto, cantidad, costo"
        )
        return redirect(url_for("kardex") + "?error=" + urllib.parse.quote(msg))

    # ── 3. PROCESAR FILAS ─────────────────────────────────────────
    data = load_data()
    importados = 0
    omitidos = 0
    hoy = _date.today().isoformat()

    data.setdefault("kardex", {})
    data.setdefault("kardex_peps", {})
    data.setdefault("productos", {})

    # Helper definido FUERA del loop para evitar closure bug de Python
    def _v(fila, col):
        if not col:
            return None
        val = fila.get(col)
        if val is None:
            return None
        s = str(val).strip()
        return s if s not in ("", "None", "nan", "NaN", "-") else None

    for fila in filas:
        producto = _v(fila, col_producto)
        if not producto:
            omitidos += 1
            continue

        cantidad = 0
        raw_cant = _v(fila, col_cantidad)
        if raw_cant:
            try:
                cantidad = int(
                    float(raw_cant.replace(",", "").replace(" ", "").replace("'", ""))
                )
            except Exception:
                cantidad = 0

        costo = 0.0
        raw_costo = _v(fila, col_costo)
        if raw_costo:
            try:
                costo = float(
                    raw_costo.replace(",", "")
                    .replace(" ", "")
                    .replace("C$", "")
                    .replace("$", "")
                    .strip()
                )
            except Exception:
                costo = 0.0

        fecha = _v(fila, col_fecha) or hoy
        if "/" in fecha:
            parts = fecha.split("/")
            try:
                if len(parts) == 3:
                    d, m, a = parts
                    if len(a) == 2:
                        a = "20" + a
                    fecha = f"{a}-{m.zfill(2)}-{d.zfill(2)}"
            except Exception:
                fecha = hoy

        descripcion = _v(fila, col_desc) or "Importación de inventario"

        if producto not in data["kardex"]:
            data["kardex"][producto] = []
            data["productos"][producto] = {
                "nombre": producto,
                "stock": 0,
                "costo_promedio": 0,
            }
            data["kardex_peps"][producto] = {
                "lotes": [],
                "stock_total": 0,
                "costo_promedio": 0,
            }

        if cantidad > 0:
            data = kardex_peps.agregar_entrada_peps(
                data, producto, fecha, cantidad, costo
            )
        importados += 1

    save_data(data)
    qs = f"importados={importados}&errores={omitidos}"
    if modo_auto:
        qs += "&auto=1"
    return redirect(url_for("kardex") + "?" + qs)


# ══════════════════════════════════════════════════════════════
#  KARDEX — CRUD: eliminar producto, editar/eliminar movimiento
# ══════════════════════════════════════════════════════════════


@app.route("/kardex/eliminar_producto", methods=["POST"])
def eliminar_producto_kardex():
    """Elimina un producto y todos sus movimientos del Kardex."""
    data = load_data()
    nombre = request.form.get("producto", "").strip()
    if nombre:
        data["kardex"].pop(nombre, None)
        data.get("kardex_peps", {}).pop(nombre, None)
        data.get("productos", {}).pop(nombre, None)
        save_data(data)
    return redirect(url_for("kardex"))


@app.route("/kardex/renombrar_producto", methods=["POST"])
def renombrar_producto_kardex():
    """Renombra un producto manteniendo todos sus movimientos."""
    data = load_data()
    nombre_old = request.form.get("nombre_old", "").strip()
    nombre_new = request.form.get("nombre_new", "").strip()
    if (
        nombre_old
        and nombre_new
        and nombre_old in data["kardex"]
        and nombre_new not in data["kardex"]
    ):
        for seccion in ("kardex", "kardex_peps", "productos"):
            d = data.get(seccion, {})
            if nombre_old in d:
                d[nombre_new] = d.pop(nombre_old)
        save_data(data)
    return redirect(url_for("kardex"))


@app.route("/kardex/editar_movimiento", methods=["POST"])
def editar_movimiento_kardex():
    """Edita un movimiento individual del Kardex (fecha, descripcion, cantidad, costo)."""
    data = load_data()
    producto = request.form.get("producto", "").strip()
    idx = int(request.form.get("idx", -1))
    k = data["kardex"].get(producto, [])
    if 0 <= idx < len(k):
        mov = k[idx]
        mov["fecha"] = request.form.get("fecha", mov.get("fecha", ""))
        mov["descripcion"] = request.form.get("descripcion", mov.get("descripcion", ""))
        try:
            nueva_cant = int(
                float(request.form.get("cantidad", mov.get("cantidad", 0)))
            )
            nuevo_costo = float(request.form.get("costo", mov.get("costo", 0)))
            mov["cantidad"] = nueva_cant
            mov["costo"] = nuevo_costo
            mov["total"] = nueva_cant * nuevo_costo
        except (ValueError, TypeError):
            pass
        # Recalcular saldos acumulados
        saldo = 0
        for m in k:
            if m.get("tipo") == "entrada":
                saldo += m.get("cantidad", 0)
            elif m.get("tipo") == "salida":
                saldo -= m.get("cantidad", 0)
            m["saldo"] = saldo
        data["kardex"][producto] = k
        # Recalcular PEPS desde cero para este producto
        data = _reconstruir_peps(data, producto)
        save_data(data)
    return redirect(url_for("kardex"))


@app.route("/kardex/eliminar_movimiento", methods=["POST"])
def eliminar_movimiento_kardex():
    """Elimina un movimiento individual y recalcula saldos."""
    data = load_data()
    producto = request.form.get("producto", "").strip()
    idx = int(request.form.get("idx", -1))
    k = data["kardex"].get(producto, [])
    if 0 <= idx < len(k):
        k.pop(idx)
        # Recalcular saldos
        saldo = 0
        for m in k:
            if m.get("tipo") == "entrada":
                saldo += m.get("cantidad", 0)
            elif m.get("tipo") == "salida":
                saldo -= m.get("cantidad", 0)
            m["saldo"] = saldo
        data["kardex"][producto] = k
        data = _reconstruir_peps(data, producto)
        save_data(data)
    return redirect(url_for("kardex"))


def _reconstruir_peps(data, producto):
    """
    Reconstruye los lotes PEPS de un producto a partir de su historial kardex.
    Procesa entradas, salidas y ajustes de forma correcta.
    """
    try:
        movs = data["kardex"].get(producto, [])
        
        # Inicializar/limpiar kardex_peps
        if "kardex_peps" not in data:
            data["kardex_peps"] = {}
        data["kardex_peps"][producto] = {
            "lotes": [],
            "stock_total": 0,
            "costo_promedio": 0,
        }
        
        # Procesar cada movimiento en orden cronológico
        for m in movs:
            if m.get("tipo") == "entrada" and m.get("cantidad", 0) > 0:
                # Agregar entrada
                data = kardex_peps.agregar_entrada_peps(
                    data,
                    producto,
                    m.get("fecha", ""),
                    m.get("cantidad", 0),
                    m.get("costo", 0),
                )
            elif m.get("tipo") == "salida" and m.get("cantidad", 0) > 0:
                # Procesar salida PEPS
                try:
                    data, _, _ = kardex_peps.procesar_salida_peps(
                        data, producto, m.get("fecha", ""), m.get("cantidad", 0)
                    )
                except Exception as e:
                    # Si hay error (stock insuficiente), continuar
                    pass
            elif m.get("tipo") == "ajuste" and m.get("cantidad", 0) > 0:
                # Para ajustes, agregar como entrada
                data = kardex_peps.agregar_entrada_peps(
                    data,
                    producto,
                    m.get("fecha", ""),
                    m.get("cantidad", 0),
                    m.get("costo", 0),
                )
    except Exception as e:
        # En caso de error, asegurar que kardex_peps existe
        if "kardex_peps" not in data:
            data["kardex_peps"] = {}
        if producto not in data["kardex_peps"]:
            data["kardex_peps"][producto] = {
                "lotes": [],
                "stock_total": 0,
                "costo_promedio": 0,
            }
    
    return data


# ── APAGAR SERVIDOR ──
@app.route("/apagar", methods=["POST"])
def apagar():
    import threading
    from flask import make_response

    def shutdown():
        import time

        time.sleep(0.6)
        try:
            os.kill(os.getpid(), signal.SIGTERM)
        except Exception:
            os._exit(0)

    t = threading.Thread(target=shutdown, daemon=True)
    t.start()

    resp = make_response(render_template("apagar.html"))
    resp.headers["Connection"] = "close"
    return resp


# ══════════════════════════════════════════════════════════════
# RUTAS API PARA GESTIÓN DE BASE DE DATOS
# ══════════════════════════════════════════════════════════════


@app.route("/api/backup_json")
def api_backup_json():
    """Exporta los datos actuales como archivo JSON"""
    data = load_data()
    from flask import make_response
    import json

    json_str = json.dumps(data, indent=2, default=str)
    resp = make_response(json_str)
    resp.headers["Content-Disposition"] = "attachment; filename=fg_backup.json"
    resp.headers["Content-Type"] = "application/json"
    return resp


@app.route("/api/copia_db")
def api_copia_db():
    """Crea una copia de la base de datos SQLite"""
    import shutil
    from flask import make_response
    from io import BytesIO

    db_path = db_internal.get_db_path()
    if os.path.exists(db_path):
        with open(db_path, "rb") as f:
            data = f.read()
        resp = make_response(data)
        resp.headers["Content-Disposition"] = "attachment; filename=fg_db_backup.db"
        resp.headers["Content-Type"] = "application/octet-stream"
        return resp
    return jsonify({"ok": False, "error": "Base de datos no encontrada"}), 404


@app.route("/api/importar_db", methods=["POST"])
def api_importar_db():
    """Importa una base de datos SQLite"""
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"ok": False, "error": "No se recibió archivo"})

    db_path = db_internal.get_db_path()
    try:
        archivo.save(db_path)
        return jsonify({"ok": True, "message": "Base de datos importada"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/ver_db")
def api_ver_db():
    """Muestra el contenido de la base de datos como HTML"""
    import sqlite3

    db_path = db_internal.get_db_path()

    html = "<html><head><title>Ver DB</title><style>"
    html += "body{font-family:monospace;background:#1a1d2e;color:#e2e4eb;padding:20px}"
    html += "table{border-collapse:collapse;width:100%}"
    html += "th,td{border:1px solid #2a2d42;padding:8px;text-align:left}"
    html += "th{background:#2a2d42}"
    html += "</style></head><body>"
    html += "<h1>Contenido de Base de Datos</h1>"

    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tablas = cursor.fetchall()

        for tabla in tablas:
            nombre_tabla = tabla[0]
            html += f"<h2>Tabla: {nombre_tabla}</h2>"
            cursor.execute(f"SELECT * FROM {nombre_tabla} LIMIT 100")
            rows = cursor.fetchall()

            if rows:
                html += "<table><tr>"
                # Headers
                for col in rows[0].keys():
                    html += f"<th>{col}</th>"
                html += "</tr>"

                # Datos
                for row in rows:
                    html += "<tr>"
                    for col in row:
                        html += f"<td>{col}</td>"
                    html += "</tr>"
                html += "</table>"
            else:
                html += "<p>Sin datos</p>"

        conn.close()
    except Exception as e:
        html += f"<p>Error: {e}</p>"

    html += "</body></html>"
    return html


@app.route("/api/restaurar_json", methods=["POST"])
def api_restaurar_json():
    """Restaura los datos desde un archivo JSON"""
    try:
        data_json = request.get_data(as_text=True)
        data = json.loads(data_json)

        # Validar estructura básica
        if "cuentas" not in data:
            return jsonify({"ok": False, "error": "Archivo JSON inválido"}), 400

        # Guardar en DB
        save_data(data)
        return jsonify({"ok": True, "message": "Datos restaurados correctamente"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/borrar_todo", methods=["POST"])
def api_borrar_todo():
    """Borra todos los datos (reset completo)"""
    try:
        data_vacio = empty_data()
        save_data(data_vacio)
        return jsonify({"ok": True, "message": "Todos los datos han sido borrados"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── NUEVA CUENTA ──
@app.route("/cuentas/nueva", methods=["POST"])
def nueva_cuenta():
    data = load_data()
    codigo = request.form.get("codigo", "").strip()
    nombre = request.form.get("nombre", "").strip()
    tipo = request.form.get("tipo", "Activo")
    if codigo and nombre and codigo not in data["cuentas"]:
        data["cuentas"][codigo] = {"nombre": nombre, "tipo": tipo, "saldo": 0}
        save_data(data)
    return redirect(request.referrer or url_for("index"))


# ── REPORTES ──
@app.route("/reportes")
def reportes():
    data = load_data()
    # Recopilar estadísticas para la pantalla
    mayor = calcular_mayor(data)
    _, td, th, tsd, tsh = calcular_balanza(data)
    di, ti, dg, tg, util = calcular_estado_resultados(data)
    activos, ta, pasivos, tp, capital, tc = calcular_balance_general(data)
    return render_template(
        "reportes.html",
        diario=data["diario"],
        cuentas=data["cuentas"],
        mayor=mayor,
        balanza_td=td,
        balanza_th=th,
        balanza_tsd=tsd,
        balanza_tsh=tsh,
        detalle_ingresos=di,
        total_ingresos=ti,
        detalle_gastos=dg,
        total_gastos=tg,
        utilidad=util,
        activos=activos,
        total_activo=ta,
        pasivos=pasivos,
        total_pasivo=tp,
        capital=capital,
        total_capital=tc,
        kardex=data["kardex"],
        pos_historial=data.get("pos_historial", []),
    )


@app.route("/reportes/exportar", methods=["POST"])
def exportar_reporte():
    """Genera un .xlsx con todas las hojas y lo guarda en reportes/"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess, sys

        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"]
        )
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    data = load_data()
    tipo_reporte = request.form.get("tipo", "todos")

    # ── Estilos ──
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1D2E")
    accent_fill = PatternFill("solid", fgColor="4F8CFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1A1D2E")
    sub_font = Font(name="Calibri", size=10, color="7A7F99")
    money_fmt = "#,##0.00"
    thin_border = Border(
        left=Side(style="thin", color="D0D3E0"),
        right=Side(style="thin", color="D0D3E0"),
        top=Side(style="thin", color="D0D3E0"),
        bottom=Side(style="thin", color="D0D3E0"),
    )
    total_fill = PatternFill("solid", fgColor="E8EAEF")
    total_font = Font(name="Calibri", size=11, bold=True, color="1A1D2E")
    green_font = Font(name="Calibri", size=11, color="27AE60")
    red_font = Font(name="Calibri", size=11, color="E74C3C")

    def style_header_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def style_data_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10, color="333333")

    def style_total_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = Border(
                left=Side(style="thin", color="D0D3E0"),
                right=Side(style="thin", color="D0D3E0"),
                top=Side(style="medium", color="4F8CFF"),
                bottom=Side(style="thin", color="D0D3E0"),
            )

    def add_title(ws, title, subtitle=""):
        ws.cell(row=1, column=1, value="F & G — Sistema Contable").font = Font(
            name="Calibri", size=12, bold=True, color="4F8CFF"
        )
        ws.cell(row=2, column=1, value=title).font = title_font
        if subtitle:
            ws.cell(row=3, column=1, value=subtitle).font = sub_font
        ws.cell(
            row=3 if not subtitle else 4,
            column=1,
            value=f"Fecha de exportación: {date.today().isoformat()}",
        ).font = sub_font
        return 5  # primera fila de datos

    wb = Workbook()
    wb.remove(wb.active)  # borrar hoja por defecto

    mayor = calcular_mayor(data)
    cuentas = data["cuentas"]

    # ══════════════ HOJA 1: LIBRO DIARIO ══════════════
    if tipo_reporte in ["todos", "diario"]:
        ws = wb.create_sheet("Libro Diario")
        ws.sheet_properties.tabColor = "4F8CFF"
        start = add_title(ws, "Libro Diario", "Registro cronológico de asientos")
        headers = [
            "#",
            "Fecha",
            "Descripción",
            "Ref",
            "Cuenta",
            "Nombre Cuenta",
            "Debe C$",
            "Haber C$",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=start, column=i, value=h)
        style_header_row(ws, start, len(headers))
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 38
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 28
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 16

        r = start + 1
        total_debe = total_haber = 0
        for entry in data["diario"]:
            for idx, mov in enumerate(entry["movimientos"]):
                ws.cell(row=r, column=1, value=entry["id"] if idx == 0 else "")
                ws.cell(row=r, column=2, value=entry["fecha"] if idx == 0 else "")
                ws.cell(row=r, column=3, value=entry["descripcion"] if idx == 0 else "")
                ws.cell(row=r, column=4, value=entry.get("ref", "") if idx == 0 else "")
                ws.cell(row=r, column=5, value=mov["cuenta"])
                ws.cell(
                    row=r,
                    column=6,
                    value=cuentas.get(mov["cuenta"], {}).get("nombre", mov["cuenta"]),
                )
                debe = mov["monto"] if mov["tipo"] == "Debe" else 0
                haber = mov["monto"] if mov["tipo"] == "Haber" else 0
                ws.cell(row=r, column=7, value=debe).number_format = money_fmt
                ws.cell(row=r, column=8, value=haber).number_format = money_fmt
                if debe:
                    ws.cell(row=r, column=7).font = green_font
                if haber:
                    ws.cell(row=r, column=8).font = red_font
                total_debe += debe
                total_haber += haber
                style_data_row(ws, r, len(headers))
                r += 1
        # Total
        ws.cell(row=r, column=6, value="TOTALES")
        ws.cell(row=r, column=7, value=total_debe).number_format = money_fmt
        ws.cell(row=r, column=8, value=total_haber).number_format = money_fmt
        style_total_row(ws, r, len(headers))

    # ══════════════ HOJA 2: LIBRO MAYOR ══════════════
    if tipo_reporte in ["todos", "mayor"]:
        ws = wb.create_sheet("Libro Mayor")
        ws.sheet_properties.tabColor = "7C5CFC"
        start = add_title(ws, "Libro Mayor", "Cuentas T — Movimientos por cuenta")
        headers = ["Cuenta", "Nombre", "Tipo", "Debe C$", "Haber C$", "Saldo C$"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=start, column=i, value=h)
        style_header_row(ws, start, len(headers))
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 16

        r = start + 1
        for codigo in sorted(cuentas.keys()):
            info = cuentas[codigo]
            debe = mayor[codigo]["debe"] if codigo in mayor else 0
            haber = mayor[codigo]["haber"] if codigo in mayor else 0
            ts = tipo_saldo(info["tipo"])
            saldo = (debe - haber) if ts == "Debe" else (haber - debe)
            ws.cell(row=r, column=1, value=codigo)
            ws.cell(row=r, column=2, value=info["nombre"])
            ws.cell(row=r, column=3, value=info["tipo"])
            ws.cell(row=r, column=4, value=debe).number_format = money_fmt
            ws.cell(row=r, column=5, value=haber).number_format = money_fmt
            ws.cell(row=r, column=6, value=saldo).number_format = money_fmt
            ws.cell(row=r, column=6).font = green_font if saldo >= 0 else red_font
            style_data_row(ws, r, len(headers))
            r += 1

    # ══════════════ HOJA 3: BALANZA DE COMPROBACIÓN ══════════════
    if tipo_reporte in ["todos", "balanza"]:
        ws = wb.create_sheet("Balanza")
        ws.sheet_properties.tabColor = "F39C12"
        start = add_title(
            ws, "Balanza de Comprobación", "Verificación de débitos y créditos"
        )
        headers = [
            "Código",
            "Cuenta",
            "Tipo",
            "Debe C$",
            "Haber C$",
            "Saldo Debe C$",
            "Saldo Haber C$",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=start, column=i, value=h)
        style_header_row(ws, start, len(headers))
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

        balanza_data, td, th, tsd, tsh = calcular_balanza(data)
        r = start + 1
        for item in balanza_data:
            ws.cell(row=r, column=1, value=item["codigo"])
            ws.cell(row=r, column=2, value=item["nombre"])
            ws.cell(row=r, column=3, value=item["tipo"])
            ws.cell(row=r, column=4, value=item["debe"]).number_format = money_fmt
            ws.cell(row=r, column=5, value=item["haber"]).number_format = money_fmt
            ws.cell(row=r, column=6, value=item["saldo_debe"]).number_format = money_fmt
            ws.cell(
                row=r, column=7, value=item["saldo_haber"]
            ).number_format = money_fmt
            style_data_row(ws, r, 7)
            r += 1
        # Totales
        ws.cell(row=r, column=2, value="TOTALES")
        ws.cell(row=r, column=4, value=td).number_format = money_fmt
        ws.cell(row=r, column=5, value=th).number_format = money_fmt
        ws.cell(row=r, column=6, value=tsd).number_format = money_fmt
        ws.cell(row=r, column=7, value=tsh).number_format = money_fmt
        style_total_row(ws, r, 7)

    # ══════════════ HOJA 4: ESTADO DE RESULTADOS ══════════════
    if tipo_reporte in ["todos", "estado_resultados"]:
        ws = wb.create_sheet("Estado de Resultados")
        ws.sheet_properties.tabColor = "2ECC71"
        start = add_title(
            ws, "Estado de Resultados", "Ingresos, Gastos y Utilidad Neta"
        )
        di, ti, dg, tg, util = calcular_estado_resultados(data)

        r = start
        ws.cell(row=r, column=1, value="INGRESOS").font = Font(
            name="Calibri", size=11, bold=True, color="27AE60"
        )
        r += 1
        for item in di:
            ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                name="Calibri", size=10, color="333333"
            )
            ws.cell(row=r, column=2, value=item["monto"]).number_format = money_fmt
            ws.cell(row=r, column=2).font = green_font
            r += 1
        ws.cell(row=r, column=1, value="Total Ingresos").font = total_font
        ws.cell(row=r, column=2, value=ti).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="27AE60"
        )
        style_total_row(ws, r, 2)
        r += 2

        ws.cell(row=r, column=1, value="GASTOS").font = Font(
            name="Calibri", size=11, bold=True, color="E74C3C"
        )
        r += 1
        for item in dg:
            ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                name="Calibri", size=10, color="333333"
            )
            ws.cell(row=r, column=2, value=item["monto"]).number_format = money_fmt
            ws.cell(row=r, column=2).font = red_font
            r += 1
        ws.cell(row=r, column=1, value="Total Gastos").font = total_font
        ws.cell(row=r, column=2, value=tg).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="E74C3C"
        )
        style_total_row(ws, r, 2)
        r += 2

        ws.cell(row=r, column=1, value="UTILIDAD NETA").font = Font(
            name="Calibri", size=13, bold=True, color="1A1D2E"
        )
        ws.cell(row=r, column=2, value=util).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri",
            size=13,
            bold=True,
            color="27AE60" if util >= 0 else "E74C3C",
        )
        style_total_row(ws, r, 2)

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 20

    # ══════════════ HOJA 5: BALANCE GENERAL ══════════════
    if tipo_reporte in ["todos", "balance_general"]:
        ws = wb.create_sheet("Balance General")
        ws.sheet_properties.tabColor = "E74C3C"
        start = add_title(ws, "Balance General", "Activos = Pasivos + Capital")
        activos, ta, pasivos, tp, capital_items, tc = calcular_balance_general(data)

        r = start
        ws.cell(row=r, column=1, value="ACTIVOS").font = Font(
            name="Calibri", size=11, bold=True, color="4F8CFF"
        )
        r += 1
        for item in activos:
            ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                name="Calibri", size=10, color="333333"
            )
            ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
            ws.cell(row=r, column=2).font = (
                green_font if item["saldo"] >= 0 else red_font
            )
            r += 1
        ws.cell(row=r, column=1, value="Total Activos").font = total_font
        ws.cell(row=r, column=2, value=ta).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="4F8CFF"
        )
        style_total_row(ws, r, 2)
        r += 2

        ws.cell(row=r, column=1, value="PASIVOS").font = Font(
            name="Calibri", size=11, bold=True, color="E74C3C"
        )
        r += 1
        for item in pasivos:
            ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                name="Calibri", size=10, color="333333"
            )
            ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
            ws.cell(row=r, column=2).font = red_font
            r += 1
        ws.cell(row=r, column=1, value="Total Pasivos").font = total_font
        ws.cell(row=r, column=2, value=tp).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="E74C3C"
        )
        style_total_row(ws, r, 2)
        r += 2

        ws.cell(row=r, column=1, value="CAPITAL").font = Font(
            name="Calibri", size=11, bold=True, color="2ECC71"
        )
        r += 1
        for item in capital_items:
            ws.cell(row=r, column=1, value=item["nombre"]).font = Font(
                name="Calibri", size=10, color="333333"
            )
            ws.cell(row=r, column=2, value=item["saldo"]).number_format = money_fmt
            ws.cell(row=r, column=2).font = (
                green_font if item["saldo"] >= 0 else red_font
            )
            r += 1
        ws.cell(row=r, column=1, value="Total Capital").font = total_font
        ws.cell(row=r, column=2, value=tc).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="2ECC71"
        )
        style_total_row(ws, r, 2)
        r += 2

        ws.cell(row=r, column=1, value="VERIFICACIÓN: Pasivos + Capital").font = Font(
            name="Calibri", size=11, bold=True, color="333333"
        )
        ws.cell(row=r, column=2, value=tp + tc).number_format = money_fmt
        ws.cell(row=r, column=2).font = Font(
            name="Calibri", size=11, bold=True, color="4F8CFF"
        )

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 20

    # ══════════════ HOJA 6: VENTAS POS ══════════════
    if tipo_reporte in ["todos"]:
        pos_historial = data.get("pos_historial", [])
        if pos_historial:
            ws = wb.create_sheet("Ventas POS")
            ws.sheet_properties.tabColor = "F39C12"
            start = add_title(
                ws, "Ventas POS", "Historial de ventas registradas en el punto de venta"
            )
            headers = [
                "Ref",
                "Fecha",
                "Cliente",
                "Forma Pago",
                "Productos",
                "Total C$",
                "Costo C$",
                "Utilidad C$",
            ]
            for i, h in enumerate(headers, 1):
                ws.cell(row=start, column=i, value=h)
            style_header_row(ws, start, len(headers))
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 40
            ws.column_dimensions["F"].width = 16
            ws.column_dimensions["G"].width = 16
            ws.column_dimensions["H"].width = 16
            r = start + 1
            total_ventas = total_costos = total_utilidades = 0
            for venta in pos_historial:
                productos_str = ", ".join(
                    f"{l.get('nombre', l.get('producto', '?'))} x{l.get('cantidad', 1)}"
                    for l in venta.get("lineas", [])
                )
                total_v = venta.get("total", venta.get("total_venta", 0))
                total_c = venta.get("costo", 0)
                utilidad = venta.get("utilidad", total_v - total_c)
                ws.cell(row=r, column=1, value=venta.get("ref", ""))
                ws.cell(row=r, column=2, value=venta.get("fecha", ""))
                ws.cell(row=r, column=3, value=venta.get("cliente", ""))
                ws.cell(row=r, column=4, value=venta.get("forma_pago", ""))
                ws.cell(row=r, column=5, value=productos_str)
                ws.cell(row=r, column=6, value=total_v).number_format = money_fmt
                ws.cell(row=r, column=6).font = green_font
                ws.cell(row=r, column=7, value=total_c).number_format = money_fmt
                ws.cell(row=r, column=8, value=utilidad).number_format = money_fmt
                ws.cell(row=r, column=8).font = (
                    green_font if utilidad >= 0 else red_font
                )
                style_data_row(ws, r, len(headers))
                total_ventas += total_v
                total_costos += total_c
                total_utilidades += utilidad
                r += 1
            ws.cell(row=r, column=5, value="TOTALES")
            ws.cell(row=r, column=6, value=total_ventas).number_format = money_fmt
            ws.cell(row=r, column=7, value=total_costos).number_format = money_fmt
            ws.cell(row=r, column=8, value=total_utilidades).number_format = money_fmt
            style_total_row(ws, r, len(headers))

    # ── Enviar archivo como descarga (el usuario elige dónde guardarlo) ──
    nombre_tipos = {
        "todos": "Completo",
        "diario": "Libro_Diario",
        "mayor": "Libro_Mayor",
        "balanza": "Balanza_Comprobacion",
        "estado_resultados": "Estado_Resultados",
        "balance_general": "Balance_General",
    }
    label = nombre_tipos.get(tipo_reporte, "Reporte")
    fecha_str = date.today().isoformat()
    filename = f"FG_{label}_{fecha_str}.xlsx"

    # Guardar en memoria (BytesIO) en lugar de disco
    from io import BytesIO

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar como descarga directa al navegador
    from flask import send_file

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5000)

# ═══════════════════════════════════════════════════════════════
#  MÓDULO: GASTOS DE COMERCIALIZACIÓN
# ═══════════════════════════════════════════════════════════════

CUENTAS_COMERCIALIZACION = [
    "6001",
    "6002",
    "6003",
    "6004",
    "6005",
    "6006",
    "6007",
    "6008",
]


def calcular_total_comercializacion(data):
    """Suma todos los movimientos Debe de las cuentas 6xxx."""
    mayor = calcular_mayor(data)
    total = 0.0
    detalle = []
    for cod in sorted(data["cuentas"].keys()):
        if not cod.startswith("6"):
            continue
        info = data["cuentas"][cod]
        if info.get("tipo") != "Gasto":
            continue
        debe = mayor[cod]["debe"] if cod in mayor else 0
        haber = mayor[cod]["haber"] if cod in mayor else 0
        saldo = debe - haber
        if saldo != 0 or debe != 0:
            detalle.append(
                {
                    "codigo": cod,
                    "nombre": info["nombre"],
                    "debe": debe,
                    "haber": haber,
                    "saldo": saldo,
                }
            )
        total += saldo
    return detalle, total


@app.route("/comercializacion")
def comercializacion():
    data = load_data()
    mayor = calcular_mayor(data)

    # Movimientos recientes de cuentas 6xxx
    movimientos = []
    for entry in data["diario"]:
        for mov in entry["movimientos"]:
            if mov["cuenta"].startswith("6"):
                movimientos.append(
                    {
                        "id": entry["id"],
                        "fecha": entry["fecha"],
                        "ref": entry.get("ref", ""),
                        "descripcion": entry["descripcion"],
                        "cuenta": mov["cuenta"],
                        "nombre_cuenta": data["cuentas"]
                        .get(mov["cuenta"], {})
                        .get("nombre", mov["cuenta"]),
                        "tipo": mov["tipo"],
                        "monto": mov["monto"],
                    }
                )
    movimientos.sort(key=lambda x: x["fecha"], reverse=True)

    detalle, total = calcular_total_comercializacion(data)

    # Cuentas 6xxx disponibles (las del catálogo)
    cuentas_6 = {
        cod: info
        for cod, info in data["cuentas"].items()
        if cod.startswith("6") and info.get("tipo") == "Gasto"
    }
    # Todas las cuentas para el formulario de asiento rápido
    todas_cuentas = data["cuentas"]

    return render_template(
        "comercializacion.html",
        movimientos=movimientos,
        detalle=detalle,
        total=total,
        cuentas_6=cuentas_6,
        todas_cuentas=todas_cuentas,
        mayor=mayor,
    )


@app.route("/comercializacion/registrar", methods=["POST"])
def comercializacion_registrar():
    """Registra un gasto de comercialización como asiento en el diario."""
    data = load_data()
    fecha = request.form.get("fecha", date.today().isoformat())
    descripcion = request.form.get("descripcion", "")
    ref = request.form.get("ref", "")
    cuenta_gasto = request.form.get("cuenta_gasto", "")
    cuenta_pago = request.form.get("cuenta_pago", "")
    monto_str = request.form.get("monto", "0")

    try:
        monto = float(monto_str)
    except ValueError:
        return redirect(url_for("comercializacion"))

    if not cuenta_gasto or not cuenta_pago or monto <= 0:
        return redirect(url_for("comercializacion"))

    entry = {
        "id": get_next_id(data, "diario"),
        "fecha": fecha,
        "descripcion": descripcion
        or f"Gasto de comercialización — {data['cuentas'].get(cuenta_gasto, {}).get('nombre', cuenta_gasto)}",
        "ref": ref,
        "movimientos": [
            {"cuenta": cuenta_gasto, "tipo": "Debe", "monto": monto},
            {"cuenta": cuenta_pago, "tipo": "Haber", "monto": monto},
        ],
    }
    data["diario"].append(entry)

    # Reflejo en caja si la cuenta pago es efectivo o banco
    if cuenta_pago in ["1001", "1002"]:
        data["caja_movimientos"].append(
            {
                "id": get_next_id(data, "caja_movimientos"),
                "fecha": fecha,
                "descripcion": entry["descripcion"],
                "tipo": "Haber",
                "monto": monto,
                "cuenta": cuenta_pago,
                "ref_diario": entry["id"],
            }
        )

    save_data(data)
    return redirect(url_for("comercializacion"))


@app.route("/comercializacion/cuenta/nueva", methods=["POST"])
def comercializacion_cuenta_nueva():
    """Agrega una nueva cuenta de comercialización al catálogo."""
    data = load_data()
    codigo = request.form.get("codigo", "").strip()
    nombre = request.form.get("nombre", "").strip()
    if codigo and nombre and codigo not in data["cuentas"]:
        data["cuentas"][codigo] = {"nombre": nombre, "tipo": "Gasto", "saldo": 0}
        save_data(data)
    return redirect(url_for("comercializacion"))


@app.route("/comercializacion/cuenta/eliminar", methods=["POST"])
def comercializacion_cuenta_eliminar():
    """Elimina una cuenta 6xxx del catálogo (solo si no tiene movimientos)."""
    data = load_data()
    codigo = request.form.get("codigo", "").strip()
    mayor = calcular_mayor(data)
    if codigo.startswith("6") and codigo in data["cuentas"]:
        tiene_mov = codigo in mayor and (
            mayor[codigo]["debe"] > 0 or mayor[codigo]["haber"] > 0
        )
        if not tiene_mov:
            del data["cuentas"][codigo]
            save_data(data)
    return redirect(url_for("comercializacion"))
