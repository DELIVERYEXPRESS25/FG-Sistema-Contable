"""
════════════════════════════════════════════════════════════════════════════
 Colectivo FG - Reportes del Kardex PEPS
════════════════════════════════════════════════════════════════════════════

Generación de reportes Excel optimizados para Kardex PEPS.
"""

from datetime import datetime, date
from io import BytesIO

def generar_reporte_kardex_excel(reporte_peps, productos_seleccionados=None):
    """
    Genera reporte Excel del Kardex PEPS.
    
    Args:
        reporte_peps: Dict con datos PEPS de productos
        productos_seleccionados: Lista de códigos de productos (None = todos)
    
    Returns:
        BytesIO: Buffer con el Excel generado
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    
    wb = Workbook()
    
    # Filtrar productos
    if productos_seleccionados:
        productos = {k: v for k, v in reporte_peps.items() if k in productos_seleccionados}
    else:
        productos = reporte_peps
    
    for idx, (codigo, info) in enumerate(productos.items()):
        # Crear hoja
        if idx == 0:
            ws = wb.active
            ws.title = _sanitize_sheet_name(codigo)
        else:
            ws = wb.create_sheet(_sanitize_sheet_name(codigo))
        
        ws.sheet_properties.tabColor = "4F8CFF"
        
        # ── ENCABEZADO ──
        ws.append(["Colectivo FG - Joyería y Muchas Más"])
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(size=16, bold=True, color="1A1D2E")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.append([f"KARDEX PEPS - {info['nombre']}"])
        ws.merge_cells('A2:H2')
        ws['A2'].font = Font(size=14, bold=True, color="4F8CFF")
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.append([f"Código: {codigo} | Generado: {date.today().isoformat()}"])
        ws.merge_cells('A3:H3')
        ws['A3'].font = Font(size=10, color="7a7f99")
        ws['A3'].alignment = Alignment(horizontal='center')
        
        ws.append([])  # Línea vacía
        
        # ── RESUMEN ──
        ws.append(["RESUMEN DEL PRODUCTO"])
        ws['A5'].font = Font(size=12, bold=True)
        ws.merge_cells('A5:B5')
        
        ws.append(["Stock Actual:", info['stock']])
        ws.append(["Costo Promedio:", f"C$ {info['costo_promedio']:.2f}"])
        ws.append(["Valor Inventario:", f"C$ {info['valor_inventario']:.2f}"])
        
        ws.append([])  # Línea vacía
        
        # ── LOTES DISPONIBLES ──
        if info['lotes']:
            ws.append(["LOTES DISPONIBLES (Método PEPS)"])
            ws['A10'].font = Font(size=12, bold=True, color="2ECC71")
            ws.merge_cells('A10:E10')
            
            headers_lotes = ["Fecha Entrada", "Cant. Original", "Cant. Disponible", "Costo Unit.", "Valor Total"]
            ws.append(headers_lotes)
            
            for cell in ws[11]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            for lote in info['lotes']:
                ws.append([
                    lote['fecha'],
                    lote['cantidad_original'],
                    lote['cantidad_disponible'],
                    lote['costo_unitario'],
                    lote['valor_total']
                ])
                
                # Formato moneda
                last_row = ws.max_row
                ws.cell(last_row, 4).number_format = '#,##0.00'
                ws.cell(last_row, 5).number_format = '#,##0.00'
        else:
            ws.append(["LOTES DISPONIBLES: Ninguno"])
            ws['A10'].font = Font(size=12, color="E74C3C")
        
        ws.append([])  # Línea vacía
        
        # ── MOVIMIENTOS ──
        row_start = ws.max_row + 1
        ws.append(["HISTORIAL DE MOVIMIENTOS"])
        ws[f'A{row_start}'].font = Font(size=12, bold=True, color="E74C3C")
        ws.merge_cells(f'A{row_start}:H{row_start}')
        
        headers_mov = ["Fecha", "Tipo", "Cantidad", "Costo Unit.", "Total", "Saldo", "Descripción"]
        ws.append(headers_mov)
        
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        if info['movimientos']:
            for mov in reversed(info['movimientos']):
                tipo_texto = f"{'📥 ' if mov['tipo'] == 'entrada' else '📤 ' if mov['tipo'] == 'salida' else '🔄 '}{mov['tipo'].title()}"
                ws.append([
                    mov['fecha'],
                    tipo_texto,
                    mov['cantidad'],
                    mov.get('costo', 0),
                    mov.get('total', 0),
                    mov['saldo'],
                    mov.get('descripcion', '')
                ])
                
                # Formato
                last_row = ws.max_row
                ws.cell(last_row, 4).number_format = '#,##0.00'
                ws.cell(last_row, 5).number_format = '#,##0.00'
        else:
            ws.append(["Sin movimientos registrados"])
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 40
    
    # Guardar en buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def _sanitize_sheet_name(name):
    """Limpia nombre de hoja Excel (max 31 caracteres, sin caracteres especiales)."""
    # Remover caracteres inválidos
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Truncar a 31 caracteres
    return name[:31]
