"""
════════════════════════════════════════════════════════════════════════════
 Generador de Excel Kardex PEPS - Formato Profesional
════════════════════════════════════════════════════════════════════════════

Genera Excel con formato idéntico al estándar contable mostrado en la imagen.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO

def generar_excel_kardex_peps(kardex_obj, empresa_nombre="EMPRESA COMERCIAL \"ABC\""):
    """
    Genera Excel del Kardex PEPS con formato profesional.
    
    Args:
        kardex_obj: Objeto KardexPEPS con movimientos
        empresa_nombre: Nombre de la empresa
    
    Returns:
        BytesIO con el Excel generado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = kardex_obj.producto_codigo[:31]
    
    # ═══════════════════════════════════════════════════════════════
    #  ESTILOS
    # ═══════════════════════════════════════════════════════════════
    
    # Header azul (MÉTODO PEPS)
    header_azul_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_azul_font = Font(name="Arial", size=24, bold=True, color="FFFFFF")
    
    # Header verde (EMPRESA)
    header_verde_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    header_verde_font = Font(name="Arial", size=12, bold=True)
    
    # Subheader amarillo (columnas)
    subheader_amarillo_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    subheader_font = Font(name="Arial", size=10, bold=True)
    
    # Subheader gris (Artículo, Fecha, etc)
    subheader_gris_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ═══════════════════════════════════════════════════════════════
    #  CONSTRUCCIÓN DEL EXCEL
    # ═══════════════════════════════════════════════════════════════
    
    # Fila 1: MÉTODO PEPS (azul)
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = "MÉTODO PEPS"
    cell.font = header_azul_font
    cell.fill = header_azul_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # Fila 2: Vacía
    ws.row_dimensions[2].height = 10
    
    # Fila 3: EMPRESA COMERCIAL (verde)
    ws.merge_cells('A3:K3')
    cell = ws['A3']
    cell.value = empresa_nombre
    cell.font = header_verde_font
    cell.fill = header_verde_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    
    # Fila 4: TARJETA DE CONTROL DE EXISTENCIAS - KARDEX
    ws.merge_cells('A4:K4')
    cell = ws['A4']
    cell.value = "TARJETA DE CONTROL DE EXISTENCIAS - KARDEX"
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    
    # Fila 5: Headers principales (amarillo)
    headers_principales = [
        ('A5', 'Artículo'),
        ('B5', 'Pantalones'),
        ('C5:E5', 'INGRESOS - COMPRAS'),
        ('F5:H5', 'EGRESOS - VENTAS'),
        ('I5:K5', 'SALDO'),
        ('L5', 'PEPS')
    ]
    
    for cell_range, text in headers_principales:
        if ':' in cell_range:
            ws.merge_cells(cell_range)
            cell = ws[cell_range.split(':')[0]]
        else:
            cell = ws[cell_range]
        
        cell.value = text
        cell.font = subheader_font
        cell.fill = subheader_amarillo_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Fila 6: Subheaders (gris)
    row_6_headers = [
        ('A6', 'Fecha'),
        ('B6', 'Concepto'),
        ('C6', 'Cantidad'),
        ('D6', 'Costo\nUnitario'),
        ('E6', 'Total'),
        ('F6', 'Cantidad'),
        ('G6', 'Costo\nUnitario'),
        ('H6', 'Total'),
        ('I6', 'Cantidad'),
        ('J6', 'Costo\nUnitario'),
        ('K6', 'Total'),
        ('L6', '')
    ]
    
    for cell_ref, text in row_6_headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.fill = subheader_gris_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[6].height = 30
    
    # ═══════════════════════════════════════════════════════════════
    #  DATOS - MOVIMIENTOS
    # ═══════════════════════════════════════════════════════════════
    
    current_row = 7
    
    # Resaltar amarillo para saldos destacados
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for mov in kardex_obj.movimientos:
        # Fecha
        ws.cell(current_row, 1).value = mov['fecha']
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center')
        ws.cell(current_row, 1).border = thin_border
        
        # Concepto
        ws.cell(current_row, 2).value = mov['concepto']
        ws.cell(current_row, 2).border = thin_border
        
        # INGRESOS
        ws.cell(current_row, 3).value = mov['ingresos']['cantidad'] if mov['ingresos']['cantidad'] != '' else ''
        ws.cell(current_row, 4).value = mov['ingresos']['costo_unitario'] if mov['ingresos']['costo_unitario'] != '' else ''
        ws.cell(current_row, 5).value = mov['ingresos']['total'] if mov['ingresos']['total'] != '' else ''
        
        if mov['ingresos']['costo_unitario'] != '':
            ws.cell(current_row, 4).number_format = '#,##0.00'
        if mov['ingresos']['total'] != '':
            ws.cell(current_row, 5).number_format = '#,##0.00'
        
        # EGRESOS  
        ws.cell(current_row, 6).value = mov['egresos']['cantidad'] if mov['egresos']['cantidad'] != '' else ''
        ws.cell(current_row, 7).value = mov['egresos']['costo_unitario'] if mov['egresos']['costo_unitario'] != '' else ''
        ws.cell(current_row, 8).value = mov['egresos']['total'] if mov['egresos']['total'] != '' else ''
        
        if mov['egresos']['costo_unitario'] != '':
            ws.cell(current_row, 7).number_format = '#,##0.00'
        if mov['egresos']['total'] != '':
            ws.cell(current_row, 8).number_format = '#,##0.00'
        
        # SALDO
        ws.cell(current_row, 9).value = mov['saldo']['cantidad']
        ws.cell(current_row, 10).value = mov['saldo']['costo_unitario']
        ws.cell(current_row, 11).value = mov['saldo']['total']
        
        ws.cell(current_row, 10).number_format = '#,##0.00'
        ws.cell(current_row, 11).number_format = '#,##0.00'
        
        # Resaltar ciertos saldos en amarillo (ejemplo: después de compras)
        if mov['tipo'] in ['compra', 'saldo_inicial']:
            ws.cell(current_row, 9).fill = highlight_fill
            ws.cell(current_row, 10).fill = highlight_fill
            ws.cell(current_row, 11).fill = highlight_fill
        
        # PEPS
        ws.cell(current_row, 12).value = mov.get('peps_info', '')
        ws.cell(current_row, 12).font = Font(name="Arial", size=8)
        
        # Bordes
        for col in range(3, 13):
            ws.cell(current_row, col).border = thin_border
            ws.cell(current_row, col).alignment = Alignment(horizontal='right' if col >= 3 else 'left')
        
        current_row += 1
    
    # ═══════════════════════════════════════════════════════════════
    #  AJUSTAR ANCHOS DE COLUMNA
    # ═══════════════════════════════════════════════════════════════
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 30
    
    # ═══════════════════════════════════════════════════════════════
    #  GUARDAR EN BUFFER
    # ═══════════════════════════════════════════════════════════════
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
